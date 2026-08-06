"""
ITR-FA-GENERATOR - Schedule FA Generator for ITR2
Copyright (C) 2025 Satvir Singh

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program. If not, see <https://www.gnu.org/licenses/>.

---

Schedule FA Generator - Web Scraping Version
Uses Selenium to scrape Yahoo Finance and download SBI rates from GitHub

Advantages:
- Fully automated (no manual downloads)
- Works on corporate networks (uses browser)
- Gets accurate data from source websites
- No API dependencies

Requirements:
- Chrome browser installed
- Internet connection
"""

import json
import os
import sys
import argparse
import pandas as pd
from datetime import datetime
import time
import warnings
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import requests

warnings.filterwarnings('ignore')

def calculate_stcg_rates_for_income(income_bracket):
    """
    Calculate STCG rates for BOTH New and Old tax regimes based on income bracket.

    Args:
        income_bracket (str): Income bracket choice (1-11)

    Returns:
        tuple: (new_regime_rate, old_regime_rate, new_display, old_display)
    """
    bracket = income_bracket.strip()

    # Income bracket to New Regime rate mapping
    new_regime_rates = {
        '1': (0.0, "0% (Nil)"),
        '2': (0.052, "5.2% (5% + 4% cess)"),
        '3': (0.104, "10.4% (10% + 4% cess)"),
        '4': (0.156, "15.6% (15% + 4% cess)"),
        '5': (0.208, "20.8% (20% + 4% cess)"),
        '6': (0.260, "26.0% (25% + 4% cess)"),
        '7': (0.312, "31.2% (30% + 4% cess)"),
        '8': (0.3432, "34.32% (30% + 10% surcharge + 4% cess)"),
        '9': (0.3588, "35.88% (30% + 15% surcharge + 4% cess)"),
        '10': (0.390, "39.0% (30% + 25% surcharge + 4% cess)"),
        '11': (0.390, "39.0% (30% + 25% surcharge + 4% cess)"),
    }

    # Income bracket to Old Regime rate mapping
    # Old regime has different base slabs up to Rs 10L
    old_regime_rates = {
        '1': (0.0, "0% (Nil)"),  # Up to 4L (old has 2.5L, but effectively 0 for both)
        '2': (0.052, "5.2% (5% + 4% cess)"),  # 4-8L (old: 2.5-5L at 5%)
        '3': (0.208, "20.8% (20% + 4% cess)"),  # 8-12L (old: 5-10L at 20%)
        '4': (0.312, "31.2% (30% + 4% cess)"),  # 12-16L (old: >10L at 30%)
        '5': (0.312, "31.2% (30% + 4% cess)"),  # 16-20L
        '6': (0.312, "31.2% (30% + 4% cess)"),  # 20-24L
        '7': (0.312, "31.2% (30% + 4% cess)"),  # 24-50L
        '8': (0.3432, "34.32% (30% + 10% surcharge + 4% cess)"),  # 50L-1Cr
        '9': (0.3588, "35.88% (30% + 15% surcharge + 4% cess)"),  # 1Cr-2Cr
        '10': (0.390, "39.0% (30% + 25% surcharge + 4% cess)"),  # 2Cr-5Cr
        '11': (0.42744, "42.744% (30% + 37% surcharge + 4% cess)"),  # Above 5Cr
    }

    if bracket not in new_regime_rates:
        raise ValueError(f"Invalid income bracket: {bracket}")

    new_rate, new_display = new_regime_rates[bracket]
    old_rate, old_display = old_regime_rates[bracket]

    return new_rate, old_rate, new_display, old_display

warnings.filterwarnings('ignore')

def save_config(config, config_file="config.json"):
    """Save configuration back to config.json file."""
    try:
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"[!] Error saving config: {e}")
        return False

def load_config(config_file="config.json"):
    """Load configuration from config.json file."""
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
        print(f"[OK] Loaded configuration from {config_file}")
        # Handle backward compatibility
        if "table_a2_custodial_accounts" in config:
            # Old multi-account format - take first account
            config["custodial_account"] = config["table_a2_custodial_accounts"][0]
        elif "table_a2_custodial_account" in config:
            # Very old single account format
            config["custodial_account"] = config["table_a2_custodial_account"]
        return config
    except FileNotFoundError:
        print(f"[!] ERROR: {config_file} not found!")
        print(f"[!] Please create config.json with your account details.")
        print(f"[!] See config.example.json for reference.")
        raise FileNotFoundError(f"Config file '{config_file}' is required. Copy config.example.json to config.json and edit your details.")
    except json.JSONDecodeError as e:
        print(f"[!] Error parsing {config_file}: {e}")
        print(f"[!] Please check your JSON syntax and try again.")
        raise

class ScheduleFAApp:
    @staticmethod
    def clean_text_for_itr(text):
        """Remove commas (ITR portal doesn't accept commas). Keep periods and asterisks."""
        if not text or not isinstance(text, str):
            return text
        # Remove only commas - periods and asterisks are allowed
        text = text.replace(',', '')
        # Remove other problematic characters if needed
        return text.strip()

    def __init__(self, calendar_year=None, stcg_rate_new=0.312, stcg_rate_old=0.312):
        """
        :param calendar_year: e.g. 2024, 2025, 2026. If None, defaults to the previous calendar year.
        :param stcg_rate_new: STCG tax rate for New Tax Regime (default 31.2% for 30% bracket)
        :param stcg_rate_old: STCG tax rate for Old Tax Regime (default 31.2% for 30% bracket)
        """
        current_year = datetime.now().year
        self.calendar_year = calendar_year if calendar_year else (current_year - 1)
        self.stcg_rate_new = stcg_rate_new  # Store New Regime STCG rate
        self.stcg_rate_old = stcg_rate_old  # Store Old Regime STCG rate
        # For backward compatibility
        self.stcg_tax_rate = stcg_rate_new

        # Schedule FA uses CALENDAR YEAR (Jan 1 - Dec 31)
        self.start_date = f"{self.calendar_year}-01-01"
        self.end_date = f"{self.calendar_year}-12-31"

        # Capital Gains (Schedule CG) uses extended range (Jan 1 - Mar 31 next year)
        # Per ITRFA.in: "If also filing Schedule CG, OS/FSI, or Form 67, set to Jan 1 2025 - Mar 31 2026 (15 months)"
        # This captures sales from Jan-Mar of next calendar year that fall in same FY
        self.cg_start_date = f"{self.calendar_year}-01-01"
        self.cg_end_date = f"{self.calendar_year + 1}-03-31"

        # Dynamic Tax Schema Years
        self.indian_fy = f"{self.calendar_year}-{str(self.calendar_year + 1)[-2:]}"
        self.assessment_year = f"{self.calendar_year + 1}-{str(self.calendar_year + 2)[-2:]}"

        print(f"\n=======================================================")
        print(f"  Schedule FA - Web Scraping Mode")
        print(f"  Target Year: {self.calendar_year}")
        print(f"  Applicable FY: {self.indian_fy} | AY: {self.assessment_year}")
        print(f"  Schedule FA: {self.start_date} to {self.end_date}")
        print(f"  Capital Gains: {self.cg_start_date} to {self.cg_end_date}")
        print(f"=======================================================\n")

        # Datasets & Caches
        self.df_sbi = self._fetch_sbi_rates_web()  # Initial download (will reload with extra dates later)
        self.company_cache = {}
        self._extra_ttbr_dates_loaded = False  # Track if we've loaded pre-FY dates

    def _fetch_sbi_rates_web(self, extra_dates=None):
        """Downloads and filters SBI forex rates using our own fetcher (SBI PDF first, GitHub fallback).

        Args:
            extra_dates: List of specific dates before FY to include (e.g., ['2024-11-08', '2024-09-15'])
        """
        # Read SBI forex rates from local CSV (updated daily by GitHub Action at 9 PM IST)
        df = None

        # Try reading local CSV first (primary source - updated by GitHub Action)
        if df is None:
            local_csv = os.path.join('data', 'SBI_FOREX_CARD_RATES_USD.csv')
            if os.path.exists(local_csv):
                try:
                    print(f"[*] Reading local SBI forex data from {local_csv}...")
                    df = pd.read_csv(local_csv)

                    if 'DATE' in df.columns and 'TT BUY' in df.columns:
                        df['Date'] = pd.to_datetime(df['DATE']).dt.strftime('%Y-%m-%d')
                        df['TTBR'] = pd.to_numeric(df['TT BUY'], errors='coerce')
                        df = df[['Date', 'TTBR']].copy()
                        print(f"[OK] Loaded {len(df)} records from local CSV")
                    else:
                        raise ValueError(f"Local CSV has unexpected format. Columns: {list(df.columns)}")
                except Exception as e:
                    print(f"[!] Failed to read local CSV: {e}")
                    df = None

        # If still no data, try legacy GitHub download
        if df is None:
            print("[*] Downloading SBI TTBR rates from GitHub (legacy method)...")

            url = "https://raw.githubusercontent.com/sahilgupta/sbi-fx-ratekeeper/main/csv_files/SBI_REFERENCE_RATES_USD.csv"

            try:
                # Download the CSV file
                response = requests.get(url, timeout=30, verify=False)
                response.raise_for_status()

                # Parse CSV
                from io import StringIO
                df = pd.read_csv(StringIO(response.text))

                print(f"[*] CSV columns found: {list(df.columns)[:5]}")

                # GitHub CSV uses: DATE, TT BUY (not Date, TTBR)
                # Normalize column names
                if 'DATE' in df.columns:
                    df['Date'] = pd.to_datetime(df['DATE']).dt.strftime('%Y-%m-%d')
                    df['TTBR'] = pd.to_numeric(df['TT BUY'], errors='coerce')
                    df = df[['Date', 'TTBR']].copy()
                elif 'Date' not in df.columns or 'TTBR' not in df.columns:
                    raise ValueError(f"CSV must have 'DATE'/'Date' and 'TT BUY'/'TTBR' columns. Found: {list(df.columns)}")

                df = df.sort_values('Date').dropna(subset=['TTBR'])
            except Exception as e:
                print(f"[ERROR] Failed to download from GitHub: {e}")
                raise

        # At this point, df should have columns: Date, TTBR (from either our fetcher, local CSV, or GitHub legacy)
        # Ensure data is clean
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
        df['TTBR'] = pd.to_numeric(df['TTBR'], errors='coerce')
        df = df.sort_values('Date').dropna(subset=['TTBR'])

        # Filter for target year only (for peak/closing calculations)
        df_year = df[(df['Date'] >= self.start_date) & (df['Date'] <= self.end_date)]

        # Add historical data for pre-calendar-year dates (if requested)
        if extra_dates:
            # Return ALL historical data from CSV (not just requested dates)
            # This allows automatic backward search to find preceding trading days
            # Filter: Include target year + all dates before target year
            df_with_history = df[df['Date'] <= self.end_date]

            # Verify requested dates exist (for logging)
            df_extra = df[df['Date'].isin(extra_dates)]
            found_dates = df_extra['Date'].tolist()
            missing_dates = [d for d in extra_dates if d not in found_dates]

            if missing_dates:
                print(f"[!] WARNING: Could not find exact TTBR for dates (will use backward search): {', '.join(sorted(missing_dates))}")

            print(f"[OK] Downloaded {len(df_year)} SBI TTBR records for {self.calendar_year}")
            print(f"[OK] Plus {len(df_with_history) - len(df_year)} historical records for backward search")
            print(f"[OK] TTBR range: {df_with_history['TTBR'].min():.2f} to {df_with_history['TTBR'].max():.2f}")
            return df_with_history[['Date', 'TTBR']]
        else:
            if df_year.empty:
                print(f"[!] WARNING: No SBI TTBR data for {self.calendar_year}")
                print(f"[!] Available date range: {df['Date'].min()} to {df['Date'].max()}")
                print("[!] Using fallback interpolated rates")
                raise ValueError(f"No data for year {self.calendar_year}")

            print(f"[OK] Downloaded {len(df_year)} SBI TTBR records for {self.calendar_year}")
            print(f"[OK] TTBR range: {df_year['TTBR'].min():.2f} to {df_year['TTBR'].max():.2f}")
            return df_year[['Date', 'TTBR']]

    def _calculate_schedule_os_fsi(self, df_dividends, df_capital_gains):
        """
        Calculate Schedule OS (Other Sources) and Schedule FSI (Foreign Source Income).

        Schedule OS uses Rule 115(1)(e): Last day of month BEFORE dividend month
        Schedule FSI aggregates dividend + capital gains income

        Returns: (df_schedule_os, df_schedule_fsi)
        """
        import math

        # CRITICAL: Schedule OS uses FINANCIAL YEAR (Apr-Mar), NOT calendar year!
        fy_start = f"{self.calendar_year}-04-01"
        fy_end = f"{self.calendar_year + 1}-03-31"

        # Filter dividends to Financial Year
        if not df_dividends.empty:
            df_div_fy = df_dividends.copy()
            df_div_fy['Date'] = pd.to_datetime(df_div_fy['Date'])
            df_div_fy = df_div_fy[(df_div_fy['Date'] >= fy_start) & (df_div_fy['Date'] <= fy_end)].copy()
        else:
            df_div_fy = pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

        # Recalculate dividends using Rule 115(1)(e) for Schedule OS
        div_os_data = []
        for _, row in df_div_fy.iterrows():
            div_date = pd.to_datetime(row['Date'])
            amount_usd = row['Amount (USD)']

            # Rule 115(1)(e): Last day of month BEFORE dividend month
            if div_date.month == 1:
                specified_year = div_date.year - 1
                specified_month = 12
            else:
                specified_year = div_date.year
                specified_month = div_date.month - 1

            import calendar
            last_day = calendar.monthrange(specified_year, specified_month)[1]
            specified_date_str = f"{specified_year}-{specified_month:02d}-{last_day:02d}"

            # Get TTBR for specified date
            ttbr_row = self.df_sbi[self.df_sbi['Date'] == specified_date_str]
            if not ttbr_row.empty:
                ttbr_os = ttbr_row['TTBR'].values[0]
            else:
                # Use backward search
                sbi_before = self.df_sbi[self.df_sbi['Date'] < specified_date_str]
                if not sbi_before.empty:
                    closest_date = sbi_before['Date'].max()
                    ttbr_os = sbi_before[sbi_before['Date'] == closest_date]['TTBR'].values[0]
                else:
                    ttbr_os = row['TTBR']  # Fallback to Schedule FA rate

            amount_inr_os = math.ceil(amount_usd * ttbr_os)

            div_os_data.append({
                'Symbol': row['Symbol'],
                'Date': div_date.strftime('%Y-%m-%d'),
                'Amount (USD)': amount_usd,
                'Specified Date (Rule 115(1)(e))': specified_date_str,
                'TTBR': round(ttbr_os, 2),
                'Amount (INR)': amount_inr_os
            })

        df_div_os = pd.DataFrame(div_os_data)

        # Calculate quarterly breakup (Section 234C)
        # Quarter 1: Apr 1 - Jun 15
        # Quarter 2: Jun 16 - Sep 15
        # Quarter 3: Sep 16 - Dec 15
        # Quarter 4: Dec 16 - Mar 15
        # Quarter 5: Mar 16 - Mar 31

        quarterly_breakup = {
            'Q1 (Apr 1 - Jun 15)': 0,
            'Q2 (Jun 16 - Sep 15)': 0,
            'Q3 (Sep 16 - Dec 15)': 0,
            'Q4 (Dec 16 - Mar 15)': 0,
            'Q5 (Mar 16 - Mar 31)': 0
        }

        for _, row in df_div_os.iterrows():
            div_date = pd.to_datetime(row['Date'])
            amount = row['Amount (INR)']
            month = div_date.month
            day = div_date.day

            # Determine quarter based on payment date
            if month <= 6 and (month < 6 or day <= 15):
                quarterly_breakup['Q1 (Apr 1 - Jun 15)'] += amount
            elif month <= 9 and (month < 9 or day <= 15):
                quarterly_breakup['Q2 (Jun 16 - Sep 15)'] += amount
            elif month <= 12 and (month < 12 or day <= 15):
                quarterly_breakup['Q3 (Sep 16 - Dec 15)'] += amount
            elif month <= 3 or (month == 3 and day <= 15):
                quarterly_breakup['Q4 (Dec 16 - Mar 15)'] += amount
            else:
                quarterly_breakup['Q5 (Mar 16 - Mar 31)'] += amount

        # Build Schedule OS DataFrame
        total_div_usd = df_div_os['Amount (USD)'].sum() if not df_div_os.empty else 0
        total_div_inr = df_div_os['Amount (INR)'].sum() if not df_div_os.empty else 0

        os_data = {
            'Indian Financial Year': [
                'Assessment Year',
                'Total Dividend Income (USD)',
                'Total Dividend Income (INR Rs.)',
                '',
                'Quarter (Section 234C)',
                '',
                'WARNINGS / NOTES'
            ],
            self.indian_fy: [
                self.assessment_year,
                round(total_div_usd, 2),
                int(total_div_inr),
                '',
                'Dividend Income (INR Rs.)',
                '',
                ''
            ]
        }

        # Add quarterly rows
        for quarter, amount in quarterly_breakup.items():
            os_data['Indian Financial Year'].append(quarter)
            os_data[self.indian_fy].append(int(amount))

        # Add warning if no dividends
        if total_div_inr == 0:
            os_data['Indian Financial Year'].append('')
            os_data[self.indian_fy].append(f"• No dividend activity found in Indian FY {self.indian_fy} (Apr-Mar)")

        df_schedule_os = pd.DataFrame(os_data)

        # Build Schedule FSI DataFrame
        # Aggregate capital gains from Financial Year sales
        total_cg_inr = 0
        if not df_capital_gains.empty:
            # Filter capital gains to Financial Year
            df_cg_fy = df_capital_gains.copy()
            df_cg_fy['Sale Date'] = pd.to_datetime(df_cg_fy['Sale Date'])
            df_cg_fy = df_cg_fy[(df_cg_fy['Sale Date'] >= fy_start) & (df_cg_fy['Sale Date'] <= fy_end)]
            total_cg_inr = int(df_cg_fy['Capital Gain (INR)'].sum()) if not df_cg_fy.empty else 0

        total_foreign_income = int(total_div_inr + total_cg_inr)

        # TODO: Extract NRA withholding from Transaction History
        total_tax_paid_usd = 0
        total_tax_paid_inr = 0

        # Build FSI summary section (only 2 columns)
        fsi_data = {
            'Indian Financial Year': [
                'Assessment Year',
                'Dividend Income (Foreign)',
                'Capital Gains Income (Foreign, per Schedule CG)',
                'Total Foreign Source Income',
                'Total Tax Paid Outside India',
                'Total Tax Relief Available (Schedule TR)',
                '',
                'WARNINGS / NOTES'
            ],
            self.indian_fy: [
                self.assessment_year,
                int(total_div_inr),
                total_cg_inr,
                total_foreign_income,
                total_tax_paid_inr,
                0,  # Tax relief calculated later
                '',
                ''
            ]
        }

        # Add warning notes if applicable
        if total_div_inr == 0 and total_cg_inr == 0:
            fsi_data['Indian Financial Year'].append('')
            fsi_data[self.indian_fy].append(f"• No dividend, NRA withholding, or capital-gains activity found in Indian FY {self.indian_fy} (Apr-Mar)")

        df_schedule_fsi = pd.DataFrame(fsi_data)

        return df_schedule_os, df_schedule_fsi, df_div_os

    def _parse_dividend_data(self, transaction_history_path=None):
        """
        Parse dividend transactions from E*TRADE Transaction History CSV.

        Returns: DataFrame with columns: Symbol, Date, Amount (USD), TTBR, Amount (INR)
        Returns empty DataFrame if file not found (dividends are optional).
        """
        if not transaction_history_path or not os.path.exists(transaction_history_path):
            print(f"[i] Transaction History file not found - skipping dividend processing")
            return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

        try:
            # Read Transaction History CSV
            df_trans = pd.read_csv(transaction_history_path)
            print(f"[*] Reading dividend data from {transaction_history_path}")

            # Filter for dividend transactions
            # E*TRADE typically uses "Dividend" or "Cash Dividend" in TransactionType or Description
            dividend_mask = (
                (df_trans.get('TransactionType', pd.Series()).str.contains('Dividend', case=False, na=False)) |
                (df_trans.get('Description', pd.Series()).str.contains('Dividend', case=False, na=False)) |
                (df_trans.get('Type', pd.Series()).str.contains('Dividend', case=False, na=False))
            )

            df_dividends = df_trans[dividend_mask].copy()

            if df_dividends.empty:
                print(f"[i] No dividend transactions found in Transaction History")
                return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

            # Extract relevant columns (handle various E*TRADE CSV formats)
            # Common column names: TransactionDate, Date, SettlementDate, PostedDate
            date_col = None
            for col in ['TransactionDate', 'Date', 'SettlementDate', 'PostedDate']:
                if col in df_dividends.columns:
                    date_col = col
                    break

            if not date_col:
                print(f"[!] Warning: Could not find date column in Transaction History")
                return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

            # Amount column (usually 'Amount', 'NetAmount', or 'Quantity')
            amount_col = None
            for col in ['Amount', 'NetAmount', 'Quantity', 'Credit']:
                if col in df_dividends.columns:
                    amount_col = col
                    break

            if not amount_col:
                print(f"[!] Warning: Could not find amount column in Transaction History")
                return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

            # Symbol column
            symbol_col = None
            for col in ['Symbol', 'SecuritySymbol', 'Ticker']:
                if col in df_dividends.columns:
                    symbol_col = col
                    break

            if not symbol_col:
                print(f"[!] Warning: Could not find symbol column in Transaction History")
                return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

            # Parse dates
            df_dividends['Date'] = pd.to_datetime(df_dividends[date_col])

            # Filter to calendar year only
            df_dividends = df_dividends[
                (df_dividends['Date'] >= self.start_date) &
                (df_dividends['Date'] <= self.end_date)
            ].copy()

            if df_dividends.empty:
                print(f"[i] No dividends found in calendar year {self.calendar_year}")
                return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

            # Extract amounts (handle negative values - dividend credits are positive)
            df_dividends['Amount (USD)'] = df_dividends[amount_col].abs()
            df_dividends['Symbol'] = df_dividends[symbol_col]

            # Get TTBR for each dividend date (exact credit date per CBDT Schedule FA instructions)
            dividend_list = []
            for _, row in df_dividends.iterrows():
                div_date = row['Date'].strftime('%Y-%m-%d')
                symbol = row['Symbol']
                amount_usd = row['Amount (USD)']

                # Get TTBR for exact dividend credit date
                ttbr_row = self.df_sbi[self.df_sbi['Date'] == div_date]

                if not ttbr_row.empty:
                    ttbr = ttbr_row['TTBR'].values[0]
                else:
                    # Dividend date not found (weekend/holiday) - use nearest preceding working day
                    sbi_before = self.df_sbi[self.df_sbi['Date'] < div_date]
                    if not sbi_before.empty:
                        closest_date = sbi_before['Date'].max()
                        ttbr = sbi_before[sbi_before['Date'] == closest_date]['TTBR'].values[0]
                        print(f"[i] Dividend {div_date} is weekend/holiday, using previous trading day {closest_date} TTBR: {ttbr:.2f}")
                    else:
                        print(f"[!] Warning: No TTBR found for dividend date {div_date}, skipping")
                        continue

                # Convert to INR using exact date TTBR
                import math
                amount_inr = math.ceil(amount_usd * ttbr)

                dividend_list.append({
                    'Symbol': symbol,
                    'Date': div_date,
                    'Amount (USD)': round(amount_usd, 2),
                    'TTBR': round(ttbr, 2),
                    'Amount (INR)': amount_inr
                })

            df_result = pd.DataFrame(dividend_list)

            if not df_result.empty:
                total_usd = df_result['Amount (USD)'].sum()
                total_inr = df_result['Amount (INR)'].sum()
                print(f"[OK] Found {len(df_result)} dividend transactions in {self.calendar_year}")
                print(f"[OK] Total dividends: ${total_usd:.2f} USD = Rs.{total_inr:,} INR")

            return df_result

        except Exception as e:
            print(f"[!] Error parsing dividend data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame(columns=['Symbol', 'Date', 'Amount (USD)', 'TTBR', 'Amount (INR)'])

    def _detect_country(self, address):
        """Detect country from address and return ITR-compliant country name and code."""
        address_upper = address.upper()

        # ITR Schedule FA Country Codes (common ones)
        country_mapping = {
            "UNITED STATES": {"name": "UNITED STATES OF AMERICA", "code": "2"},
            "USA": {"name": "UNITED STATES OF AMERICA", "code": "2"},
            "U.S.A": {"name": "UNITED STATES OF AMERICA", "code": "2"},
            "CANADA": {"name": "CANADA", "code": "3"},
            "UNITED KINGDOM": {"name": "UNITED KINGDOM", "code": "1"},
            "UK": {"name": "UNITED KINGDOM", "code": "1"},
            "GERMANY": {"name": "GERMANY", "code": "4"},
            "FRANCE": {"name": "FRANCE", "code": "5"},
            "JAPAN": {"name": "JAPAN", "code": "6"},
            "AUSTRALIA": {"name": "AUSTRALIA", "code": "7"},
            "SWITZERLAND": {"name": "SWITZERLAND", "code": "8"},
            "NETHERLANDS": {"name": "NETHERLANDS", "code": "9"},
            "SINGAPORE": {"name": "SINGAPORE", "code": "10"},
            "HONG KONG": {"name": "HONG KONG", "code": "11"},
            "CHINA": {"name": "CHINA", "code": "12"},
            "SOUTH KOREA": {"name": "SOUTH KOREA", "code": "13"},
            "KOREA": {"name": "SOUTH KOREA", "code": "13"},
            "TAIWAN": {"name": "TAIWAN", "code": "14"},
            "INDIA": {"name": "INDIA", "code": "15"},
            "BRAZIL": {"name": "BRAZIL", "code": "16"},
            "MEXICO": {"name": "MEXICO", "code": "17"},
            "ISRAEL": {"name": "ISRAEL", "code": "18"},
            "IRELAND": {"name": "IRELAND", "code": "19"},
            "SPAIN": {"name": "SPAIN", "code": "20"},
            "ITALY": {"name": "ITALY", "code": "21"},
        }

        # Try to detect country from address
        for country_key, country_data in country_mapping.items():
            if country_key in address_upper:
                return country_data["name"], country_data["code"]

        # Default to USA if not detected
        print(f"     [!] Could not detect country from address, defaulting to USA")
        return "UNITED STATES OF AMERICA", "2"

    def _scrape_company_profile(self, symbol):
        """Scrapes company profile information from Yahoo Finance."""
        print(f"[*] Fetching company profile for {symbol}...")

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        driver = None
        try:
            driver = webdriver.Chrome(options=chrome_options)

            # Go to Yahoo Finance profile page
            url = f"https://finance.yahoo.com/quote/{symbol}/profile"
            driver.get(url)
            time.sleep(3)  # Wait for page to load

            # Extract company info from asset-profile section
            company_name = f"{symbol} Inc."  # Default
            full_address = "United States"  # Default
            zip_code = "00000"  # Default
            country_name = "UNITED STATES OF AMERICA"  # Default
            country_code = "2"  # Default (USA)

            try:
                # Find the asset-profile section using data-testid
                profile_section = driver.find_element(By.CSS_SELECTOR, "[data-testid='asset-profile']")

                # Extract company name (h3 within asset-profile)
                try:
                    company_name = profile_section.find_element(By.CSS_SELECTOR, "h3").text.strip()
                except:
                    print(f"[!] Could not extract company name from h3")

                # Extract address from nested classes: .company-details > .company-info > .address
                # The .address class has 3 divs: [street, city+state+zip, country]
                # We only need first 2 divs (street + city/state/zip)
                try:
                    # All are class names, not tags
                    address_element = profile_section.find_element(By.CSS_SELECTOR, ".company-details .company-info .address")

                    # Get all div children (3 divs: street, city+state+zip, country)
                    address_divs = address_element.find_elements(By.TAG_NAME, "div")

                    # Extract country from third div (for country detection)
                    country_text = ""
                    if len(address_divs) >= 3:
                        country_text = address_divs[2].text.strip()

                    # Extract zip from the second div (city, state, zip)
                    # Zip is the last word (after last space) that's 5 digits
                    if len(address_divs) >= 2:
                        city_state_zip = address_divs[1].text.strip()
                        # Split by space and get last element as zip
                        parts = city_state_zip.split()
                        if parts and len(parts[-1]) == 5 and parts[-1].isdigit():
                            zip_code = parts[-1]
                            # Remove zip from city_state_zip for address
                            city_state_no_zip = " ".join(parts[:-1])
                        else:
                            city_state_no_zip = city_state_zip

                        # Build full address from first div (street) + second div without zip
                        street = address_divs[0].text.strip() if len(address_divs) >= 1 else ""
                        full_address = f"{street} {city_state_no_zip}".strip()

                        # Use country text for detection (more accurate than address)
                        if country_text:
                            country_name, country_code = self._detect_country(country_text)
                    else:
                        # Fallback if structure is different
                        address_parts = [address_divs[i].text.strip() for i in range(min(2, len(address_divs))) if address_divs[i].text.strip()]
                        full_address = " ".join(address_parts)

                except Exception as e:
                    print(f"[!] Could not extract address element: {e}")

            except Exception as e:
                print(f"[!] Could not find asset-profile section: {e}")

            # Detect country from address (only if not already detected from third div)
            try:
                if not country_name or country_name == "":
                    country_name, country_code = self._detect_country(full_address)
            except:
                country_name, country_code = self._detect_country(full_address)

            # If zip code wasn't extracted from div structure, try regex as fallback
            if not zip_code or zip_code == "":
                import re
                zip_match = re.search(r'\b(\d{5})\b', full_address)
                zip_code = zip_match.group(1) if zip_match else "00000"

            print(f"[OK] {company_name}")
            print(f"     Country: {country_name} (Code: {country_code})")
            print(f"     Address: {full_address}")
            print(f"     Zip: {zip_code}")

            driver.quit()

            return {
                "company_name": company_name,
                "company_address": full_address,
                "zip_code": zip_code,
                "country_name": country_name,
                "country_code": country_code
            }

        except Exception as e:
            print(f"[!] Error scraping profile for {symbol}: {e}")
            if driver:
                driver.quit()
            return {
                "company_name": f"{symbol} Inc.",
                "company_address": "United States",
                "zip_code": "00000",
                "country_name": "UNITED STATES OF AMERICA",
                "country_code": "2"
            }

    def _read_client_statement(self, client_statement_path=None):
        """Reads closing balance from E*TRADE ClientStatement PDF."""
        # Auto-find ClientStatements_*.pdf in inputs folder
        if client_statement_path is None:
            import glob
            pattern = "etrade_inputs/ClientStatements_*.pdf"
            matching_files = glob.glob(pattern)

            if matching_files:
                client_statement_path = matching_files[0]  # Use first match
                print(f"[*] Found ClientStatement: {client_statement_path}")
            else:
                print(f"[i] No ClientStatement PDF found (etrade_inputs/ClientStatements_*.pdf)")
                print(f"[i] Will use calculated closing balance from holdings")
                return None

        if not os.path.exists(client_statement_path):
            print(f"[i] ClientStatement not found: {client_statement_path}")
            print(f"[i] Will use calculated closing balance from holdings")
            return None

        try:
            import PyPDF2
            import re

            print(f"[*] Reading ClientStatement: {client_statement_path}")

            with open(client_statement_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)

                # Read first page
                page = pdf_reader.pages[0]
                text = page.extract_text()

                # Extract account number (E*TRADE format: "146 - 239025 - 205 - 4 - 1" at top of statement)
                # The main account number is the first 9 digits (146239025)
                account_match = re.search(r'(\d{3})\s*-\s*(\d{6})\s*-\s*\d{3}\s*-\s*\d\s*-\s*\d', text)
                if account_match:
                    account_number = account_match.group(1) + account_match.group(2)  # Combine first two parts
                    print(f"[OK] Found Account Number: {account_number}")
                    # Store it for later use
                    self.extracted_account_number = account_number
                else:
                    self.extracted_account_number = None
                    print(f"[!] Could not extract account number from ClientStatement")

                # Extract "Ending Total Value (as of MM/DD/YY) $XX,XXX.XX"
                match = re.search(r'Ending Total Value.*?\$([0-9,]+\.[0-9]{2})', text)

                if match:
                    ending_value_str = match.group(1).replace(',', '')
                    ending_value_usd = float(ending_value_str)

                    print(f"[OK] Found Ending Total Value: ${ending_value_usd:,.2f}")
                    return ending_value_usd
                else:
                    print(f"[!] Could not find Ending Total Value in ClientStatement")
                    return None

        except ImportError:
            print(f"[!] PyPDF2 not installed. Run: pip install PyPDF2")
            print(f"[i] Will use calculated closing balance from holdings")
            return None
        except Exception as e:
            print(f"[!] Error reading ClientStatement: {e}")
            print(f"[i] Will use calculated closing balance from holdings")
            return None

    def _scrape_yahoo_finance(self, symbol):
        """Scrapes stock price data from Yahoo Finance using Selenium."""
        print(f"[*] Scraping {symbol} stock prices from Yahoo Finance...")
        print(f"[*] This may take 30-60 seconds, please wait...")

        # Setup Chrome options
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # Run in background
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.page_load_strategy = 'eager'  # Don't wait for all resources

        driver = None
        try:
            # Initialize driver - use system Chrome directly (no ChromeDriverManager for corporate networks)
            driver = webdriver.Chrome(options=chrome_options)

            # Navigate to history page
            url = f"https://finance.yahoo.com/quote/{symbol}/history/"
            print(f"[*] Opening {url}")
            driver.get(url)
            time.sleep(5)

            print("[*] Setting custom date range...")
            try:
                # Click on the date picker to open it
                date_picker = driver.find_element(By.CSS_SELECTOR, "div[data-testid='history-date-picker']")
                date_picker.click()
                time.sleep(3)

                # Find and fill the startDate field (format: mm/dd/yyyy)
                # Yahoo Finance might exclude boundaries, so request one day earlier
                start_date_str = f"12/31/{self.calendar_year - 1}"
                print(f"[*] Setting start date to {start_date_str} (to ensure we get 01/01/{self.calendar_year})...")
                start_date_input = driver.find_element(By.CSS_SELECTOR, "input[name='startDate']")
                start_date_input.clear()
                time.sleep(0.5)
                start_date_input.send_keys(start_date_str)
                time.sleep(1)

                # Find and fill the endDate field (format: mm/dd/yyyy)
                # Request one day later to ensure we get 12/31 of target year
                end_date_str = f"01/01/{self.calendar_year + 1}"
                print(f"[*] Setting end date to {end_date_str} (to ensure we get 12/31/{self.calendar_year})...")
                end_date_input = driver.find_element(By.CSS_SELECTOR, "input[name='endDate']")
                end_date_input.clear()
                time.sleep(0.5)
                end_date_input.send_keys(end_date_str)
                time.sleep(1)

                # Try multiple selectors for Done button
                print("[*] Looking for Done button...")
                done_clicked = False
                done_selectors = [
                    "//button[text()='Done']",
                    "//button[contains(text(), 'Done')]",
                    "//span[text()='Done']/parent::button",
                    "//button[@type='submit']",
                    "button[type='submit']"
                ]

                for selector in done_selectors:
                    try:
                        if selector.startswith('//'):
                            done_button = driver.find_element(By.XPATH, selector)
                        else:
                            done_button = driver.find_element(By.CSS_SELECTOR, selector)
                        done_button.click()
                        done_clicked = True
                        print(f"[OK] Set date range: {start_date_str} - {end_date_str} (will filter to {self.calendar_year}), clicked Done")
                        break
                    except:
                        continue

                if not done_clicked:
                    print("[!] Could not find Done button, pressing Enter on end date field...")
                    from selenium.webdriver.common.keys import Keys
                    end_date_input.send_keys(Keys.RETURN)

                # Wait for table to reload with full year data
                print("[*] Waiting for table to reload with full year data...")
                time.sleep(10)

            except Exception as e:
                print(f"[!] Could not set date range: {str(e)[:80]}")
                print("[!] Using default view (last 12 months only)")

            # Scrape table data
            print("[*] Scraping price table...")
            rows = driver.find_elements(By.XPATH, "//table[@data-test='historical-prices']//tbody//tr")

            if not rows:
                # Try alternative selector
                rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

            if not rows:
                raise Exception("Could not find price table")

            print(f"[*] Found {len(rows)} rows in table, filtering for year {self.calendar_year}...")

            data = []
            for row in rows:
                try:
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if len(cols) >= 6:
                        date_str = cols[0].text.strip()
                        # Use Adj Close (Column 5) instead of Close (Column 4)
                        # Adj Close accounts for splits, dividends, and other corporate actions
                        adj_close_price = cols[5].text.replace(',', '').strip()

                        # Parse date
                        date_obj = pd.to_datetime(date_str, errors='coerce')
                        if pd.isna(date_obj):
                            continue

                        date_formatted = date_obj.strftime('%Y-%m-%d')

                        # Filter for target year only
                        if self.start_date <= date_formatted <= self.end_date:
                            data.append({
                                'Date': date_formatted,
                                'Stock_Close_USD': float(adj_close_price)
                            })
                except:
                    continue

            if not data:
                raise Exception(f"No data found for {self.calendar_year}")

            df = pd.DataFrame(data)
            df = df.sort_values('Date')
            print(f"[OK] Scraped {len(df)} trading days for {symbol} in {self.calendar_year}")
            return df

        except Exception as e:
            print(f"[!] ERROR scraping Yahoo Finance: {str(e)}")
            print(f"[!] Falling back to yfinance API")

            # Fallback to yfinance API
            try:
                import yfinance as yf
                ticker = yf.Ticker(symbol)
                df_stock = yf.download(symbol, start=self.start_date, end=self.end_date, progress=False)

                if not df_stock.empty:
                    if isinstance(df_stock.columns, pd.MultiIndex):
                        df_stock = df_stock['Close']
                    df_stock = df_stock.reset_index()
                    df_stock['Date'] = pd.to_datetime(df_stock['Date']).dt.strftime('%Y-%m-%d')
                    df_stock = df_stock.rename(columns={'Close': 'Stock_Close_USD'})
                    df_stock = df_stock[['Date', 'Stock_Close_USD']]
                    print(f"[OK] Fetched {len(df_stock)} days via yfinance API")
                    return df_stock
            except:
                pass

            return pd.DataFrame()

        finally:
            if driver:
                driver.quit()

    def get_company_details(self, symbol):
        """Fetches company info and stock prices via web scraping."""
        if symbol in self.company_cache:
            return self.company_cache[symbol]

        print(f"\n[*] Processing stock: {symbol}")

        # Default company details
        company_name = f"{symbol} Inc."
        address = "United States"
        zip_code = ""

        # Get company details from config
        config = load_config()
        if symbol in config.get("table_a3_companies", {}):
            company_info = config["table_a3_companies"][symbol]
            company_name = company_info.get("company_name", f"{symbol} Inc.")
            address = company_info.get("company_address", "United States")
            zip_code = company_info.get("zip_code", "")
        else:
            # Scrape company profile from Yahoo Finance if not in config
            profile = self._scrape_company_profile(symbol)
            company_name = profile.get("company_name", f"{symbol} Inc.")
            address = profile.get("company_address", "United States")
            zip_code = profile.get("zip_code", "")

        # Scrape stock prices
        df_stock = self._scrape_yahoo_finance(symbol)

        if df_stock.empty:
            raise ValueError(f"Could not fetch stock prices for {symbol}")

        # Build valuation matrix - use ALL US trading days
        # Per Indian tax law (Rule 115): when SBI rate unavailable (weekend/holiday),
        # use the last preceding trading day's rate

        # Remove duplicates from stock data and SBI data before merge (keep last occurrence)
        df_stock_clean = df_stock[['Date', 'Stock_Close_USD']].drop_duplicates(subset=['Date'], keep='last')
        df_sbi_clean = self.df_sbi[['Date', 'TTBR']].drop_duplicates(subset=['Date'], keep='last')

        df_matrix = pd.merge(
            df_stock_clean,
            df_sbi_clean,
            on='Date',
            how='left'  # LEFT join = keep all US trading days, even if SBI rate missing
        )

        # Sort by date first (required for forward-fill)
        df_matrix = df_matrix.sort_values('Date').reset_index(drop=True)

        # Forward-fill missing TTBR with previous day's rate
        # This implements Rule 115: "use rate from last preceding working day"
        df_matrix['TTBR'] = df_matrix['TTBR'].ffill()

        # Drop any rows still missing TTBR (start of dataset, before any SBI data available)
        rows_before = len(df_matrix)
        df_matrix = df_matrix.dropna(subset=['TTBR'])
        rows_after = len(df_matrix)
        if rows_before > rows_after:
            print(f"[i] Dropped {rows_before - rows_after} dates before first available SBI rate")

        # Calculate INR valuation per share (for all US trading days)
        df_matrix['Valuation_Per_Share_INR'] = df_matrix['Stock_Close_USD'] * df_matrix['TTBR']

        data = {
            "name": company_name,
            "address": address,
            "zip": zip_code,
            "matrix": df_matrix
        }
        self.company_cache[symbol] = data
        return data

    def calculate_tranche_values(self, symbol, qty, acq_date_str, sell_date_str=None, unit_cost_usd=0.0):
        """Computes Initial, Peak, and Closing values in INR for a lot based on holding period."""
        comp_data = self.get_company_details(symbol)
        df_matrix = comp_data["matrix"]

        hold_start = max(self.start_date, acq_date_str)
        hold_end = sell_date_str if (sell_date_str and sell_date_str <= self.end_date) else self.end_date

        window = df_matrix[(df_matrix['Date'] >= hold_start) & (df_matrix['Date'] <= hold_end)]
        if window.empty:
            window = df_matrix.tail(1)

        # 1. Initial Value (Cost basis at acquisition)
        # IMPORTANT: Use ACTUAL acquisition date for initial value, not hold_start
        # For shares acquired before FY (e.g., Nov 2024), we need the Nov 2024 TTBR
        init_row = df_matrix[df_matrix['Date'] == acq_date_str]

        if not init_row.empty:
            ttbr_init = init_row['TTBR'].values[0]
        else:
            # Acquisition date not in our matrix (before FY start or non-trading day)
            # Try to get TTBR from SBI data directly
            acq_ttbr_row = self.df_sbi[self.df_sbi['Date'] == acq_date_str]
            if not acq_ttbr_row.empty:
                ttbr_init = acq_ttbr_row['TTBR'].values[0]
            else:
                # Acquisition date not found (weekend/holiday)
                # Use nearest previous trading day TTBR
                sbi_before = self.df_sbi[self.df_sbi['Date'] < acq_date_str]
                if not sbi_before.empty:
                    closest_date = sbi_before['Date'].max()
                    ttbr_init = sbi_before[sbi_before['Date'] == closest_date]['TTBR'].values[0]
                    print(f"[i] {acq_date_str} is weekend/holiday, using previous trading day {closest_date} TTBR: {ttbr_init:.2f}")
                else:
                    # Really can't find anything, use fallback
                    ttbr_init = 83.50
                    print(f"[!] WARNING: TTBR not found for {acq_date_str}, using fallback {ttbr_init}")

        if unit_cost_usd > 0:
            initial_val = round(qty * unit_cost_usd * ttbr_init, 2)
        else:
            # Use market price on acquisition date if no unit cost provided
            if not init_row.empty:
                initial_val = round(qty * init_row['Valuation_Per_Share_INR'].values[0], 2)
            else:
                # Can't calculate without acquisition date data
                initial_val = 0
                print(f"[!] WARNING: Cannot calculate initial value for {acq_date_str}")

        # 2. Peak Value during holding window
        peak_idx = window['Valuation_Per_Share_INR'].idxmax()
        peak_val = round(qty * window['Valuation_Per_Share_INR'].max(), 2)

        # Get peak details for reporting
        peak_row = window.loc[peak_idx]
        peak_date = peak_row['Date']
        peak_price_usd = peak_row['Stock_Close_USD']
        peak_ttbr = peak_row['TTBR']
        peak_per_share_inr = peak_row['Valuation_Per_Share_INR']

        # 3. Closing Value
        # If sold within this FY: Closing = 0 (no longer holding)
        # If not sold or sold after FY: Closing = value on Dec 31 (still holding)
        if sell_date_str and sell_date_str <= self.end_date:
            # Sold within the FY -> Closing balance is 0
            closing_val = 0
        else:
            # Still holding on Dec 31 -> Use Dec 31 value
            dec31_row = df_matrix[df_matrix['Date'] == self.end_date]
            close_per_share = dec31_row['Valuation_Per_Share_INR'].values[0] if not dec31_row.empty else df_matrix['Valuation_Per_Share_INR'].iloc[-1]
            closing_val = round(qty * close_per_share, 2)

        # Return values + peak details dict
        peak_details = {
            'peak_date': peak_date,
            'peak_price_usd': peak_price_usd,
            'peak_ttbr': peak_ttbr,
            'peak_per_share_inr': peak_per_share_inr,
            'hold_start': hold_start,
            'hold_end': hold_end
        }

        return initial_val, peak_val, closing_val, peak_details

    def _scan_and_reload_ttbr_if_needed(self, bystatus_path, gl_path):
        """Scan E*TRADE files for pre-FY acquisition dates and reload TTBR if needed."""
        if self._extra_ttbr_dates_loaded:
            return  # Already loaded

        extra_dates = []

        # Scan ByStatus for pre-calendar-year acquisitions
        try:
            df_bystatus = pd.read_excel(bystatus_path, sheet_name='Sellable')
            for _, row in df_bystatus.iterrows():
                if row.get('Sellable Qty.', 0) > 0:
                    acq_date_raw = row.get('Date Acquired')
                    if pd.notna(acq_date_raw):
                        acq_date = pd.to_datetime(acq_date_raw)
                        if acq_date.strftime('%Y-%m-%d') < self.start_date:
                            # Add only exact date - CSV lookup will search backward if weekend/holiday
                            extra_dates.append(acq_date.strftime('%Y-%m-%d'))
        except Exception as e:
            print(f"[!] Could not scan ByStatus for pre-calendar-year dates: {e}")

        # Scan G&L for pre-calendar-year acquisitions
        if gl_path and os.path.exists(gl_path):
            try:
                df_gl = pd.read_excel(gl_path, sheet_name='G&L_Expanded')
                df_sold = df_gl[df_gl['Record Type'] == 'Sell']
                for _, row in df_sold.iterrows():
                    acq_date_raw = row.get('Date Acquired')
                    if pd.notna(acq_date_raw):
                        acq_date = pd.to_datetime(acq_date_raw)
                        if acq_date.strftime('%Y-%m-%d') < self.start_date:
                            # Add only exact date - CSV lookup will search backward if weekend/holiday
                            extra_dates.append(acq_date.strftime('%Y-%m-%d'))
            except Exception as e:
                print(f"[!] Could not scan G&L for pre-calendar-year dates: {e}")

        # If we found pre-calendar-year dates, reload TTBR with them
        if extra_dates:
            unique_extra = sorted(set(extra_dates))
            print(f"[*] Found acquisition dates before calendar year {self.calendar_year}, loading historical TTBR")
            print(f"[*] Dates needed: {', '.join(unique_extra)}")
            self.df_sbi = self._fetch_sbi_rates_web(extra_dates=unique_extra)
            self._extra_ttbr_dates_loaded = True

    def process_etrade_exports(self, bystatus_path=None, gl_path=None, transaction_history_path=None, account_no=None, config=None):
        # Default paths - check inputs folder first, then root folder
        # ALL FILES ARE NOW OPTIONAL - handle partial file scenarios
        if bystatus_path is None:
            if os.path.exists("etrade_inputs/ByStatus_expanded.xlsx"):
                bystatus_path = "etrade_inputs/ByStatus_expanded.xlsx"
            elif os.path.exists("ByStatus_expanded.xlsx"):
                bystatus_path = "ByStatus_expanded.xlsx"
            else:
                bystatus_path = None  # ByStatus not found - Table A3 will be empty (no holdings)

        # G&L file is OPTIONAL - only needed if stocks were sold during the financial year
        if gl_path is None:
            if os.path.exists("etrade_inputs/G&L_Expanded.xlsx"):
                gl_path = "etrade_inputs/G&L_Expanded.xlsx"
            elif os.path.exists("G&L_Expanded.xlsx"):
                gl_path = "G&L_Expanded.xlsx"
            else:
                gl_path = None  # G&L file not found - Table A3 incomplete, Capital Gains empty

        # Transaction History CSV is OPTIONAL - only needed if dividends were received
        if transaction_history_path is None:
            if os.path.exists("etrade_inputs/Transaction_History.csv"):
                transaction_history_path = "etrade_inputs/Transaction_History.csv"
            elif os.path.exists("Transaction_History.csv"):
                transaction_history_path = "Transaction_History.csv"
            else:
                transaction_history_path = None  # Transaction History not found - this is OK

        # Note: We allow processing even without ByStatus/G&L if companies exist in config
        # This enables ClientStatement-only mode for Table A2 generation

        # IMPORTANT: Scan files for pre-FY acquisition dates and reload TTBR if needed
        # This ensures we have TTBR for dates like Nov 2024 for initial value calculations
        self._scan_and_reload_ttbr_if_needed(bystatus_path, gl_path)

        # Parse dividend data from Transaction History (optional - dividends may not exist)
        df_dividends = self._parse_dividend_data(transaction_history_path)

        # Store dividend details for per-lot allocation later
        # We'll calculate each lot's dividend share based on holdings on dividend date
        dividend_transactions = []
        if not df_dividends.empty:
            print(f"[*] Found {len(df_dividends)} dividend transactions")
            for _, div_row in df_dividends.iterrows():
                dividend_transactions.append({
                    'symbol': div_row['Symbol'],
                    'date': div_row['Date'],
                    'amount_usd': div_row['Amount (USD)'],
                    'amount_inr': div_row['Amount (INR)']
                })
                print(f"    {div_row['Symbol']}: {div_row['Date']} - ${div_row['Amount (USD)']} = INR {div_row['Amount (INR)']:,.2f}")

        # Read ByStatus file if available
        df_open = pd.DataFrame()
        if bystatus_path:
            try:
                df_bystatus = pd.read_excel(bystatus_path, sheet_name='Sellable')
                df_open = df_bystatus[df_bystatus['Record Type'].isin(['Purchase', 'Grant'])].copy()
            except Exception as e:
                print(f"[!] WARNING: Error reading ByStatus file: {e}")
                print(f"[!] Table A3 holdings will be empty")
                df_open = pd.DataFrame()
        else:
            print(f"[!] WARNING: ByStatus_expanded.xlsx not found")
            print(f"[!] Table A3 will not include current holdings (only sales from G&L if available)")

        # Try to read G&L file if it exists
        df_sold = pd.DataFrame()  # Empty dataframe if no sales
        df_sold_future = pd.DataFrame()  # Shares that will be sold after this FY
        if gl_path and os.path.exists(gl_path):
            try:
                df_gl = pd.read_excel(gl_path, sheet_name='G&L_Expanded')
                df_sold_all = df_gl[df_gl['Record Type'] == 'Sell'].copy()

                # IMPORTANT: We need to handle TWO types of sold shares:
                # 1. Sold IN this FY: Show Closing Balance = 0, Gross Proceeds = actual proceeds
                # 2. Sold AFTER this FY: Show Closing Balance > 0, Gross Proceeds = 0

                df_sold_all['Date Acquired'] = pd.to_datetime(df_sold_all['Date Acquired'])
                df_sold_all['Date Sold'] = pd.to_datetime(df_sold_all['Date Sold'])

                # CRITICAL DISTINCTION:
                # - Table A3 (Schedule FA) uses CALENDAR YEAR (Jan 1 - Dec 31)
                # - Capital Gains (Schedule CG) uses EXTENDED PERIOD (Jan 1 - Mar 31 next year)
                # These are TWO DIFFERENT things per ITRFA.in guidance!

                # === FOR TABLE A3 (Calendar Year: Jan 1 - Dec 31) ===
                # Category 1: Sold within CALENDAR YEAR (for Table A3 disclosure)
                df_sold_calendar = df_sold_all[
                    (df_sold_all['Date Acquired'] <= self.end_date) &
                    (df_sold_all['Date Sold'] >= self.start_date) &
                    (df_sold_all['Date Sold'] <= self.end_date)
                ].copy()

                # Category 2: Held in CALENDAR YEAR but sold AFTER calendar year ends
                df_sold_future = df_sold_all[
                    (df_sold_all['Date Acquired'] <= self.end_date) &
                    (df_sold_all['Date Sold'] > self.end_date)
                ].copy()

                # Category 3: Sales BEFORE this calendar year (excluded from A3 for tracking)
                df_sold_before_fy = df_sold_all[
                    (df_sold_all['Date Sold'] < pd.to_datetime(self.start_date))
                ].copy()

                # Use calendar year sales for Table A3
                df_sold = df_sold_calendar

                # === FOR CAPITAL GAINS SHEET (Extended: Jan 1 - Mar 31 next year) ===
                # This captures sales from Jan-Mar next calendar year that fall in same Indian FY
                df_sold_cg = df_sold_all[
                    (df_sold_all['Date Acquired'] <= self.cg_end_date) &
                    (df_sold_all['Date Sold'] >= self.cg_start_date) &
                    (df_sold_all['Date Sold'] <= self.cg_end_date)
                ].copy()

                print(f"[OK] Table A3: {len(df_sold_calendar)} sales in calendar year {self.calendar_year}")
                print(f"[OK] Capital Gains: {len(df_sold_cg)} sales in extended period (Jan {self.calendar_year} - Mar {self.calendar_year + 1})")
                print(f"[OK] Found {len(df_sold_future)} holdings that will be sold AFTER {self.cg_end_date}")

                excluded = len(df_sold_before_fy)
                if excluded > 0:
                    print(f"[i] Excluded {excluded} sales from BEFORE FY {self.indian_fy} (will appear in 'Excluded from A3' sheet)")

            except Exception as e:
                print(f"[!] WARNING: Error reading G&L file: {e}")
                print("[i] Continuing without sold shares data")
                df_sold = pd.DataFrame()
        else:
            print("")
            print("=" * 70)
            print("[WARNING] G&L_Expanded.xlsx not found!")
            print("")
            print("This file is REQUIRED if you sold ANY shares during the year.")
            print("Without it:")
            print("  - Table A3 will be INCOMPLETE (missing sold shares)")
            print("  - Capital Gains will be EMPTY (missing tax calculations)")
            print("")
            print("Only continue if you are CERTAIN you had ZERO sales this year.")
            print("=" * 70)
            print("")
            print("[i] Continuing without sales data...")

        equity_tranches = []

        # =====================================================================
        # Table A3 Structure: SEPARATE ROWS for partial sales
        # - Partial sale = TWO separate rows (one for holding, one for sold)
        # - Each row has genuinely different peak and closing values
        # - NEVER consolidate into single row with reduced share count
        # =====================================================================

        # 1. Parse Open Lots (Unsold shares) - SEPARATE ROW for each
        for _, row in df_open.iterrows():
            qty = int(row['Sellable Qty.'])
            if qty == 0:
                continue

            # IMPORTANT: Only include holdings acquired ON OR BEFORE the end of the calendar year
            acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
            if acq_date > self.end_date:
                continue

            # Get symbol from row, or infer if only one company in portfolio
            if pd.notna(row['Symbol']):
                symbol = str(row['Symbol']).strip()
            elif len(self.company_cache) == 1:
                symbol = list(self.company_cache.keys())[0]
                print(f"[i] Symbol missing for row, using {symbol} (only company in portfolio)")
            else:
                raise ValueError(f"Symbol column missing for row with acquisition date {acq_date} and multiple companies in portfolio. Cannot determine which company.")

            comp_info = self.get_company_details(symbol)

            plan_type = str(row['Plan Type'])
            nature_prefix = "ESPP" if "ESPP" in plan_type else "RSU"
            nature = f"{nature_prefix} ({qty} shares)" if qty != 1 else f"{nature_prefix} ({qty} share)"

            # Use Purchase Date FMV for initial value
            purchase_fmv_str = str(row['Purchase Date FMV']).replace('$', '').replace(',', '').strip()
            unit_cost = float(purchase_fmv_str) if purchase_fmv_str and purchase_fmv_str != '--' else float(row['Est. Cost Basis (per share):'])

            init_val, peak_val, close_val, peak_details = self.calculate_tranche_values(symbol, qty, acq_date, unit_cost_usd=unit_cost)

            equity_tranches.append({
                "CountryName": "UNITED STATES OF AMERICA",
                "CountryCodeExcludingIndia": 2,
                "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                "ZipCode": str(comp_info["zip"]),
                "NatureOfEntity": nature,
                "InterestAcquiringDate": acq_date,
                "InitialValOfInvstmnt": init_val,
                "PeakBalanceDuringPeriod": peak_val,
                "ClosingBalance": close_val,
                "_FMV_USD": unit_cost,
                "_total_qty": qty,  # Store for dividend calculation
                "_open_qty": qty,   # All shares are open (holding)
                "_sold_qty": 0,
                "_sold_details": [],
                "_peak_details": peak_details,  # Store peak date and details
                "TotGrossAmtPaidCredited": 0,  # Will be calculated below
                "TotGrossProceeds": 0
            })

        # 2. Parse Sold Lots (Actually sold WITHIN this FY) - SEPARATE ROW for each
        # These have: Closing Balance = 0 (no longer holding)
        #            Gross Proceeds = actual proceeds from sale
        for _, row in df_sold.iterrows():
            qty = int(row['Quantity'])

            # Get symbol from row, or infer if only one company in portfolio
            if pd.notna(row['Symbol']):
                symbol = str(row['Symbol']).strip()
            elif len(self.company_cache) == 1:
                symbol = list(self.company_cache.keys())[0]
                print(f"[i] Symbol missing for sold row, using {symbol} (only company in portfolio)")
            else:
                raise ValueError(f"Symbol column missing for sold row with quantity {qty} and multiple companies in portfolio. Cannot determine which company.")

            comp_info = self.get_company_details(symbol)

            # Determine if it's RSU or ESPP based on Plan Type column
            plan_type = str(row.get('Plan Type', ''))
            is_espp = 'ESPP' in plan_type.upper() or 'EMPLOYEE STOCK PURCHASE' in plan_type.upper()

            nature = f"ESPP ({qty} shares) Sold" if is_espp else f"RSU ({qty} shares) Sold"
            if qty == 1:
                nature = f"ESPP ({qty} share) Sold" if is_espp else f"RSU ({qty} share) Sold"

            acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
            sell_date = pd.to_datetime(row['Date Sold']).strftime('%Y-%m-%d')

            # CRITICAL: Use correct FMV per Section 49(2AA)
            if is_espp and 'Purchase Date Fair Mkt. Value' in row and pd.notna(row['Purchase Date Fair Mkt. Value']):
                unit_cost = float(row['Purchase Date Fair Mkt. Value'])
            else:
                unit_cost = float(row['Adjusted Cost Basis Per Share'])

            proceeds_usd = float(row['Total Proceeds'])

            init_val, peak_val, close_val, peak_details = self.calculate_tranche_values(symbol, qty, acq_date, sell_date_str=sell_date, unit_cost_usd=unit_cost)

            df_matrix = comp_info["matrix"]
            sell_row = df_matrix[df_matrix['Date'] == sell_date]
            sell_ttbr = sell_row['TTBR'].values[0] if not sell_row.empty else 89.47
            proceeds_inr = round(proceeds_usd * sell_ttbr, 2)

            equity_tranches.append({
                "CountryName": "UNITED STATES OF AMERICA",
                "CountryCodeExcludingIndia": 2,
                "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                "ZipCode": str(comp_info["zip"]),
                "NatureOfEntity": nature,
                "InterestAcquiringDate": acq_date,
                "InitialValOfInvstmnt": init_val,
                "PeakBalanceDuringPeriod": peak_val,
                "ClosingBalance": close_val,  # 0 because sold
                "_FMV_USD": unit_cost,
                "_SaleDate": sell_date,
                "_GrossProceeds": proceeds_inr,
                "_total_qty": qty,  # Store for dividend calculation
                "_open_qty": 0,     # None are open (all sold)
                "_sold_qty": qty,
                "_sold_details": [{
                    'qty': qty,
                    'sell_date': sell_date,
                    'proceeds_usd': proceeds_usd,
                    'proceeds_inr': proceeds_inr
                }],
                "_peak_details": peak_details,  # Store peak date and details
                "TotGrossAmtPaidCredited": 0,  # Will be calculated below
                "TotGrossProceeds": proceeds_inr
            })

        # 3. Parse Future-Sold Lots (Held in this FY but sold AFTER FY ends)
        # These have: Closing Balance = closing value on Dec 31 (still holding)
        #            Gross Proceeds = 0 (not sold yet in this FY)
        for _, row in df_sold_future.iterrows():
            qty = int(row['Quantity'])

            # Get symbol from row, or infer if only one company in portfolio
            if pd.notna(row['Symbol']):
                symbol = str(row['Symbol']).strip()
            elif len(self.company_cache) == 1:
                symbol = list(self.company_cache.keys())[0]
                print(f"[i] Symbol missing for future-sold row, using {symbol} (only company in portfolio)")
            else:
                raise ValueError(f"Symbol column missing for future-sold row with quantity {qty} and multiple companies in portfolio. Cannot determine which company.")

            comp_info = self.get_company_details(symbol)

            plan_type = str(row.get('Plan Type', ''))
            is_espp = 'ESPP' in plan_type.upper() or 'EMPLOYEE STOCK PURCHASE' in plan_type.upper()
            nature_prefix = "ESPP" if is_espp else "RSU"
            nature = f"{nature_prefix} ({qty} shares) - Sold" if qty != 1 else f"{nature_prefix} ({qty} share) - Sold"

            acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
            sell_date = pd.to_datetime(row['Date Sold']).strftime('%Y-%m-%d')

            # CRITICAL: Use correct FMV per Section 49(2AA)
            if is_espp and 'Purchase Date Fair Mkt. Value' in row and pd.notna(row['Purchase Date Fair Mkt. Value']):
                unit_cost = float(row['Purchase Date Fair Mkt. Value'])
            else:
                unit_cost = float(row['Adjusted Cost Basis Per Share'])

            proceeds_usd = float(row['Total Proceeds'])

            # Get TTBR for sell date (may need to fetch if not in current year)
            df_matrix = comp_info["matrix"]
            sell_row = df_matrix[df_matrix['Date'] == sell_date]
            if not sell_row.empty:
                sell_ttbr = sell_row['TTBR'].values[0]
            else:
                sell_ttbr = 89.47  # Fallback
            proceeds_inr = round(proceeds_usd * sell_ttbr, 2)

            # Calculate values WITHOUT sell date (so closing balance is > 0)
            init_val, peak_val, close_val, peak_details = self.calculate_tranche_values(symbol, qty, acq_date, unit_cost_usd=unit_cost)

            equity_tranches.append({
                "CountryName": "UNITED STATES OF AMERICA",
                "CountryCodeExcludingIndia": 2,
                "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                "ZipCode": str(comp_info["zip"]),
                "NatureOfEntity": nature,
                "InterestAcquiringDate": acq_date,
                "InitialValOfInvstmnt": init_val,
                "PeakBalanceDuringPeriod": peak_val,
                "ClosingBalance": close_val,  # > 0 because still holding on Dec 31
                "_FMV_USD": unit_cost,
                "_SaleDate": sell_date,
                "_GrossProceeds": proceeds_inr,
                "_total_qty": qty,
                "_open_qty": qty,   # All shares still holding (future sold)
                "_sold_qty": 0,
                "_sold_details": [],
                "_peak_details": peak_details,  # Store peak date and details
                "TotGrossAmtPaidCredited": 0,  # Will be calculated below
                "TotGrossProceeds": 0  # 0 because not sold yet in this FY
            })

        # 4. Calculate dividends per lot based on holdings on dividend payment date
        print("\n[*] Calculating dividends per lot...")
        if dividend_transactions:
            for div in dividend_transactions:
                div_symbol = div['symbol']
                div_date = div['date']
                div_amount_inr = div['amount_inr']

                print(f"\n  Dividend: {div_symbol} on {div_date} - INR {div_amount_inr:,.2f}")

                # Find all lots of this symbol that were held on dividend date
                lots_held = []
                for tranche in equity_tranches:
                    if 'Beneficial Interest' in tranche['NatureOfEntity']:
                        continue  # Skip unvested

                    # Extract symbol from NameOfEntity (or use stored symbol)
                    # Match dividend symbol to tranche company
                    tranche_symbol = div_symbol  # Assuming single symbol for now

                    acq_date = tranche['InterestAcquiringDate']

                    # Check if this lot was held on dividend date
                    if acq_date <= div_date:
                        # Check if sold before dividend date
                        was_sold_before_div = False
                        shares_on_div_date = tranche['_total_qty']

                        # Check sold details
                        for sale in tranche['_sold_details']:
                            if sale['sell_date'] <= div_date:
                                # Sold before or on dividend date - reduce shares
                                shares_on_div_date -= sale['qty']
                                if shares_on_div_date == 0:
                                    was_sold_before_div = True
                                    break

                        if not was_sold_before_div and shares_on_div_date > 0:
                            lots_held.append({
                                'tranche': tranche,
                                'shares': shares_on_div_date
                            })

                # Calculate total shares held on dividend date
                total_shares_on_div_date = sum(lot['shares'] for lot in lots_held)

                if total_shares_on_div_date == 0:
                    print(f"    WARNING: No shares held on {div_date} for {div_symbol}")
                    continue

                print(f"    Total shares held on {div_date}: {total_shares_on_div_date}")

                # Allocate dividend proportionally to each lot
                for lot_info in lots_held:
                    tranche = lot_info['tranche']
                    shares = lot_info['shares']

                    # Calculate this lot's share of dividend
                    lot_dividend = round((shares / total_shares_on_div_date) * div_amount_inr, 2)

                    # Add to existing dividend (in case of multiple dividend payments)
                    tranche['TotGrossAmtPaidCredited'] += lot_dividend

                    print(f"      {tranche['NatureOfEntity']:40s} | {shares:3d} shares | INR {lot_dividend:,.2f}")

        else:
            print("  No dividend transactions found")

        # 5. Process Unvested RSUs (Beneficial Interest) - OPTIONAL
        # Per ITRFA.in: "conservative practice; some CAs defer until vesting"
        # Controlled by config.json: "disclose_unvested_rsu" (default: false)
        config = load_config()
        if config.get('disclose_unvested_rsu', False):
            try:
                print("\n[*] Checking for unvested RSUs (conservative disclosure enabled)...")
                df_unvested = pd.read_excel(bystatus_path, sheet_name='Unvested')

                # Get all companies from company_cache
                symbols_in_portfolio = sorted(self.company_cache.keys())

                # Process unvested RSUs for EACH company
                for symbol in symbols_in_portfolio:
                    comp_info = self.get_company_details(symbol)
                    df_matrix = comp_info["matrix"]

                    # Find unvested grants for this company
                    unvested_grants = df_unvested[
                        (df_unvested['Symbol'] == symbol) &
                        (df_unvested['Plan Type'].notna()) &
                        (df_unvested['Unvested Qty.'].notna()) &
                        (df_unvested['Unvested Qty.'] > 0)
                    ]

                    if unvested_grants.empty:
                        continue  # No unvested RSUs for this company

                    total_unvested = int(unvested_grants['Unvested Qty.'].sum())

                    # Get earliest grant date
                    grant_dates = pd.to_datetime(unvested_grants['Grant Date'], errors='coerce').dropna()
                    if not grant_dates.empty:
                        earliest_grant_date = grant_dates.min().strftime('%Y-%m-%d')
                    else:
                        # Use Dec 31 of current year if no grant date found
                        earliest_grant_date = self.end_date

                    print(f"[OK] Found {total_unvested} unvested {symbol} RSU units")
                    print(f"    Earliest grant date: {earliest_grant_date}")

                    # Calculate peak value (unvested qty x peak price x peak TTBR)
                    if not df_matrix.empty:
                        peak_idx = df_matrix['Valuation_Per_Share_INR'].idxmax()
                        peak_price_inr = df_matrix.loc[peak_idx, 'Valuation_Per_Share_INR']
                        peak_val = round(total_unvested * peak_price_inr, 2)
                    else:
                        peak_val = 0

                    # Calculate closing value (unvested qty x Dec 31 price x Dec 31 TTBR)
                    close_row = df_matrix[df_matrix['Date'] == self.end_date]
                    if not close_row.empty:
                        close_price_inr = close_row['Valuation_Per_Share_INR'].values[0]
                        close_val = round(total_unvested * close_price_inr, 2)
                    else:
                        # Use last available price
                        close_price_inr = df_matrix['Valuation_Per_Share_INR'].iloc[-1]
                        close_val = round(total_unvested * close_price_inr, 2)

                    # Add unvested RSUs as one aggregated A3 row per company
                    # ITR portal limit: 34 characters for NatureOfEntity
                    # "Beneficial Interest" = legal term for future right to receive shares (not yet acquired)
                    equity_tranches.append({
                        "CountryName": "UNITED STATES OF AMERICA",
                        "CountryCodeExcludingIndia": 2,
                        "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                        "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                        "ZipCode": str(comp_info["zip"]),
                        "NatureOfEntity": f"Beneficial Interest ({total_unvested} shares)",
                        "InterestAcquiringDate": earliest_grant_date,
                        "InitialValOfInvstmnt": 0,  # 0 because not acquired yet (only a promise)
                        "PeakBalanceDuringPeriod": peak_val,
                        "ClosingBalance": close_val,
                        "_FMV_USD": 0,  # N/A for unvested
                        "TotGrossAmtPaidCredited": 0,  # No dividends on unvested RSUs
                        "TotGrossProceeds": 0
                    })

                    print(f"    Initial Value: Rs. 0 (not acquired yet)")
                    print(f"    Peak Value: Rs. {peak_val:,.2f}")
                    print(f"    Closing Value: Rs. {close_val:,.2f}")

            except Exception as e:
                print(f"[i] Could not process unvested RSUs: {e}")
                print("[i] Continuing without unvested RSU data")
        else:
            print("[i] Unvested RSU disclosure disabled (config: disclose_unvested_rsu = false)")
            print("[i] Per ITRFA.in: 'some CAs defer until vesting' - this is the practical approach")
            print("[i] To enable conservative disclosure, set 'disclose_unvested_rsu': true in config.json")

        # Table A2 Custodial Account Aggregation
        # CORRECT METHOD: Calculate daily total account value and find maximum
        # (Not sum of individual peaks, since they occur on different dates)

        # Get all companies from company_cache
        symbols_in_portfolio = sorted(self.company_cache.keys())

        # Build a mapping of company name to symbol for matching tranches
        # We'll need to match tranches to their company
        # IMPORTANT: Use cleaned name (without commas) to match tranches
        company_name_to_symbol = {}
        for symbol in symbols_in_portfolio:
            comp_info = self.get_company_details(symbol)
            cleaned_name = self.clean_text_for_itr(comp_info["name"])
            company_name_to_symbol[cleaned_name] = symbol

        # Get the daily matrix (dates and TTBR) - skip if no holdings/sales
        if not symbols_in_portfolio:
            # No ByStatus or G&L - create empty daily matrix with just TTBR data
            df_daily = self.df_sbi[['Date', 'TTBR']].copy()
            df_daily['Total Account Value (USD)'] = 0
            df_daily['Total Account Value (INR)'] = 0
            total_peak_account_inr = 0
            total_peak_account_usd = 0
            peak_date = self.end_date
            peak_ttbr = 0
            print(f"[i] No holdings/sales data - A2 Peak will be 0")
        else:
            first_symbol = symbols_in_portfolio[0]
            first_company = self.get_company_details(first_symbol)
            df_daily = first_company["matrix"][['Date', 'TTBR']].copy()

            # For each day, calculate total account value across ALL companies
            daily_account_usd = []
            daily_account_inr = []

            import re
            for _, day_row in df_daily.iterrows():
                date = day_row['Date']
                ttbr = day_row['TTBR']

                # Sum value across all companies
                total_value_usd = 0

                for symbol in symbols_in_portfolio:
                    comp_info = self.get_company_details(symbol)
                    df_matrix = comp_info["matrix"]

                    # Get stock price for this date
                    price_row = df_matrix[df_matrix['Date'] == date]
                    if price_row.empty:
                        continue  # Skip if no price data for this date
                    stock_price = price_row['Stock_Close_USD'].values[0]

                    # Sum shares of this company owned on this date
                    company_shares = 0
                    for tranche in equity_tranches:
                        # Skip Beneficial Interest (unvested - not part of custodial account balance)
                        if "Beneficial Interest" in tranche['NatureOfEntity']:
                            continue

                        # Match tranche to company
                        tranche_company_name = tranche['NameOfEntity']
                        if company_name_to_symbol.get(tranche_company_name) != symbol:
                            continue  # This tranche belongs to a different company

                        acq_date = tranche['InterestAcquiringDate']
                        sell_date = tranche.get('_SaleDate')  # None if not sold

                        # Determine if we owned this holding on this date
                        # Must be: acquired by this date AND not yet sold (or sold after this date)
                        if acq_date <= date and (sell_date is None or sell_date > date):
                            # Extract quantity from nature string
                            nature = tranche['NatureOfEntity']
                            qty_match = re.search(r'\((\d+)\s+shares?\)', nature)
                            if qty_match:
                                qty = int(qty_match.group(1))
                                company_shares += qty

                    # Add this company's value to total
                    total_value_usd += company_shares * stock_price

                # Calculate total account value in INR
                total_value_inr = total_value_usd * ttbr

                daily_account_usd.append(total_value_usd)
                daily_account_inr.append(total_value_inr)

            # Add to dataframe for reference
            df_daily['Total Account Value (USD)'] = daily_account_usd
            df_daily['Total Account Value (INR)'] = daily_account_inr

            # Find peak account value
            peak_idx = df_daily['Total Account Value (INR)'].idxmax()
            peak_row = df_daily.loc[peak_idx]

            total_peak_account_inr = round(peak_row['Total Account Value (INR)'], 2)
            total_peak_account_usd = round(peak_row['Total Account Value (USD)'], 2)
            peak_date = peak_row['Date']
            peak_ttbr = round(peak_row['TTBR'], 2)

            print(f"[*] A2 Peak calculated from daily account values:")
            print(f"    Peak Date: {peak_date}")
            print(f"    Account Value: ${total_peak_account_usd:.2f} x {peak_ttbr:.2f} = {total_peak_account_inr:.2f} INR")

        # Store the daily matrix with account values for the reference sheet
        self._daily_account_matrix = df_daily

        # Try to get closing balance from ClientStatement PDF (more accurate)
        client_statement_closing_usd = self._read_client_statement()

        if client_statement_closing_usd:
            # Use ClientStatement value and convert to INR
            # Get TTBR rate for Dec 31
            if symbols_in_portfolio:
                # Get from first company's matrix (TTBR is same for all)
                first_company = self.get_company_details(symbols_in_portfolio[0])
                df_matrix = first_company["matrix"]
                dec31_row = df_matrix[df_matrix['Date'] == self.end_date]
                closing_ttbr = dec31_row['TTBR'].values[0] if not dec31_row.empty else 89.47
            else:
                # No holdings/sales - get TTBR from SBI data directly
                dec31_row = self.df_sbi[self.df_sbi['Date'] == self.end_date]
                closing_ttbr = dec31_row['TTBR'].values[0] if not dec31_row.empty else 89.47

            total_closing_account_inr = round(client_statement_closing_usd * closing_ttbr, 2)
            print(f"[OK] Using ClientStatement closing: ${client_statement_closing_usd:.2f} x {closing_ttbr:.2f} = {total_closing_account_inr:.2f} INR")
        else:
            # No ClientStatement: mark as missing
            total_closing_account_inr = 0
            print(f"[!] ClientStatement PDF not found - Table A2 closing balance will be 0")

        # Build Table A2 entry from config (single custodial account)
        acc_config = config.get("custodial_account", {})

        # Use extracted account number from ClientStatement if available, then config, then parameter, then empty
        extracted = getattr(self, 'extracted_account_number', None)
        from_config = acc_config.get("account_number")

        if extracted:
            final_account_no = extracted
            print(f"[i] Using account number from ClientStatement: {final_account_no}")
        elif from_config:
            final_account_no = from_config
            print(f"[i] Using account number from config.json: {final_account_no}")
        else:
            final_account_no = ""
            print(f"[!] Account number not found - Table A2 account number will be empty")

        # Calculate total dividends (all symbols combined)
        total_dividends_inr = int(df_dividends['Amount (INR)'].sum()) if not df_dividends.empty else 0

        # Calculate total sale proceeds (all symbols combined)
        total_sale_proceeds_inr = sum(t.get("TotGrossProceeds", 0) for t in equity_tranches)

        # Build Table A2 - Create separate rows for dividends and sale proceeds if both exist
        # Per ITRFA.in guidance: "If both dividend AND sales, create TWO A2 rows"
        custodial_accounts = []

        # Base account info (same for all rows)
        base_account_info = {
            "CountryName": acc_config.get("country_name", "UNITED STATES OF AMERICA"),
            "CountryCodeExcludingIndia": int(acc_config.get("country_code", 2)),
            "FinancialInstName": self.clean_text_for_itr(acc_config.get("financial_institution_name", "E*TRADE Securities LLC")),
            "FinancialInstAddress": self.clean_text_for_itr(acc_config.get("financial_institution_address", "1271 Avenue of the Americas New York NY 10020 United States")),
            "ZipCode": str(acc_config.get("zip_code", "10020")),
            "AccountNumber": str(final_account_no),
            "Status": acc_config.get("status", "BENEFICIAL_OWNER"),
            "AccOpenDate": acc_config.get("account_opening_date", ""),
            "PeakBalanceDuringPeriod": total_peak_account_inr,
            "ClosingBalance": total_closing_account_inr
        }

        # Case 1: Both dividends AND sales exist → Create TWO rows
        if total_dividends_inr > 0 and total_sale_proceeds_inr > 0:
            # Row 1: Dividend
            custodial_accounts.append({
                **base_account_info,
                "GrossAmtPaidCredited": total_dividends_inr,
                "NatureOfAmount": "D"  # D = Dividend
            })
            # Row 2: Sale proceeds
            custodial_accounts.append({
                **base_account_info,
                "GrossAmtPaidCredited": total_sale_proceeds_inr,
                "NatureOfAmount": "P"  # P = Proceeds from Sale or Redemption of Financial Assets
            })
            print(f"[i] Table A2: Creating TWO rows (Dividend: Rs.{total_dividends_inr:,}, Sale Proceeds: Rs.{total_sale_proceeds_inr:,})")

        # Case 2: Only dividends (no sales) → Create ONE row
        elif total_dividends_inr > 0:
            custodial_accounts.append({
                **base_account_info,
                "GrossAmtPaidCredited": total_dividends_inr,
                "NatureOfAmount": "D"  # D = Dividend
            })
            print(f"[i] Table A2: Creating ONE row (Dividend only: Rs.{total_dividends_inr:,})")

        # Case 3: Only sales (no dividends) → Create ONE row
        elif total_sale_proceeds_inr > 0:
            custodial_accounts.append({
                **base_account_info,
                "GrossAmtPaidCredited": total_sale_proceeds_inr,
                "NatureOfAmount": "P"  # P = Proceeds from Sale
            })
            print(f"[i] Table A2: Creating ONE row (Sale Proceeds only: Rs.{total_sale_proceeds_inr:,})")

        # Case 4: No dividends AND no sales → Create ONE row with N (No Amount)
        else:
            custodial_accounts.append({
                **base_account_info,
                "GrossAmtPaidCredited": 0,
                "NatureOfAmount": "N"  # N = No Amount
            })
            print(f"[i] Table A2: Creating ONE row (No dividends or sales)")


        # Sort equity_tranches by acquisition date (oldest first)
        # Keep "Beneficial Interest" at the end (it has earliest grant date but should be last)
        print("\n[*] Sorting Table A3 by acquisition date...")

        # Separate beneficial interest from regular tranches
        beneficial_interest = [t for t in equity_tranches if "Beneficial Interest" in t.get("NatureOfEntity", "")]
        regular_tranches = [t for t in equity_tranches if "Beneficial Interest" not in t.get("NatureOfEntity", "")]

        # Sort regular tranches by acquisition date (oldest first)
        regular_tranches.sort(key=lambda x: x["InterestAcquiringDate"])

        # Recombine: regular tranches first (sorted), then beneficial interest at end
        equity_tranches = regular_tranches + beneficial_interest

        # Clean internal fields (those starting with _) before exporting to JSON
        equity_tranches_clean = []
        for tranche in equity_tranches:
            clean_tranche = {k: v for k, v in tranche.items() if not k.startswith('_')}
            equity_tranches_clean.append(clean_tranche)

        output_data = {
            "ScheduleFA": {
                "AssessmentYear": self.assessment_year,
                "IndianFY": self.indian_fy,
                "DtlsForeignCustodialAcc": custodial_accounts,
                "DtlsForeignEquityDebtInterest": equity_tranches_clean,
                "DetailsOfTrustOutIndiaTrustee": []
            }
        }

        # Create outputs folder
        config = load_config()
        output_dir = "etrade_outputs"
        os.makedirs(output_dir, exist_ok=True)

        json_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}.json")
        excel_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}.xlsx")

        # Clean output directory BEFORE writing new files
        import glob
        output_pattern = os.path.join(output_dir, "schedule_fa_*")
        for old_file in glob.glob(output_pattern):
            try:
                os.remove(old_file)
                print(f"[*] Removed old file: {os.path.basename(old_file)}")
            except Exception as e:
                print(f"[!] Could not remove {old_file}: {e}")

        with open(json_filename, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2)

        df_a2 = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignCustodialAcc"])
        df_a3_raw = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignEquityDebtInterest"])

        # Track if A3 is empty for later processing
        a3_is_empty = df_a3_raw.empty
        df_a3 = df_a3_raw

        # Create Peak Value Details sheet
        peak_details_data = []
        for tranche in equity_tranches:
            if 'Beneficial Interest' in tranche['NatureOfEntity']:
                continue  # Skip unvested shares

            peak_info = tranche.get('_peak_details', {})
            peak_details_data.append({
                'Nature of Entity': tranche['NatureOfEntity'],
                'Acquisition Date': tranche['InterestAcquiringDate'],
                'Quantity': tranche['_total_qty'],
                'Peak Calculation Window Start': peak_info.get('hold_start', ''),
                'Peak Calculation Window End': peak_info.get('hold_end', ''),
                'Peak Date': peak_info.get('peak_date', ''),
                'Peak Price (USD)': round(peak_info.get('peak_price_usd', 0), 2),
                'Peak TTBR': round(peak_info.get('peak_ttbr', 0), 2),
                'Peak Value per Share (INR)': round(peak_info.get('peak_per_share_inr', 0), 2),
                'Peak Value Total (INR)': tranche['PeakBalanceDuringPeriod']
            })

        df_peak_details = pd.DataFrame(peak_details_data)

        # If A3 is empty, peak details should also show note
        if a3_is_empty or df_peak_details.empty:
            df_peak_details = pd.DataFrame({
                'Note': [f'No shares are vested in calendar year {self.calendar_year}']
            })

        # Add USD columns to A2 (for reference like ITRFA.in)
        # Calculate USD values by dividing INR by approximate TTBR
        if not df_a2.empty and 'PeakBalanceDuringPeriod' in df_a2.columns:
            # Use the peak TTBR from calculation
            peak_ttbr = peak_ttbr if 'peak_ttbr' in locals() else 89.47
            df_a2.insert(df_a2.columns.get_loc('PeakBalanceDuringPeriod'), 'Peak Balance (USD)',
                        round(df_a2['PeakBalanceDuringPeriod'] / peak_ttbr, 2))
            df_a2.insert(df_a2.columns.get_loc('ClosingBalance'), 'Closing Balance (USD)',
                        round(df_a2['ClosingBalance'] / peak_ttbr, 2))

        # Add USD columns to A3 (for reference like ITRFA.in)
        # Use the stored _FMV_USD values from equity_tranches
        # Skip if A3 is empty (will be replaced with note later)
        if not a3_is_empty and not df_a3.empty:
            # Calculate USD values from the tranches
            usd_initial = []
            usd_peak = []
            usd_closing = []
            usd_gross_amt = []
            usd_gross_proceeds = []

            for tranche in equity_tranches:
                fmv_usd = tranche.get('_FMV_USD', 0)
                qty_match = re.search(r'\((\d+)\s+shares?\)', tranche['NatureOfEntity'])
                qty = int(qty_match.group(1)) if qty_match else 1

                # Get TTBR for this tranche
                acq_date = tranche['InterestAcquiringDate']
                acq_ttbr_row = self.df_sbi[self.df_sbi['Date'] == acq_date]
                ttbr = acq_ttbr_row['TTBR'].values[0] if not acq_ttbr_row.empty else 89.47

                usd_initial.append(round(fmv_usd * qty, 2))
                usd_peak.append(round(tranche['PeakBalanceDuringPeriod'] / ttbr, 2) if tranche['PeakBalanceDuringPeriod'] > 0 else 0)
                usd_closing.append(round(tranche['ClosingBalance'] / ttbr, 2) if tranche['ClosingBalance'] > 0 else 0)
                usd_gross_amt.append(0)  # No dividends currently

                # Gross proceeds in USD
                if tranche.get('_GrossProceeds', 0) > 0:
                    # Get sale TTBR
                    if tranche.get('_SaleDate'):
                        sale_ttbr_row = self.df_sbi[self.df_sbi['Date'] == tranche['_SaleDate']]
                        sale_ttbr = sale_ttbr_row['TTBR'].values[0] if not sale_ttbr_row.empty else ttbr
                    else:
                        sale_ttbr = ttbr
                    usd_gross_proceeds.append(round(tranche['_GrossProceeds'] / sale_ttbr, 2))
                else:
                    usd_gross_proceeds.append(0)

            # Insert USD columns before INR columns
            df_a3.insert(df_a3.columns.get_loc('InitialValOfInvstmnt'), 'Initial Investment (USD)', usd_initial)
            df_a3.insert(df_a3.columns.get_loc('PeakBalanceDuringPeriod'), 'Peak Value (USD)', usd_peak)
            df_a3.insert(df_a3.columns.get_loc('ClosingBalance'), 'Closing Value (USD)', usd_closing)
            df_a3.insert(df_a3.columns.get_loc('TotGrossAmtPaidCredited'), 'Gross Amount Paid/Credited (USD)', usd_gross_amt)
            df_a3.insert(df_a3.columns.get_loc('TotGrossProceeds'), 'Gross Proceeds from Sale (USD)', usd_gross_proceeds)

        # Round all numeric values to integers for Excel display (A2 and A3 INR columns)
        numeric_cols_a2 = ['PeakBalanceDuringPeriod', 'ClosingBalance', 'GrossAmtPaidCredited']
        for col in numeric_cols_a2:
            if col in df_a2.columns:
                df_a2[col] = df_a2[col].round(0).astype(int)

        numeric_cols_a3 = ['InitialValOfInvstmnt', 'PeakBalanceDuringPeriod', 'ClosingBalance',
                           'TotGrossAmtPaidCredited', 'TotGrossProceeds']
        for col in numeric_cols_a3:
            if col in df_a3.columns:
                df_a3[col] = df_a3[col].round(0).astype(int)

        # Force ZipCode, AccountNumber, and CountryCode to be text (prevent Excel auto-conversion)
        if 'ZipCode' in df_a2.columns:
            df_a2['ZipCode'] = df_a2['ZipCode'].astype(str)
        if 'AccountNumber' in df_a2.columns:
            df_a2['AccountNumber'] = df_a2['AccountNumber'].astype(str)

        if 'ZipCode' in df_a3.columns:
            df_a3['ZipCode'] = df_a3['ZipCode'].astype(str)
        if 'CountryCodeExcludingIndia' in df_a3.columns:
            df_a3['CountryCodeExcludingIndia'] = df_a3['CountryCodeExcludingIndia'].astype(str)

        # Create reference sheet with daily stock prices and TTBR rates for ALL companies
        # Format: Date | SBI TTBR | Symbol1 Price (USD) | Symbol1 Value (INR) | Symbol2 Price (USD) | Symbol2 Value (INR) | ...

        # Get all unique symbols from equity_tranches
        symbols_in_portfolio = set()
        for tranche in equity_tranches:
            # Extract symbol from company name or use stored symbol
            # For now, we'll get it from company_cache keys
            pass

        # Get symbols from company cache (these are the ones we actually fetched data for)
        symbols_in_portfolio = sorted(self.company_cache.keys())

        # Create reference sheet - if no companies, create simple note sheet
        if not symbols_in_portfolio:
            df_reference = pd.DataFrame({
                'Note': ['No company data fetched',
                         f'ByStatus and G&L files does not have data for CY {self.calendar_year}']
            })
        else:
            # Start with Date and TTBR columns from any company's matrix (TTBR is same for all)
            first_symbol = symbols_in_portfolio[0]
            first_company = self.get_company_details(first_symbol)
            df_reference = first_company["matrix"][['Date', 'TTBR']].copy()
            df_reference = df_reference.rename(columns={'TTBR': 'SBI TTBR'})

            # Add columns for each company (Symbol Price USD, Symbol Value INR)
            for symbol in symbols_in_portfolio:
                comp_info = self.get_company_details(symbol)
                df_company = comp_info["matrix"][['Date', 'Stock_Close_USD', 'Valuation_Per_Share_INR']].copy()

                # Merge with main reference dataframe on Date
                df_reference = pd.merge(
                    df_reference,
                    df_company,
                    on='Date',
                    how='left'
                )

                # Rename columns with company symbol
                df_reference = df_reference.rename(columns={
                    'Stock_Close_USD': f'{symbol} (USD)',
                    'Valuation_Per_Share_INR': f'{symbol} (INR)'
                })

            # Round all numeric columns to 2 decimal places
            for col in df_reference.columns:
                if col != 'Date':
                    df_reference[col] = df_reference[col].round(2)

        # Create Pre-FY Acquisitions sheet for holdings acquired before the financial year
        pre_fy_data = []
        for tranche in equity_tranches:
            acq_date = tranche['InterestAcquiringDate']
            if acq_date < self.start_date:  # Acquired before FY start
                # Get the quantity from nature field
                nature = tranche['NatureOfEntity']

                # Extract actual FMV from E*TRADE file (stored during processing)
                fmv_usd = tranche.get('_FMV_USD', 0)
                initial_inr = tranche['InitialValOfInvstmnt']

                # Try to get acquisition date from our TTBR data
                acq_ttbr_row = self.df_sbi[self.df_sbi['Date'] == acq_date]
                if not acq_ttbr_row.empty:
                    ttbr_used = acq_ttbr_row['TTBR'].values[0]
                    ttbr_source = acq_date
                else:
                    # It was a weekend/holiday, find what we used
                    sbi_before = self.df_sbi[self.df_sbi['Date'] < acq_date]
                    if not sbi_before.empty:
                        ttbr_source = sbi_before['Date'].max()
                        ttbr_used = sbi_before[sbi_before['Date'] == ttbr_source]['TTBR'].values[0]
                    else:
                        ttbr_source = 'Fallback'
                        ttbr_used = 83.50

                # Extract quantity from nature string for per-share FMV
                import re
                qty_match = re.search(r'\((\d+)\s+shares?\)', nature)
                qty = int(qty_match.group(1)) if qty_match else 1

                pre_fy_data.append({
                    'Nature': nature,
                    'Acquisition Date': acq_date,
                    'Day of Week': pd.to_datetime(acq_date).strftime('%A'),
                    'Quantity': qty,
                    'FMV per Share (USD)': round(fmv_usd, 2),
                    'Total FMV (USD)': round(fmv_usd * qty, 2),
                    'TTBR Used': round(ttbr_used, 2),
                    'TTBR Date': ttbr_source,
                    'Initial Value (INR)': initial_inr,
                    'Note': 'Weekend/Holiday' if ttbr_source != acq_date else 'Trading Day'
                })

        df_pre_fy = pd.DataFrame(pre_fy_data) if pre_fy_data else pd.DataFrame({
            'Note': ['No acquisitions before FY start']
        })

        # Create A2 Peak Calculation sheet - showing daily account values
        # Check if we have actual account value columns (only present when holdings exist)
        if 'Total Account Value (USD)' in self._daily_account_matrix.columns:
            df_a2_peak = self._daily_account_matrix[['Date', 'TTBR',
                                                       'Total Account Value (USD)',
                                                       'Total Account Value (INR)']].copy()
            df_a2_peak = df_a2_peak.rename(columns={
                'Date': 'Date',
                'TTBR': 'SBI TTBR',
                'Total Account Value (USD)': 'Total Account Value (USD)',
                'Total Account Value (INR)': 'Total Account Value (INR)'
            })
            # Round values
            df_a2_peak['SBI TTBR'] = df_a2_peak['SBI TTBR'].round(2)
            df_a2_peak['Total Account Value (USD)'] = df_a2_peak['Total Account Value (USD)'].round(2)
            df_a2_peak['Total Account Value (INR)'] = df_a2_peak['Total Account Value (INR)'].round(2)

            # Add peak summary (using plain values, not formatted strings, to avoid Excel errors)
            peak_account_idx = df_a2_peak['Total Account Value (INR)'].idxmax()
            peak_account_row = df_a2_peak.loc[peak_account_idx]

            # Create summary columns with proper initialization
            summary_labels = []
            summary_values = []

            # Add peak summary data
            summary_items = [
                ('Peak Date', str(peak_account_row['Date'])),
                ('TTBR', float(peak_account_row['SBI TTBR'])),
                ('Account Value (USD)', float(peak_account_row['Total Account Value (USD)'])),
                ('Account Value (INR)', round(peak_account_row['Total Account Value (INR)'], 2)),
                ('', ''),
                ('A2 Peak Balance (INR)', round(peak_account_row['Total Account Value (INR)'], 2))
            ]

            for label, value in summary_items:
                summary_labels.append(label)
                summary_values.append(value)

            # Pad with empty strings to match DataFrame length
            while len(summary_labels) < len(df_a2_peak):
                summary_labels.append('')
                summary_values.append('')

            # Add summary columns with proper data
            df_a2_peak['--'] = ''
            df_a2_peak['--'] = ''
            df_a2_peak['PEAK SUMMARY'] = summary_labels
            df_a2_peak['Value'] = summary_values
        else:
            # No holdings - create simple sheet showing "No holdings data"
            df_a2_peak = pd.DataFrame({
                'Note': ['No ByStatus or G&L files provided',
                         'Cannot calculate daily account values',
                         'Peak is set to 0 in Table A2']
            })

        # Create Capital Gains sheet using EXTENDED PERIOD (Jan 1 - Mar 31 next year)
        # This calculates advance tax obligations based on sale date
        # Advance Tax Schedule (Income Tax Rule 234C):
        #   - By July 15: 15% of tax
        #   - By Sep 15: 45% of tax (cumulative)
        #   - By Dec 15: 75% of tax (cumulative)
        #   - By Mar 15: 100% of tax (cumulative)
        capital_gains_data = []

        # Build Capital Gains from df_sold_cg (extended period) instead of equity_tranches (calendar year)
        # This ensures we capture sales from Jan-Mar next year that are in same Indian FY
        if 'df_sold_cg' in locals() and not df_sold_cg.empty:
            for _, row in df_sold_cg.iterrows():
                sale_date = pd.to_datetime(row['Date Sold'])
                acq_date = pd.to_datetime(row['Date Acquired'])

                # Extract quantity and nature
                qty = int(row['Quantity Sold'])
                symbol = row['Symbol']
                plan_type = row.get('Plan Type', 'Stock')

                # Determine nature prefix
                if 'RSU' in str(plan_type).upper() or 'RESTRICTED' in str(plan_type).upper():
                    nature = f"RSU ({qty} shares)"
                elif 'ESPP' in str(plan_type).upper() or 'EMPLOYEE' in str(plan_type).upper():
                    nature = f"ESPP ({qty} shares)"
                else:
                    nature = f"Stock ({qty} shares)"

                # Calculate holding period in CALENDAR MONTHS (not days!)
                # Section 2(42A): "not more than 24 months" for unlisted securities (foreign shares)
                # CRITICAL: Sale on 24-month anniversary = STILL short-term (not more than 24 months includes the anniversary)
                # CRITICAL: Use calendar months, NOT 730 days (leap years: 24 months can be 731 days)
                holding_months = (sale_date.year - acq_date.year) * 12 + (sale_date.month - acq_date.month)

                # Determine tax type, rate, and section
                # Foreign company shares = UNLISTED securities (no STT on Indian exchange)
                # Threshold: 24 months (NOT 12 months which applies to STT-paid Indian listed equity)
                #
                # LTCG (Long Term): > 24 months → Section 112 (no indexation per Finance Act 2024)
                # STCG (Short Term): ≤ 24 months → Section 48 (taxed at slab rate)
                if holding_months > 24:
                    tax_type = "LTCG"
                    tax_section = "Section 112"
                    tax_rate = 0.125  # 12.5% (no indexation benefit)
                else:
                    tax_type = "STCG"
                    tax_section = "Section 48"
                    tax_rate = self.stcg_tax_rate  # User's income tax slab rate (selected at start)

                # Calculate proceeds and cost basis in INR
                import math
                proceeds_usd = float(row['Total Proceeds'])

                # CRITICAL: Use correct FMV per Section 49(2AA) for cost basis
                # - RSU: "Adjusted Cost Basis Per Share" is correct (equals FMV at vest)
                # - ESPP: Must use "Purchase Date Fair Mkt. Value" (NOT "Adjusted Cost Basis"!)
                is_espp = 'ESPP' in str(plan_type).upper() or 'EMPLOYEE' in str(plan_type).upper()
                if is_espp and 'Purchase Date Fair Mkt. Value' in row and pd.notna(row['Purchase Date Fair Mkt. Value']):
                    unit_cost_basis = float(row['Purchase Date Fair Mkt. Value'])  # FMV on purchase date (correct for ESPP)
                else:
                    # For RSU: Adjusted Cost Basis = FMV at vest (correct)
                    unit_cost_basis = float(row['Adjusted Cost Basis Per Share'])

                cost_basis_usd = unit_cost_basis * qty

                # CRITICAL: Income-tax Rule 115(1)(f) for Schedule CG (Capital Gains)
                # "For income chargeable under the head 'Capital gains', the specified date is
                # the last day of the month immediately preceding the month in which the capital
                # asset is transferred (sold)"
                #
                # This is DIFFERENT from Schedule FA which uses exact date!
                # - Schedule FA (Table A3): Exact acquisition/sale date (CBDT filing instructions)
                # - Schedule CG: Last day of month BEFORE sale month (Rule 115(1)(f))
                #
                # Example: Sale on Aug 15, 2025 → Use Jul 31, 2025 TTBR
                # Example: Sale on Jan 1, 2026 → Use Dec 31, 2025 TTBR
                #
                # BOTH proceeds and cost basis use the SAME rate (the specified date rate)

                # Calculate specified date per Rule 115(1)(f)
                if sale_date.month == 1:
                    # Sale in January → last day of December previous year
                    specified_year = sale_date.year - 1
                    specified_month = 12
                else:
                    # Last day of previous month
                    specified_year = sale_date.year
                    specified_month = sale_date.month - 1

                # Get last day of the specified month
                import calendar
                last_day = calendar.monthrange(specified_year, specified_month)[1]
                specified_date = datetime(specified_year, specified_month, last_day).strftime('%Y-%m-%d')

                # Get TTBR for the specified date (same rate for BOTH proceeds and cost basis)
                specified_ttbr_df = self.df_sbi[self.df_sbi['Date'] == specified_date]
                if not specified_ttbr_df.empty:
                    specified_ttbr = specified_ttbr_df['TTBR'].values[0]
                else:
                    # Forward-fill: use previous available rate (for weekends/holidays)
                    prior_dates = self.df_sbi[self.df_sbi['Date'] < specified_date].sort_values('Date', ascending=False)
                    specified_ttbr = prior_dates['TTBR'].values[0] if not prior_dates.empty else 85.0

                # Use SAME rate for both proceeds and cost basis (per Rule 115(1)(f))
                gross_proceeds = math.ceil(proceeds_usd * specified_ttbr)  # Round UP proceeds
                cost_basis = math.ceil(cost_basis_usd * specified_ttbr)  # Round UP cost basis (same rate!)
                capital_gain = gross_proceeds - cost_basis  # Already rounded up

                # Store base info WITHOUT tax calculation (will calculate for both regimes later)
                capital_gains_data.append({
                    'Nature': nature,
                    'Quantity': qty,
                    'Acquisition Date': acq_date_str,
                    'Sale Date': sale_date_str,
                    'Rule 115(1)(f) Specified Date': specified_date,
                    'TTBR (INR/USD)': round(specified_ttbr, 2),
                    'Holding Period (months)': holding_months,
                    'Tax Type': tax_type,
                    'Section': tax_section,
                    'Cost Basis (INR)': cost_basis,
                    'Sale Proceeds (INR)': gross_proceeds,
                    'Capital Gain (INR)': capital_gain,
                    # Tax calculation done separately for both regimes
                    'STCG_Rate_New': self.stcg_rate_new if tax_type == "STCG" else 0.125,
                    'STCG_Rate_Old': self.stcg_rate_old if tax_type == "STCG" else 0.125,
                })

        # Add future sales (after current FY) for advance tax planning
        # These sales haven't happened yet but help user plan for next year's advance tax
        if 'df_sold_future' in locals() and not df_sold_future.empty:
            for _, row in df_sold_future.iterrows():
                sale_date = pd.to_datetime(row['Date Sold'])
                acq_date = pd.to_datetime(row['Date Acquired'])

                qty = int(row['Quantity'])  # G&L file has 'Quantity', not 'Quantity Sold'
                symbol = row['Symbol']
                plan_type = row.get('Plan Type', 'Stock')

                # Determine nature prefix
                if 'RSU' in str(plan_type).upper() or 'RESTRICTED' in str(plan_type).upper():
                    nature = f"RSU ({qty} shares) - FUTURE"
                elif 'ESPP' in str(plan_type).upper() or 'EMPLOYEE' in str(plan_type).upper():
                    nature = f"ESPP ({qty} shares) - FUTURE"
                else:
                    nature = f"Stock ({qty} shares) - FUTURE"

                # Calculate holding period in CALENDAR MONTHS (not days!)
                holding_months = (sale_date.year - acq_date.year) * 12 + (sale_date.month - acq_date.month)

                # Determine tax type, rate, and section
                if holding_months > 24:
                    tax_type = "LTCG"
                    tax_section = "Section 112"
                    tax_rate = 0.125
                else:
                    tax_type = "STCG"
                    tax_section = "Section 48"
                    tax_rate = 0.312

                # Calculate proceeds and cost basis
                import math
                proceeds_usd = float(row['Total Proceeds'])

                # Use correct FMV per Section 49(2AA)
                is_espp = 'ESPP' in str(plan_type).upper() or 'EMPLOYEE' in str(plan_type).upper()
                if is_espp and 'Purchase Date Fair Mkt. Value' in row and pd.notna(row['Purchase Date Fair Mkt. Value']):
                    unit_cost_basis = float(row['Purchase Date Fair Mkt. Value'])
                else:
                    unit_cost_basis = float(row['Adjusted Cost Basis Per Share'])

                cost_basis_usd = unit_cost_basis * qty

                # CRITICAL: Income-tax Rule 115(1)(f) for Schedule CG
                # Calculate specified date per Rule 115(1)(f) - same logic as current sales
                sale_date_str = sale_date.strftime('%Y-%m-%d')
                acq_date_str = acq_date.strftime('%Y-%m-%d')

                if sale_date.month == 1:
                    specified_year = sale_date.year - 1
                    specified_month = 12
                else:
                    specified_year = sale_date.year
                    specified_month = sale_date.month - 1

                import calendar
                last_day = calendar.monthrange(specified_year, specified_month)[1]
                specified_date = datetime(specified_year, specified_month, last_day).strftime('%Y-%m-%d')

                # Get TTBR for specified date (may be in future, use latest available as estimate)
                specified_ttbr_df = self.df_sbi[self.df_sbi['Date'] == specified_date]
                if not specified_ttbr_df.empty:
                    specified_ttbr = specified_ttbr_df['TTBR'].values[0]
                else:
                    # For future sales, use latest TTBR as estimate
                    prior_dates = self.df_sbi[self.df_sbi['Date'] < specified_date].sort_values('Date', ascending=False)
                    if not prior_dates.empty:
                        specified_ttbr = prior_dates['TTBR'].values[0]
                    else:
                        # If specified date is beyond our TTBR data, use latest available
                        specified_ttbr = self.df_sbi.sort_values('Date', ascending=False)['TTBR'].values[0] if not self.df_sbi.empty else 89.47

                # Use SAME rate for both proceeds and cost basis (per Rule 115(1)(f))
                gross_proceeds = math.ceil(proceeds_usd * specified_ttbr)
                cost_basis = math.ceil(cost_basis_usd * specified_ttbr)  # Same rate!
                capital_gain = gross_proceeds - cost_basis

                # Store base info WITHOUT tax calculation (will calculate for both regimes later)
                capital_gains_data.append({
                    'Nature': nature,
                    'Quantity': qty,
                    'Acquisition Date': acq_date_str,
                    'Sale Date': sale_date_str,
                    'Rule 115(1)(f) Specified Date': specified_date,
                    'TTBR (INR/USD)': round(specified_ttbr, 2),
                    'Holding Period (months)': holding_months,
                    'Tax Type': tax_type,
                    'Section': tax_section,
                    'Cost Basis (INR)': cost_basis,
                    'Sale Proceeds (INR)': gross_proceeds,
                    'Capital Gain (INR)': capital_gain,
                    # Tax calculation done separately for both regimes
                    'STCG_Rate_New': self.stcg_rate_new if tax_type == "STCG" else 0.125,
                    'STCG_Rate_Old': self.stcg_rate_old if tax_type == "STCG" else 0.125,
                })

        # Helper function to calculate tax and advance tax for a given rate
        def calculate_tax_for_regime(sales_data, regime_name, use_old_regime=False):
            import math
            result = []
            for item in sales_data:
                # Get appropriate tax rate
                tax_rate = item['STCG_Rate_Old'] if use_old_regime else item['STCG_Rate_New']
                capital_gain = item['Capital Gain (INR)']
                tax_amount = math.ceil(capital_gain * tax_rate)

                # Calculate advance tax schedule
                sale_date = pd.to_datetime(item['Sale Date'])
                sale_month = sale_date.month
                if sale_month <= 6:
                    adv_jul = math.ceil(tax_amount * 0.15)
                    adv_sep = math.ceil(tax_amount * 0.45)
                    adv_dec = math.ceil(tax_amount * 0.75)
                    adv_mar = tax_amount
                elif sale_month <= 8:
                    adv_jul = 0
                    adv_sep = math.ceil(tax_amount * 0.45)
                    adv_dec = math.ceil(tax_amount * 0.75)
                    adv_mar = tax_amount
                elif sale_month <= 11:
                    adv_jul = 0
                    adv_sep = 0
                    adv_dec = math.ceil(tax_amount * 0.75)
                    adv_mar = tax_amount
                else:
                    adv_jul = 0
                    adv_sep = 0
                    adv_dec = 0
                    adv_mar = tax_amount

                result.append({
                    **item,  # Include all base fields
                    'Tax Rate': f"{tax_rate*100}%",
                    'Tax Amount (INR)': tax_amount,
                    'Adv Tax by Jul 15 (15%)': adv_jul,
                    'Adv Tax by Sep 15 (45%)': adv_sep,
                    'Adv Tax by Dec 15 (75%)': adv_dec,
                    'Adv Tax by Mar 15 (100%)': adv_mar,
                })
            return result

        # Calculate for BOTH regimes
        capital_gains_new_regime = calculate_tax_for_regime(capital_gains_data, "New Regime", use_old_regime=False)
        capital_gains_old_regime = calculate_tax_for_regime(capital_gains_data, "Old Regime", use_old_regime=True)

        # Generate Schedule OS and FSI (use New Regime data for now)
        print(f"\n[*] Generating Schedule OS and Schedule FSI...")
        df_schedule_os, df_schedule_fsi, df_div_os = self._calculate_schedule_os_fsi(df_dividends, pd.DataFrame(capital_gains_new_regime))
        print(f"[OK] Schedule OS: Total dividend income Rs.{df_schedule_os[self.indian_fy][2]:,}" if len(df_schedule_os) > 2 else "[OK] Schedule OS: No dividends")
        print(f"[OK] Schedule FSI: Total foreign income Rs.{df_schedule_fsi[self.indian_fy][3]:,}" if len(df_schedule_fsi) > 3 else "[OK] Schedule FSI: No foreign income")

        # Create sale details DataFrames for BOTH regimes
        df_sale_details_new = pd.DataFrame([{
            'Nature': item['Nature'],
            'Quantity': item['Quantity'],
            'Acquisition Date': item['Acquisition Date'],
            'Sale Date': item['Sale Date'],
            'Rule 115(1)(f) Specified Date': item['Rule 115(1)(f) Specified Date'],
            'TTBR (INR/USD)': item['TTBR (INR/USD)'],
            'Holding Period (months)': item['Holding Period (months)'],
            'Tax Type': item['Tax Type'],
            'Section': item['Section'],
            'Cost Basis (INR)': item['Cost Basis (INR)'],
            'Sale Proceeds (INR)': item['Sale Proceeds (INR)'],
            'Capital Gain (INR)': item['Capital Gain (INR)'],
            'Tax Rate': item['Tax Rate'],
            'Tax Amount (INR)': item['Tax Amount (INR)']
        } for item in capital_gains_new_regime])

        df_sale_details_old = pd.DataFrame([{
            'Nature': item['Nature'],
            'Quantity': item['Quantity'],
            'Acquisition Date': item['Acquisition Date'],
            'Sale Date': item['Sale Date'],
            'Rule 115(1)(f) Specified Date': item['Rule 115(1)(f) Specified Date'],
            'TTBR (INR/USD)': item['TTBR (INR/USD)'],
            'Holding Period (months)': item['Holding Period (months)'],
            'Tax Type': item['Tax Type'],
            'Section': item['Section'],
            'Cost Basis (INR)': item['Cost Basis (INR)'],
            'Sale Proceeds (INR)': item['Sale Proceeds (INR)'],
            'Capital Gain (INR)': item['Capital Gain (INR)'],
            'Tax Rate': item['Tax Rate'],
            'Tax Amount (INR)': item['Tax Amount (INR)']
        } for item in capital_gains_old_regime])

        # If no capital gains, show note in sale details
        if df_sale_details_new.empty:
            df_sale_details_new = pd.DataFrame({
                'Note': [f'No capital gains in FY {self.indian_fy}',
                         'All sales in G&L file are either in current FY or future years']
            })

        if df_sale_details_old.empty:
            df_sale_details_old = pd.DataFrame({
                'Note': [f'No capital gains in FY {self.indian_fy}',
                         'All sales in G&L file are either in current FY or future years']
            })

        # For backward compatibility, keep df_sale_details as New Regime
        df_sale_details = df_sale_details_new

        # Table 2: Advance Tax Schedule - GROUPED by deadline applicability
        # Helper function to calculate advance tax for any regime
        def calculate_advance_tax_schedule(capital_gains_list):
            """Calculate advance tax schedule for a given capital gains list."""
            if not capital_gains_list:
                return pd.DataFrame(columns=[
                    'Sale Period', 'Financial Year', 'Tax Type', 'Total Tax (INR)', 'By Jul 15',
                    'By Sep 15', 'By Dec 15', 'By Mar 15', 'Note'
                ])

            advance_tax_rows = []

            # Helper function to get FY year from a date
            def get_fy_year(date):
                return date.year if date.month >= 4 else date.year - 1

            # Group 1: Sales from Apr 1 to Jul 15 - All 4 deadlines apply
            group1_sales = [item for item in capital_gains_list
                           if (pd.to_datetime(item['Sale Date']).month >= 4 and pd.to_datetime(item['Sale Date']).month <= 6) or
                              (pd.to_datetime(item['Sale Date']).month == 7 and pd.to_datetime(item['Sale Date']).day <= 15)]
            if group1_sales:
                group1_tax = sum(item['Tax Amount (INR)'] for item in group1_sales)
                group1_jul = sum(item['Adv Tax by Jul 15 (15%)'] for item in group1_sales)
                group1_sep = sum(item['Adv Tax by Sep 15 (45%)'] for item in group1_sales)
                group1_dec = sum(item['Adv Tax by Dec 15 (75%)'] for item in group1_sales)
                group1_mar = sum(item['Adv Tax by Mar 15 (100%)'] for item in group1_sales)

                # Get FY year from first sale in group
                first_sale_date = pd.to_datetime(group1_sales[0]['Sale Date'])
                fy_year = get_fy_year(first_sale_date)

                advance_tax_rows.append({
                    'Sale Period': f'Apr 1 - Jul 15, {fy_year}',
                    'Financial Year': f'FY {fy_year}-{str(fy_year+1)[-2:]}',
                    'Tax Type': 'Advance Tax',
                    'Total Tax (INR)': group1_tax,
                    'By Jul 15': group1_jul,
                    'By Sep 15': group1_sep,
                    'By Dec 15': group1_dec,
                    'By Mar 15': group1_mar,
                    'Note': 'All 4 deadlines apply'
                })

            # Group 2: Sales from Jul 16 to Sep 15 - 3 deadlines (Jul passed)
            group2_sales = [item for item in capital_gains_list
                           if (pd.to_datetime(item['Sale Date']).month == 7 and pd.to_datetime(item['Sale Date']).day > 15) or
                              (pd.to_datetime(item['Sale Date']).month == 8) or
                              (pd.to_datetime(item['Sale Date']).month == 9 and pd.to_datetime(item['Sale Date']).day <= 15)]
            if group2_sales:
                group2_tax = sum(item['Tax Amount (INR)'] for item in group2_sales)
                group2_jul = sum(item['Adv Tax by Jul 15 (15%)'] for item in group2_sales)  # Will be 0
                group2_sep = sum(item['Adv Tax by Sep 15 (45%)'] for item in group2_sales)
                group2_dec = sum(item['Adv Tax by Dec 15 (75%)'] for item in group2_sales)
                group2_mar = sum(item['Adv Tax by Mar 15 (100%)'] for item in group2_sales)

                first_sale_date = pd.to_datetime(group2_sales[0]['Sale Date'])
                fy_year = get_fy_year(first_sale_date)

                advance_tax_rows.append({
                    'Sale Period': f'Jul 16 - Sep 15, {fy_year}',
                    'Financial Year': f'FY {fy_year}-{str(fy_year+1)[-2:]}',
                    'Tax Type': 'Advance Tax',
                    'Total Tax (INR)': group2_tax,
                    'By Jul 15': group2_jul,
                    'By Sep 15': group2_sep,
                    'By Dec 15': group2_dec,
                    'By Mar 15': group2_mar,
                    'Note': 'Jul 15 deadline passed'
                })

            # Group 3: Sales from Sep 16 to Dec 15 - 2 deadlines (Jul/Sep passed)
            group3_sales = [item for item in capital_gains_list
                           if (pd.to_datetime(item['Sale Date']).month == 9 and pd.to_datetime(item['Sale Date']).day > 15) or
                              (pd.to_datetime(item['Sale Date']).month == 10) or
                              (pd.to_datetime(item['Sale Date']).month == 11) or
                              (pd.to_datetime(item['Sale Date']).month == 12 and pd.to_datetime(item['Sale Date']).day <= 15)]
            if group3_sales:
                group3_tax = sum(item['Tax Amount (INR)'] for item in group3_sales)
                group3_jul = sum(item['Adv Tax by Jul 15 (15%)'] for item in group3_sales)  # Will be 0
                group3_sep = sum(item['Adv Tax by Sep 15 (45%)'] for item in group3_sales)  # Will be 0
                group3_dec = sum(item['Adv Tax by Dec 15 (75%)'] for item in group3_sales)
                group3_mar = sum(item['Adv Tax by Mar 15 (100%)'] for item in group3_sales)

                first_sale_date = pd.to_datetime(group3_sales[0]['Sale Date'])
                fy_year = get_fy_year(first_sale_date)

                advance_tax_rows.append({
                    'Sale Period': f'Sep 16 - Dec 15, {fy_year}',
                    'Financial Year': f'FY {fy_year}-{str(fy_year+1)[-2:]}',
                    'Tax Type': 'Advance Tax',
                    'Total Tax (INR)': group3_tax,
                    'By Jul 15': group3_jul,
                    'By Sep 15': group3_sep,
                    'By Dec 15': group3_dec,
                    'By Mar 15': group3_mar,
                    'Note': 'Jul/Sep deadlines passed'
                })

            # Group 4: Sales from Dec 16 to Mar 15 - Only Mar 15 deadline applies
            group4_sales = [item for item in capital_gains_list
                           if (pd.to_datetime(item['Sale Date']).month == 12 and pd.to_datetime(item['Sale Date']).day > 15) or
                              (pd.to_datetime(item['Sale Date']).month in [1, 2]) or
                              (pd.to_datetime(item['Sale Date']).month == 3 and pd.to_datetime(item['Sale Date']).day <= 15)]
            if group4_sales:
                group4_tax = sum(item['Tax Amount (INR)'] for item in group4_sales)
                group4_jul = sum(item['Adv Tax by Jul 15 (15%)'] for item in group4_sales)  # Will be 0
                group4_sep = sum(item['Adv Tax by Sep 15 (45%)'] for item in group4_sales)  # Will be 0
                group4_dec = sum(item['Adv Tax by Dec 15 (75%)'] for item in group4_sales)  # Will be 0
                group4_mar = sum(item['Adv Tax by Mar 15 (100%)'] for item in group4_sales)

                first_sale_date = pd.to_datetime(group4_sales[0]['Sale Date'])
                fy_year = get_fy_year(first_sale_date)

                # Period spans across calendar years
                if first_sale_date.month >= 4:
                    period_str = f'Dec 16, {fy_year} - Mar 15, {fy_year+1}'
                else:
                    period_str = f'Dec 16, {fy_year-1} - Mar 15, {fy_year}'

                advance_tax_rows.append({
                    'Sale Period': period_str,
                    'Financial Year': f'FY {fy_year}-{str(fy_year+1)[-2:]}',
                    'Tax Type': 'Advance Tax',
                    'Total Tax (INR)': group4_tax,
                    'By Jul 15': group4_jul,
                    'By Sep 15': group4_sep,
                    'By Dec 15': group4_dec,
                    'By Mar 15': group4_mar,
                    'Note': 'Only Mar 15 deadline applies'
                })

            # Group 5: Sales from Mar 16 to Mar 31 - Pay by Mar 31 (self-assessment tax)
            group5_sales = [item for item in capital_gains_list
                           if pd.to_datetime(item['Sale Date']).month == 3 and pd.to_datetime(item['Sale Date']).day > 15]
            if group5_sales:
                group5_tax = sum(item['Tax Amount (INR)'] for item in group5_sales)
                group5_jul = sum(item['Adv Tax by Jul 15 (15%)'] for item in group5_sales)  # Will be 0
                group5_sep = sum(item['Adv Tax by Sep 15 (45%)'] for item in group5_sales)  # Will be 0
                group5_dec = sum(item['Adv Tax by Dec 15 (75%)'] for item in group5_sales)  # Will be 0
                group5_mar = sum(item['Adv Tax by Mar 15 (100%)'] for item in group5_sales)  # Will be 0

                first_sale_date = pd.to_datetime(group5_sales[0]['Sale Date'])
                fy_year = get_fy_year(first_sale_date)

                advance_tax_rows.append({
                    'Sale Period': f'Mar 16-31, {first_sale_date.year}',
                    'Financial Year': f'FY {fy_year}-{str(fy_year+1)[-2:]}',
                    'Tax Type': 'Self-Assessment Tax',
                    'Total Tax (INR)': group5_tax,
                    'By Jul 15': group5_jul,
                    'By Sep 15': group5_sep,
                    'By Dec 15': group5_dec,
                    'By Mar 15': group5_mar,
                    'Note': 'Pay by Mar 31 (self-assessment)'
                })

            # Add TOTAL row at the end
            if advance_tax_rows:
                total_tax = int(sum(row['Total Tax (INR)'] for row in advance_tax_rows))
                total_jul = int(sum(row['By Jul 15'] for row in advance_tax_rows))
                total_sep = int(sum(row['By Sep 15'] for row in advance_tax_rows))
                total_dec = int(sum(row['By Dec 15'] for row in advance_tax_rows))
                total_mar = int(sum(row['By Mar 15'] for row in advance_tax_rows))

                advance_tax_rows.append({
                    'Sale Period': 'TOTAL',
                    'Financial Year': '',
                    'Tax Type': '',
                    'Total Tax (INR)': total_tax,
                    'By Jul 15': total_jul,
                    'By Sep 15': total_sep,
                    'By Dec 15': total_dec,
                    'By Mar 15': total_mar,
                    'Note': 'Sum across all groups'
                })

            # If no sales, return empty DataFrame (note will be in sale details section only)
            if not advance_tax_rows:
                return pd.DataFrame(columns=[
                    'Sale Period', 'Financial Year', 'Tax Type', 'Total Tax (INR)', 'By Jul 15',
                    'By Sep 15', 'By Dec 15', 'By Mar 15', 'Note'
                ])

            return pd.DataFrame(advance_tax_rows)

        # Generate advance tax schedules for BOTH regimes
        df_advance_tax_new = calculate_advance_tax_schedule(capital_gains_new_regime)
        df_advance_tax_old = calculate_advance_tax_schedule(capital_gains_old_regime)

        # For backward compatibility
        df_advance_tax = df_advance_tax_new

        # If no sales, create empty DataFrames with columns
        if df_sale_details.empty:
            df_sale_details = pd.DataFrame(columns=[
                'Nature', 'Quantity', 'Acquisition Date', 'Sale Date',
                'Rule 115(1)(f) Specified Date', 'TTBR (INR/USD)',
                'Holding Period (months)', 'Tax Type', 'Section',
                'Cost Basis (INR)', 'Sale Proceeds (INR)', 'Capital Gain (INR)',
                'Tax Rate', 'Tax Amount (INR)'
            ])

        # Create "Excluded from A3" sheet - shows sales from previous years
        # This helps track what got removed from A3 as years progress
        excluded_a3_data = []
        if 'df_sold_before_fy' in locals() and not df_sold_before_fy.empty:
            for _, row in df_sold_before_fy.iterrows():
                qty = int(row['Quantity'])
                plan_type = str(row.get('Plan Type', ''))
                nature_prefix = "ESPP" if "ESPP" in plan_type else "RSU"

                acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
                sell_date = pd.to_datetime(row['Date Sold']).strftime('%Y-%m-%d')

                excluded_a3_data.append({
                    'Nature': f"{nature_prefix} ({qty} shares)" if qty != 1 else f"{nature_prefix} ({qty} share)",
                    'Quantity': qty,
                    'Acquisition Date': acq_date,
                    'Sale Date': sell_date,
                    'Year Sold': pd.to_datetime(sell_date).year,
                    'Reason': f"Sold in {pd.to_datetime(sell_date).year} (before FY {self.calendar_year})",
                    'Total Proceeds (USD)': round(float(row['Total Proceeds']), 2),
                    'Adjusted Cost Basis Per Share (USD)': round(float(row['Adjusted Cost Basis Per Share']), 2)
                })

        df_excluded_a3 = pd.DataFrame(excluded_a3_data) if excluded_a3_data else pd.DataFrame({
            'Note': [f'No sales excluded from A3 for FY {self.calendar_year}',
                     'All sales in G&L file are either in current FY or future years']
        })

        # Round numeric values in Excluded A3 sheet
        if excluded_a3_data:
            numeric_cols_excluded = ['Initial Value (INR)', 'Peak Value (INR)', 'Sale Proceeds (INR)',
                                      'Adjusted Cost Basis Per Share (USD)']
            for col in numeric_cols_excluded:
                if col in df_excluded_a3.columns:
                    if col == 'Adjusted Cost Basis Per Share (USD)':
                        df_excluded_a3[col] = df_excluded_a3[col].round(2)  # Keep 2 decimals for USD per share
                    else:
                        df_excluded_a3[col] = df_excluded_a3[col].round(0).astype(int)  # Round INR to integers

        # If Table A3 is empty, replace with note DataFrame
        if a3_is_empty:
            df_a3 = pd.DataFrame({
                'Note': [f'No shares are vested in calendar year {self.calendar_year}']
            })

        with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
            df_a2.to_excel(writer, sheet_name="Table A2 Custodial Acc", index=False)
            df_a3.to_excel(writer, sheet_name="Table A3 Equity Interest", index=False)

            # Write Capital Gains sheet
            # If no capital gains, show a single note instead of dual regime tables
            if df_sale_details_new.columns.tolist() == ['Note']:
                # No capital gains - show single note
                df_sale_details_new.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=0, startcol=0)
            else:
                # Has capital gains - show BOTH regimes for comparison
                current_row = 0

                # Write first table to create the sheet
                df_sale_details_new.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=1, startcol=0)

                # Now access the worksheet and add regime header
                worksheet = writer.sheets["Capital Gains"]
                worksheet.cell(row=1, column=1, value="NEW TAX REGIME - Capital Gains")

                # Update current_row to continue after sale details
                current_row = 1 + len(df_sale_details_new) + 1  # header + data + spacing
                current_row += 3  # +3 blank rows spacing

                # Table 2: New Regime Advance Tax
                df_advance_tax_new.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=current_row, startcol=0)
                current_row += len(df_advance_tax_new) + 6  # +1 header, +5 spacing between regimes

                # OLD TAX REGIME Section
                # Add header row (directly at current_row+1, no extra blank row)
                worksheet.cell(row=current_row+1, column=1, value="OLD TAX REGIME - Capital Gains")
                current_row += 1  # Just move past the header (no blank row)

                # Table 3: Old Regime Sale Details
                df_sale_details_old.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=current_row, startcol=0)
                current_row += len(df_sale_details_old) + 4  # +1 header, +3 spacing

                # Table 4: Old Regime Advance Tax
                df_advance_tax_old.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=current_row, startcol=0)

                # For backward compatibility with formatting code
                start_row_table2 = len(df_sale_details_new) + 4

            # Insert Schedule OS and FSI after Capital Gains
            df_schedule_os.to_excel(writer, sheet_name="Schedule OS", index=False)
            df_schedule_fsi.to_excel(writer, sheet_name="Schedule FSI", index=False)

            df_excluded_a3.to_excel(writer, sheet_name="Excluded from A3", index=False)
            df_a2_peak.to_excel(writer, sheet_name="A2 Peak Calculation", index=False)
            df_peak_details.to_excel(writer, sheet_name="A3 Peak Value Details", index=False)
            pre_fy_sheet_name = f"Pre-{self.calendar_year} Holdings Init Val"
            df_pre_fy.to_excel(writer, sheet_name=pre_fy_sheet_name, index=False)

            # Add Dividend Reference sheets if dividends exist
            if not df_dividends.empty:
                df_dividends.to_excel(writer, sheet_name="Dividends (Schedule FA)", index=False)
            if not df_div_os.empty:
                df_div_os.to_excel(writer, sheet_name="Dividends (Schedule OS)", index=False)

            # Daily Rates sheet at the end (moved from middle)
            reference_sheet_name = f"{self.calendar_year} - Daily Rates"
            df_reference.to_excel(writer, sheet_name=reference_sheet_name, index=False)

            # Apply beautiful formatting to all sheets
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Color scheme
            header_fill = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")  # Dark teal
            header_font = Font(bold=True, color="FFFFFF", size=11)  # White bold
            alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
            total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
            total_font = Font(bold=True, size=11)
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]

                # Special formatting for Capital Gains sheet with dual regimes
                if sheet_name == "Capital Gains":
                    # Check if this is a note-only sheet (no capital gains)
                    has_capital_gains = df_sale_details_new.columns.tolist() != ['Note']

                    if has_capital_gains:
                        # Colors for Capital Gains dual-regime layout
                        regime_header_fill = PatternFill(start_color="0277BD", end_color="0277BD", fill_type="solid")  # Blue (regime headers)
                        regime_header_font = Font(bold=True, color="FFFFFF", size=12)
                        advance_tax_header_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")  # Dark gray (advance tax headers)
                        advance_tax_header_font = Font(bold=True, color="FFFFFF", size=11)
                        no_border = Border()  # Empty border (no lines)

                        # Row 1: NEW TAX REGIME header (blue, centered, NO borders)
                        for cell in ws[1]:
                            cell.fill = regime_header_fill
                            cell.font = regime_header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = no_border

                        # Row 2: Sale details header (teal)
                        for cell in ws[2]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = border_thin

                        # Calculate row numbers (they shifted after removing blank rows)
                        # Old: row 11, New: row 2 + len(sale_details) + 1 = approximately row 10
                        advance_tax_row_new = 2 + len(df_sale_details_new) + 1 + 3  # +3 for spacing
                        regime2_header_row = advance_tax_row_new + len(df_advance_tax_new) + 1 + 5  # +5 for spacing
                        sale_details_row_old = regime2_header_row + 1
                        advance_tax_row_old = sale_details_row_old + len(df_sale_details_old) + 1 + 3

                        # Advance tax header NEW - dark gray ONLY for columns with content (A-I), J-N no color/border
                        for cell in ws[advance_tax_row_new]:
                            if cell.value is not None and str(cell.value).strip() != '':
                                cell.fill = advance_tax_header_fill
                                cell.font = advance_tax_header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                cell.border = border_thin
                            else:
                                # Empty columns - no color, no border
                                cell.border = no_border

                        # OLD TAX REGIME header (blue, centered, NO borders)
                        for cell in ws[regime2_header_row]:
                            cell.fill = regime_header_fill
                            cell.font = regime_header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = no_border

                        # Sale details header OLD (teal)
                        for cell in ws[sale_details_row_old]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = border_thin

                        # Advance tax header OLD - dark gray ONLY for columns with content (A-I), J-N no color/border
                        for cell in ws[advance_tax_row_old]:
                            if cell.value is not None and str(cell.value).strip() != '':
                                cell.fill = advance_tax_header_fill
                                cell.font = advance_tax_header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                cell.border = border_thin
                            else:
                                # Empty columns - no color, no border
                                cell.border = no_border

                        # Format all data rows
                        header_rows = {1, 2, advance_tax_row_new, regime2_header_row, sale_details_row_old, advance_tax_row_old}

                        # Calculate row ranges for different sections
                        sale_details_rows_new = set(range(3, 2 + len(df_sale_details_new) + 1))  # Rows after header 2
                        advance_tax_data_rows_new = set(range(advance_tax_row_new + 1, advance_tax_row_new + len(df_advance_tax_new) + 1))
                        sale_details_rows_old = set(range(sale_details_row_old + 1, sale_details_row_old + len(df_sale_details_old) + 1))
                        advance_tax_data_rows_old = set(range(advance_tax_row_old + 1, advance_tax_row_old + len(df_advance_tax_old) + 1))

                        all_advance_tax_data_rows = advance_tax_data_rows_new | advance_tax_data_rows_old
                        all_sale_details_rows = sale_details_rows_new | sale_details_rows_old

                        # Professional color scheme for sale details
                        light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Very light blue
                        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                            is_blank_row = all(cell.value is None or str(cell.value).strip() == '' for cell in row)

                            # Skip formatting for headers (already done above)
                            if row_idx in header_rows:
                                continue

                            # Determine if this is an odd or even sale details row (for alternating colors)
                            if row_idx in sale_details_rows_new:
                                row_position = row_idx - 3  # Position within sale details (0-indexed)
                            elif row_idx in sale_details_rows_old:
                                row_position = row_idx - (sale_details_row_old + 1)
                            else:
                                row_position = None

                            # Apply borders and center alignment
                            for cell in row:
                                if is_blank_row:
                                    # Blank separator rows - no borders
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                elif row_idx in all_advance_tax_data_rows:
                                    # Advance tax data rows - border only for cells with content
                                    if cell.value is not None and str(cell.value).strip() != '':
                                        cell.border = border_thin
                                    else:
                                        cell.border = no_border
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                elif row_idx in all_sale_details_rows:
                                    # Sale details rows - alternating colors (light blue / white)
                                    if row_position is not None and row_position % 2 == 0:
                                        cell.fill = white_fill
                                    else:
                                        cell.fill = light_blue_fill
                                    cell.border = border_thin
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                else:
                                    # Other regular data rows
                                    cell.border = border_thin
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # No capital gains - format the Note sheet like Table A3
                        # Row 1: Note header (teal)
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = border_thin

                        # Row 2+: Note content (white background)
                        for row_idx in range(2, ws.max_row + 1):
                            for cell in ws[row_idx]:
                                if cell.value is not None:
                                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                                    cell.border = border_thin
                else:
                    # Standard formatting for other sheets
                    # Format header row
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border_thin

                    # Format data rows
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                        # Check if this is a blank separator row (all cells are None or empty)
                        is_blank_row = all(cell.value is None or str(cell.value).strip() == '' for cell in row)

                        # Alternating row colors
                        if row_idx % 2 == 0:
                            for cell in row:
                                if cell.value is not None:
                                    cell.fill = alt_row_fill

                        # Apply borders and alignment (skip borders for blank separator rows)
                        for cell in row:
                            if not is_blank_row:
                                cell.border = border_thin
                            cell.alignment = Alignment(vertical='center')

                # Format specific columns based on content
                for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                    header_value = col[0].value
                    col_letter = get_column_letter(col_idx)

                    # TTBR and Rate columns - show 2 decimal places (CHECK FIRST - these contain "USD" but aren't USD values!)
                    if header_value and any(keyword in str(header_value) for keyword in
                                             ['TTBR', 'Rate']):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'  # Show 2 decimal places

                    # USD columns - show 2 decimal places with $ symbol (CHECK AFTER TTBR to avoid conflicts)
                    elif header_value and '(USD)' in str(header_value):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '$#,##0.00'  # Show 2 decimal places

                    # INR Currency columns - different format for A2/A3 vs other sheets
                    elif header_value and any(keyword in str(header_value) for keyword in
                                           ['INR', 'Value', 'Balance', 'Amount', 'Proceeds', 'Cost', 'Gain', 'Tax', 'Invstmnt']):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and isinstance(cell.value, (int, float)):
                                # A2, A3, and Capital Gains sheets: integers only (ITR portal requirement)
                                # Other sheets: show decimals for accuracy
                                if sheet_name in ["Table A2 Custodial Acc", "Table A3 Equity Interest", "Excluded from A3", "Capital Gains"]:
                                    cell.number_format = '"Rs."#,##0'  # No decimals for ITR sheets
                                else:
                                    cell.number_format = '"Rs."#,##0.00'  # Show 2 decimal places for reference sheets

                    # Date columns
                    elif header_value and 'Date' in str(header_value):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and not isinstance(cell.value, str):
                                cell.number_format = 'YYYY-MM-DD'

                    # Text columns (ZipCode, AccountNumber)
                    elif header_value in ['ZipCode', 'AccountNumber']:
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value is not None:
                                cell.value = str(cell.value)
                                cell.number_format = '@'

                # Special formatting for Peak Info summary tables in Reference and A2 Peak sheets
                # This MUST run AFTER column formatting to override the "Value" column default
                if sheet_name in ["Reference - Daily Rates", "A2 Peak Calculation"]:
                    # Find where Peak Info columns start (look for "Peak Info" or "PEAK SUMMARY" header)
                    peak_info_col = None
                    value_col = None
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value in ['Peak Info', 'PEAK SUMMARY']:
                            peak_info_col = col_idx
                        elif cell.value == 'Value' and peak_info_col:
                            value_col = col_idx
                            break

                    if peak_info_col and value_col:
                        # Override formatting row by row based on label
                        for row_idx in range(2, ws.max_row + 1):
                            label_cell = ws.cell(row=row_idx, column=peak_info_col)
                            value_cell = ws.cell(row=row_idx, column=value_col)

                            if label_cell.value and value_cell.value and isinstance(value_cell.value, (int, float)):
                                label = str(label_cell.value)

                                # Apply specific formatting based on label content
                                if '(USD)' in label:
                                    # Stock Price (USD) or Account Value (USD)
                                    value_cell.number_format = '$#,##0.00'
                                elif label == 'TTBR':
                                    # Plain TTBR rate
                                    value_cell.number_format = '#,##0.00'
                                elif 'INR' in label or 'Peak Balance' in label:
                                    # Account Value (INR), Peak INR Value, A2 Peak Balance (INR)
                                    value_cell.number_format = '"Rs."#,##0.00'

                # Special formatting for Schedule OS and FSI sheets
                if sheet_name in ["Schedule OS", "Schedule FSI"]:
                    # These sheets have a different layout - first column is labels, second column is data
                    # First row should have header formatting
                    for cell in ws[1]:
                        if cell.value:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = border_thin

                    # Format label column (Column A) - bold
                    label_font = Font(bold=True, size=10)
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=1)  # Column A
                        if cell.value:
                            cell.font = label_font
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    # Format data column (Column B) - numbers with proper formatting
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=2)  # Column B (indian_fy column)
                        if cell.value and isinstance(cell.value, (int, float)):
                            # Check the label in column A to determine format
                            label_cell = ws.cell(row=row_idx, column=1)
                            label = str(label_cell.value) if label_cell.value else ""

                            if 'USD' in label:
                                cell.number_format = '$#,##0.00'
                            elif 'INR' in label or 'Income' in label or 'Tax' in label or 'Relief' in label:
                                cell.number_format = '"Rs."#,##0'
                            else:
                                cell.number_format = '#,##0'

                    # For Schedule FSI, format the country details table (rows with headers in row with "Country")
                    if sheet_name == "Schedule FSI":
                        # Find the row with country details header
                        for row_idx in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row_idx, column=1)
                            if cell.value == "Country":
                                # This is the header row for country details table
                                for col_idx in range(1, ws.max_column + 1):
                                    header_cell = ws.cell(row=row_idx, column=col_idx)
                                    if header_cell.value:
                                        header_cell.fill = header_fill
                                        header_cell.font = header_font
                                        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                        header_cell.border = border_thin
                                break

                    # Format warnings/notes section (usually at the bottom)
                    warning_font = Font(italic=True, size=9, color="FF0000")
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=1)
                        if cell.value and 'WARNING' in str(cell.value).upper():
                            cell.font = Font(bold=True, size=10)
                        # Format warning text rows
                        cell = ws.cell(row=row_idx, column=2)
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('•'):
                            cell.font = warning_font
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    # Auto-adjust column widths for Schedule OS and FSI
                    ws.column_dimensions['A'].width = 45  # Label column
                    ws.column_dimensions['B'].width = 25  # Data column
                    for col_idx in range(3, ws.max_column + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 18

                # Special formatting for Capital Gains sheet (only when has capital gains)
                if sheet_name == "Capital Gains" and df_sale_details_new.columns.tolist() != ['Note']:
                    # Table 2 header and data (bold + colored)
                    table2_header_row = start_row_table2 + 1
                    table2_first_data_row = start_row_table2 + 2
                    table2_last_data_row = start_row_table2 + 1 + len(df_advance_tax)

                    # Format Table 2 header
                    for cell in ws[table2_header_row]:
                        if cell.value:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Format ALL Table 2 data rows (multiple rows now, not just one)
                    for row_num in range(table2_first_data_row, table2_last_data_row + 1):
                        is_total_row = ws.cell(row_num, 1).value == 'TOTAL'

                        for cell in ws[row_num]:
                            if cell.value is not None:
                                # TOTAL row: bold + colored background
                                if is_total_row:
                                    cell.fill = total_fill
                                    cell.font = total_font

                                cell.alignment = Alignment(horizontal='center', vertical='center')

                                # Force number format for numeric columns (prevents Excel date auto-conversion)
                                if isinstance(cell.value, (int, float)) and cell.value != 0:
                                    cell.number_format = '"Rs."#,##0'  # No decimals for ITR format
                                elif isinstance(cell.value, (int, float)) and cell.value == 0:
                                    cell.number_format = '"Rs."#,##0'  # Show 0 as Rs.0

                # Freeze first row (except Capital Gains sheet)
                if sheet_name != 'Capital Gains':
                    ws.freeze_panes = ws['A2']

            # Auto-adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # Special handling for Capital Gains note-only sheet
                if sheet_name == "Capital Gains" and df_sale_details_new.columns.tolist() == ['Note']:
                    # Set a wide column for the note text to wrap properly
                    worksheet.column_dimensions['A'].width = 80
                    continue

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    header_cell = column[0]

                    # Calculate width based on header and data
                    for cell in column:
                        try:
                            if cell.value is not None:
                                # For formatted numbers, account for the formatting
                                if isinstance(cell.value, (int, float)):
                                    # Currency with symbol (Rs.1,234.56 or $1,234.56)
                                    if cell.number_format and ('Rs.' in cell.number_format or '$' in cell.number_format):
                                        # Account for currency symbol, commas, and decimals
                                        formatted_length = len(f"{cell.value:,.2f}") + 2  # +2 for symbol and space
                                    # Plain numbers with decimals
                                    elif cell.number_format and '#,##0' in cell.number_format:
                                        formatted_length = len(f"{cell.value:,.2f}")
                                    else:
                                        formatted_length = len(str(cell.value))
                                    cell_length = formatted_length
                                else:
                                    cell_length = len(str(cell.value))

                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                    # Ensure minimum width for header text
                    if header_cell.value:
                        header_length = len(str(header_cell.value))
                        max_length = max(max_length, header_length)

                    # Set column width with padding, minimum 10, maximum 50
                    adjusted_width = min(max(max_length + 3, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        # Generate CSV files with custom headers matching ITR format
        csv_a2_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}_table_a2.csv")
        csv_a3_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}_table_a3.csv")

        # Table A2 CSV with ITR-compliant headers and column order
        # First, round all numeric values in the source data to integers
        for acc in output_data["ScheduleFA"]["DtlsForeignCustodialAcc"]:
            acc["PeakBalanceDuringPeriod"] = round(acc["PeakBalanceDuringPeriod"])
            acc["ClosingBalance"] = round(acc["ClosingBalance"])
            acc["GrossAmtPaidCredited"] = round(acc["GrossAmtPaidCredited"])

        df_a2_csv = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignCustodialAcc"])
        df_a2_csv = df_a2_csv.rename(columns={
            "CountryName": "Country/Region name",
            "CountryCodeExcludingIndia": "Country Name and Code",
            "FinancialInstName": "Name of financial institution",
            "FinancialInstAddress": "Address of financial institution",
            "ZipCode": "ZIP Code",
            "AccountNumber": "Account Number",
            "Status": "Status",
            "AccOpenDate": "Account opening date",
            "PeakBalanceDuringPeriod": "Peak Balance During the Period",
            "ClosingBalance": "Closing balance",
            "NatureOfAmount": "Nature of Amount",
            "GrossAmtPaidCredited": "Amount"
        })
        # Format dates to DD/MM/YYYY (ITR portal requirement)
        if "Account opening date" in df_a2_csv.columns:
            df_a2_csv["Account opening date"] = pd.to_datetime(df_a2_csv["Account opening date"], errors='coerce').dt.strftime('%d/%m/%Y')

        # Round numeric columns to integers (ITR portal requirement)
        df_a2_csv["Peak Balance During the Period"] = df_a2_csv["Peak Balance During the Period"].round(0).astype(int)
        df_a2_csv["Closing balance"] = df_a2_csv["Closing balance"].round(0).astype(int)
        df_a2_csv["Amount"] = df_a2_csv["Amount"].round(0).astype(int)

        # Reorder columns to match ITR format exactly
        df_a2_csv = df_a2_csv[[
            "Country/Region name", "Country Name and Code", "Name of financial institution",
            "Address of financial institution", "ZIP Code", "Account Number", "Status",
            "Account opening date", "Peak Balance During the Period", "Closing balance",
            "Nature of Amount", "Amount"
        ]]
        # Write A2 CSV (ITR portal format requirement)
        # Try simple format without quotes or trailing commas
        df_a2_csv.to_csv(csv_a2_filename, index=False, quoting=0)  # quoting=0 = QUOTE_MINIMAL

        # Table A3 CSV with ITR-compliant headers
        # First, round all numeric values in the source data to integers
        for holding in output_data["ScheduleFA"]["DtlsForeignEquityDebtInterest"]:
            holding["InitialValOfInvstmnt"] = round(holding["InitialValOfInvstmnt"])
            holding["PeakBalanceDuringPeriod"] = round(holding["PeakBalanceDuringPeriod"])
            holding["ClosingBalance"] = round(holding["ClosingBalance"])
            holding["TotGrossAmtPaidCredited"] = round(holding["TotGrossAmtPaidCredited"])
            holding["TotGrossProceeds"] = round(holding["TotGrossProceeds"])

        df_a3_csv = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignEquityDebtInterest"])
        df_a3_csv = df_a3_csv.rename(columns={
            "CountryName": "Country/Region name",
            "CountryCodeExcludingIndia": "Country Name and Code",
            "NameOfEntity": "Name of entity",
            "AddressOfEntity": "Address of entity",
            "ZipCode": "ZIP Code",
            "NatureOfEntity": "Nature of entity",
            "InterestAcquiringDate": "Date of acquiring the interest",
            "InitialValOfInvstmnt": "Initial value of the investment",
            "PeakBalanceDuringPeriod": "Peak value of investment during the Period",
            "ClosingBalance": "Closing balance",
            "TotGrossAmtPaidCredited": "Total gross amount paid/credited with respect to the holding during the period",
            "TotGrossProceeds": "Total gross proceeds from sale or redemption of investment during the period"
        })

        # Keep dates in YYYY-MM-DD ISO format (ITR portal requirement)
        # This avoids DD/MM vs MM/DD ambiguity
        if "Date of acquiring the interest" in df_a3_csv.columns:
            df_a3_csv["Date of acquiring the interest"] = pd.to_datetime(df_a3_csv["Date of acquiring the interest"], errors='coerce').dt.strftime('%Y-%m-%d')

        # Force text columns to be treated as text (prevent Excel "leading zeros" warning)
        # Add ="value" format for text fields
        text_cols_a3 = ["Country Name and Code", "ZIP Code"]
        for col in text_cols_a3:
            if col in df_a3_csv.columns:
                df_a3_csv[col] = df_a3_csv[col].astype(str)

        # Round numeric columns to integers (ITR portal requirement) - only if columns exist
        if "Initial value of the investment" in df_a3_csv.columns:
            df_a3_csv["Initial value of the investment"] = df_a3_csv["Initial value of the investment"].round(0).astype(int)
            df_a3_csv["Peak value of investment during the Period"] = df_a3_csv["Peak value of investment during the Period"].round(0).astype(int)
            df_a3_csv["Closing balance"] = df_a3_csv["Closing balance"].round(0).astype(int)
            df_a3_csv["Total gross amount paid/credited with respect to the holding during the period"] = df_a3_csv["Total gross amount paid/credited with respect to the holding during the period"].round(0).astype(int)
            df_a3_csv["Total gross proceeds from sale or redemption of investment during the period"] = df_a3_csv["Total gross proceeds from sale or redemption of investment during the period"].round(0).astype(int)

        # Write A3 CSV (ITR portal format requirement)
        # Try simple format without quotes or trailing commas
        df_a3_csv.to_csv(csv_a3_filename, index=False, quoting=0)  # quoting=0 = QUOTE_MINIMAL

        # Count total sheets
        base_sheets = 10  # A2, A3, OS, FSI, Excluded, CG, Reference, A2 Peak, A3 Peak Details, Pre-FY
        dividend_sheets = 0
        if not df_dividends.empty:
            dividend_sheets = 2  # Dividends (FA) and Dividends (OS)
        total_sheets = base_sheets + dividend_sheets

        print(f"\n[SUCCESS] Finished processing calendar year {self.calendar_year}!")
        print(f"    - JSON Output:  {json_filename}")
        print(f"    - Excel Output: {excel_filename} ({total_sheets} sheets)")
        print(f"        - Table A2 Custodial Acc")
        print(f"        - Table A3 Equity Interest")
        print(f"        - Capital Gains (Current + Future sales)")
        print(f"        - Schedule OS (Other Sources - Dividend Income)")
        print(f"        - Schedule FSI (Foreign Source Income)")
        print(f"        - Excluded from A3 (Sales from previous years)")
        print(f"        - A2 Peak Calculation (Daily account values)")
        print(f"        - A3 Peak Value Details (Peak date and value breakdown per lot)")
        print(f"        - Pre-{self.calendar_year} Holdings Init Val")
        if not df_dividends.empty:
            print(f"        - Dividends (Schedule FA) - {len(df_dividends)} payments, exact date TTBR")
            print(f"        - Dividends (Schedule OS) - {len(df_div_os)} payments, Rule 115(1)(e) TTBR")
        print(f"        - {self.calendar_year} - Daily Rates (Stock prices + SBI TTBR)")
        print(f"    - CSV A2:       {csv_a2_filename}")
        print(f"    - CSV A3:       {csv_a3_filename}")
        print(f"    - Total Equity Tranches: {len(equity_tranches)}")
        print(f"    - Open Holdings: {len([t for t in equity_tranches if t['ClosingBalance'] > 0])}")
        print(f"    - Sold Holdings: {len([t for t in equity_tranches if 'Sold' in t['NatureOfEntity']])}")
        if 'df_sold_before_fy' in locals():
            print(f"    - Excluded from A3: {len(df_sold_before_fy)} sales from previous years\n")
        else:
            print()
        return output_data

# =====================================================================
# EXECUTION ENTRY POINT
# =====================================================================
if __name__ == "__main__":

    # CONFIGURATION
    # Load configuration from config.json
    config = load_config()

    # Get target year from config (required field, validated by BAT file)
    # Fallback: auto-detect as (current year - 1) for ITR filing
    from datetime import datetime
    TARGET_YEAR = config.get("target_year", datetime.now().year - 1)
    # Get account number from custodial_account
    custodial_acc = config.get("custodial_account", {})
    ACCOUNT_NUMBER = custodial_acc.get("account_number", "")

    # Account number can be extracted from ClientStatement PDF, so it's optional in config
    if not ACCOUNT_NUMBER or ACCOUNT_NUMBER == "ENTER_YOUR_ETRADE_ACCOUNT_NUMBER":
        print("[i] Account number not in config.json - will extract from ClientStatement PDF")
        ACCOUNT_NUMBER = ""  # Will be extracted from ClientStatement

    # Input file paths - auto-detect from inputs folder
    BYSTATUS_FILE = "etrade_inputs/ByStatus_expanded.xlsx"
    GL_FILE = "etrade_inputs/G&L_Expanded.xlsx"
    TRANSACTION_HISTORY_FILE = "etrade_inputs/Transaction_History.csv"

    # Fallback to root if inputs folder doesn't have them
    if not os.path.exists(BYSTATUS_FILE):
        BYSTATUS_FILE = "ByStatus_expanded.xlsx"
    if not os.path.exists(GL_FILE):
        GL_FILE = "G&L_Expanded.xlsx"
    if not os.path.exists(TRANSACTION_HISTORY_FILE):
        TRANSACTION_HISTORY_FILE = "Transaction_History.csv"

    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Schedule FA Generator')
    parser.add_argument('--income-bracket', type=str, help='Income bracket choice (1-11)')
    args = parser.parse_args()

    print("\n[*] Starting Schedule FA generation (WEB SCRAPING MODE)...")
    print(f"[*] This will open Chrome browser in background")
    print(f"[*] Looking for E*TRADE files: {BYSTATUS_FILE}, {GL_FILE}")
    print()

    # Get income bracket from command-line argument
    income_bracket = args.income_bracket.strip() if args.income_bracket else None

    # If empty, no G&L file exists, use defaults (won't be used anyway)
    if not income_bracket:
        print("[i] No capital gains - using default STCG rates (will not be applied)")
        stcg_rate_new = 0.312
        stcg_rate_old = 0.312
        print(f"[OK] Default rates set (not used - no sales)")
        print()
    else:
        # Calculate rates for BOTH regimes based on income bracket
        stcg_rate_new, stcg_rate_old, new_display, old_display = calculate_stcg_rates_for_income(income_bracket)
        print(f"[OK] New Tax Regime STCG: {new_display}")
        print(f"[OK] Old Tax Regime STCG: {old_display}")
        print(f"[i] Capital Gains sheet will show BOTH regimes for comparison")
        print()

    # For backward compatibility, use new regime rate as default
    # (This variable is used when initializing the app - represents New Regime for now)
    stcg_rate = stcg_rate_new

    try:
        # Read E*TRADE files to discover company symbols
        print("[*] Reading E*TRADE export files to discover companies...")

        # ByStatus file is optional - only read if it exists
        df_open = pd.DataFrame()
        if os.path.exists(BYSTATUS_FILE):
            df_open = pd.read_excel(BYSTATUS_FILE)
        else:
            print("[!] WARNING: ByStatus_expanded.xlsx not found")
            print("[!] Table A3 will not include current holdings")

        # G&L file is optional - only read if it exists
        df_sold = pd.DataFrame()
        if os.path.exists(GL_FILE):
            df_sold = pd.read_excel(GL_FILE)
        else:
            print("[!] WARNING: G&L_Expanded.xlsx not found")
            print("[!] Table A3 will not include sold shares")
            print("[!] Capital Gains will be empty")

        # Auto-discover unique symbols from input files
        symbols = set()
        if not df_open.empty and 'Symbol' in df_open.columns:
            symbols.update(df_open['Symbol'].dropna().unique())
        if not df_sold.empty and 'Symbol' in df_sold.columns:
            symbols.update(df_sold['Symbol'].dropna().unique())

        # If no symbols found from files, use companies from config.json
        if not symbols:
            print("[i] No ByStatus or G&L files found - will use companies from config.json")
            config_companies = config.get("table_a3_companies", {})
            symbols = set(k for k in config_companies.keys() if not k.startswith('_'))
            if not symbols:
                print("")
                print("=" * 70)
                print("[ERROR] No company symbols found!")
                print("Need either:")
                print("  - etrade_inputs/ByStatus_expanded.xlsx (for holdings)")
                print("  - etrade_inputs/G&L_Expanded.xlsx (for sales)")
                print("  - OR companies configured in config.json")
                print("=" * 70)
                raise FileNotFoundError("No company symbols available for processing.")

        print(f"[OK] Discovered {len(symbols)} unique symbols: {', '.join(sorted(str(s).strip() for s in symbols))}")

        # Initialize app to access scraping methods
        app = ScheduleFAApp(calendar_year=TARGET_YEAR, stcg_rate_new=stcg_rate_new, stcg_rate_old=stcg_rate_old)

        # Scrape company info for each discovered symbol not in config
        config_companies = config.get("table_a3_companies", {})
        # Filter out entries that start with underscore (comments/notes)
        config_companies = {k: v for k, v in config_companies.items() if not k.startswith('_')}

        for symbol in sorted(symbols):
            symbol = str(symbol).strip()
            if symbol and symbol not in config_companies:
                print(f"\n[+] Auto-discovering company info for: {symbol}")
                company_info = app._scrape_company_profile(symbol)
                config_companies[symbol] = company_info
                time.sleep(2)  # Be respectful to Yahoo Finance
            elif symbol:
                print(f"[>] Using existing config for: {symbol}")

        # Update config with discovered companies
        config["table_a3_companies"] = config_companies

        # Save updated config back to file for next run
        save_config(config)
        print("[OK] Saved company info to config.json for next run")
        print()

        # Now process E*TRADE exports with updated config
        # Pass None for GL_FILE if it doesn't exist
        gl_file_to_use = GL_FILE if os.path.exists(GL_FILE) else None
        app.process_etrade_exports(
            bystatus_path=BYSTATUS_FILE,
            gl_path=gl_file_to_use,
            transaction_history_path=TRANSACTION_HISTORY_FILE,
            account_no=ACCOUNT_NUMBER,
            config=config
        )

        print("[*] Process complete! Check the generated files:")
        print(f"    - schedule_fa_{app.indian_fy}.json")
        print(f"    - schedule_fa_{app.indian_fy}.xlsx")
        print(f"    - schedule_fa_{app.indian_fy}_table_a2.csv")
        print(f"    - schedule_fa_{app.indian_fy}_table_a3.csv")
        print("\n[*] You can now upload the CSV/JSON to the ITR e-filing portal.")
        print("[*] Review the Excel file to verify all values before filing.\n")

    except Exception as e:
        print(f"\n[ERROR] {str(e)}\n")
