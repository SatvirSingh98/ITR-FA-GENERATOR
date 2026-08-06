"""
ITR-FA-GENERATOR - Schedule FA Generator for ITR2/ITR3
Copyright (c) 2024-2026 Satvinder Singh
Licensed under GNU General Public License

ExcelFormatter
Handles professional Excel formatting for all output sheets
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Apply professional formatting to Excel workbooks"""

    def __init__(self):
        # Define color scheme
        self.header_fill = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")  # Dark teal
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
        self.total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
        self.total_font = Font(bold=True, size=11)

        # Borders
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.no_border = Border()

        # Capital Gains specific colors
        self.regime_header_fill = PatternFill(start_color="0277BD", end_color="0277BD", fill_type="solid")  # Blue
        self.regime_header_font = Font(bold=True, color="FFFFFF", size=12)
        self.sale_header_fill = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")  # Teal
        self.sale_header_font = Font(bold=True, color="FFFFFF", size=11)
        self.advance_tax_header_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")  # Dark gray
        self.advance_tax_header_font = Font(bold=True, color="FFFFFF", size=11)
        self.light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    def format_capital_gains_sheet(self, ws, df_sale_details_new, df_advance_tax_new, df_sale_details_old, df_advance_tax_old):
        """
        Apply dual-regime formatting to Capital Gains sheet

        Args:
            ws: Worksheet object
            df_sale_details_new: New regime sale details DataFrame
            df_advance_tax_new: New regime advance tax DataFrame
            df_sale_details_old: Old regime sale details DataFrame
            df_advance_tax_old: Old regime advance tax DataFrame
        """
        # Row 1: NEW TAX REGIME header (blue, centered, NO borders)
        for cell in ws[1]:
            cell.fill = self.regime_header_fill
            cell.font = self.regime_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.no_border

        # Row 2: Sale details header (teal)
        for cell in ws[2]:
            cell.fill = self.sale_header_fill
            cell.font = self.sale_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border_thin

        # Calculate row positions
        advance_tax_row_new = 2 + len(df_sale_details_new) + 1 + 3
        regime2_header_row = advance_tax_row_new + len(df_advance_tax_new) + 1 + 5
        sale_details_row_old = regime2_header_row + 1
        advance_tax_row_old = sale_details_row_old + len(df_sale_details_old) + 1 + 3

        # Advance tax header NEW - selective formatting
        for cell in ws[advance_tax_row_new]:
            if cell.value is not None and str(cell.value).strip() != '':
                cell.fill = self.advance_tax_header_fill
                cell.font = self.advance_tax_header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.border_thin
            else:
                cell.border = self.no_border

        # OLD TAX REGIME header
        for cell in ws[regime2_header_row]:
            cell.fill = self.regime_header_fill
            cell.font = self.regime_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.no_border

        # Sale details header OLD
        for cell in ws[sale_details_row_old]:
            cell.fill = self.sale_header_fill
            cell.font = self.sale_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border_thin

        # Advance tax header OLD - selective formatting
        for cell in ws[advance_tax_row_old]:
            if cell.value is not None and str(cell.value).strip() != '':
                cell.fill = self.advance_tax_header_fill
                cell.font = self.advance_tax_header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.border_thin
            else:
                cell.border = self.no_border

        # Define row sets
        header_rows = {1, 2, advance_tax_row_new, regime2_header_row, sale_details_row_old, advance_tax_row_old}
        sale_details_rows_new = set(range(3, 2 + len(df_sale_details_new) + 1))
        advance_tax_data_rows_new = set(range(advance_tax_row_new + 1, advance_tax_row_new + len(df_advance_tax_new) + 1))
        sale_details_rows_old = set(range(sale_details_row_old + 1, sale_details_row_old + len(df_sale_details_old) + 1))
        advance_tax_data_rows_old = set(range(advance_tax_row_old + 1, advance_tax_row_old + len(df_advance_tax_old) + 1))

        all_advance_tax_data_rows = advance_tax_data_rows_new | advance_tax_data_rows_old
        all_sale_details_rows = sale_details_rows_new | sale_details_rows_old

        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            is_blank_row = all(cell.value is None or str(cell.value).strip() == '' for cell in row)

            if row_idx in header_rows:
                continue

            # Determine row position for alternating colors
            if row_idx in sale_details_rows_new:
                row_position = row_idx - 3
            elif row_idx in sale_details_rows_old:
                row_position = row_idx - (sale_details_row_old + 1)
            else:
                row_position = None

            # Apply formatting to each cell
            for cell in row:
                if is_blank_row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx in all_advance_tax_data_rows:
                    if cell.value is not None and str(cell.value).strip() != '':
                        cell.border = self.border_thin
                    else:
                        cell.border = self.no_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx in all_sale_details_rows:
                    if row_position is not None and row_position % 2 == 0:
                        cell.fill = self.white_fill
                    else:
                        cell.fill = self.light_blue_fill
                    cell.border = self.border_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.border = self.border_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    def format_standard_sheet(self, ws):
        """
        Apply standard formatting to non-Capital Gains sheets

        Args:
            ws: Worksheet object
        """
        # Format header row
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border_thin

        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            is_blank_row = all(cell.value is None or str(cell.value).strip() == '' for cell in row)

            # Alternating row colors
            if row_idx % 2 == 0:
                for cell in row:
                    if cell.value is not None:
                        cell.fill = self.alt_row_fill

            # Apply borders and alignment
            for cell in row:
                if not is_blank_row:
                    cell.border = self.border_thin
                cell.alignment = Alignment(vertical='center')

    def auto_size_columns(self, ws):
        """
        Auto-size columns based on content

        Args:
            ws: Worksheet object
        """
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            header_value = col[0].value
            col_letter = get_column_letter(col_idx)

            # Set column width based on header content
            if header_value:
                header_str = str(header_value)
                if 'Date' in header_str or 'Period' in header_str:
                    ws.column_dimensions[col_letter].width = 18
                elif 'Amount' in header_str or 'Value' in header_str or 'Tax' in header_str:
                    ws.column_dimensions[col_letter].width = 16
                elif 'Rate' in header_str or 'TTBR' in header_str:
                    ws.column_dimensions[col_letter].width = 14
                elif 'Nature' in header_str or 'Type' in header_str or 'Section' in header_str:
                    ws.column_dimensions[col_letter].width = 20
                elif 'Note' in header_str:
                    ws.column_dimensions[col_letter].width = 35
                else:
                    ws.column_dimensions[col_letter].width = 15

    def format_workbook(self, writer, sheet_configs):
        """
        Format entire workbook with all sheets

        Args:
            writer: ExcelWriter object
            sheet_configs (dict): Configuration for special sheets
                Example: {
                    'Capital Gains': {
                        'type': 'dual_regime',
                        'df_sale_details_new': df1,
                        'df_advance_tax_new': df2,
                        'df_sale_details_old': df3,
                        'df_advance_tax_old': df4
                    }
                }
        """
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Apply special formatting for Capital Gains sheet
            if sheet_name == "Capital Gains" and sheet_name in sheet_configs:
                config = sheet_configs[sheet_name]
                self.format_capital_gains_sheet(
                    ws,
                    config['df_sale_details_new'],
                    config['df_advance_tax_new'],
                    config['df_sale_details_old'],
                    config['df_advance_tax_old']
                )
            else:
                # Standard formatting
                self.format_standard_sheet(ws)

            # Auto-size columns for all sheets
            self.auto_size_columns(ws)
