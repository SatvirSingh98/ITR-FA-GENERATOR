"""
Schedule FA Generator - Web Scraping Version
Uses Selenium to scrape Yahoo Finance and download SBI rates from GitHub

Advantages:
- Fully automated (no manual downloads)
- Works on corporate networks (uses browser)
- Gets accurate data from source websites
- No API dependencies

Requirements:
- Chrome browser installed
- Internet connection
"""

import json
import os
import pandas as pd
from datetime import datetime
import time
import warnings
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import requests

warnings.filterwarnings('ignore')

def save_config(config, config_file="config.json"):
    """Save configuration back to config.json file."""
    try:
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"[!] Error saving config: {e}")
        return False

def load_config(config_file="config.json"):
    """Load configuration from config.json file."""
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
        print(f"[OK] Loaded configuration from {config_file}")
        # Handle backward compatibility
        if "table_a2_custodial_accounts" in config:
            # Old multi-account format - take first account
            config["custodial_account"] = config["table_a2_custodial_accounts"][0]
        elif "table_a2_custodial_account" in config:
            # Very old single account format
            config["custodial_account"] = config["table_a2_custodial_account"]
        return config
    except FileNotFoundError:
        print(f"[!] ERROR: {config_file} not found!")
        print(f"[!] Please create config.json with your account details.")
        print(f"[!] See config.example.json for reference.")
        raise FileNotFoundError(f"Config file '{config_file}' is required. Copy config.example.json to config.json and edit your details.")
    except json.JSONDecodeError as e:
        print(f"[!] Error parsing {config_file}: {e}")
        print(f"[!] Please check your JSON syntax and try again.")
        raise

class ScheduleFAApp:
    @staticmethod
    def clean_text_for_itr(text):
        """Remove commas, periods, asterisks, and other special characters that ITR portal doesn't accept."""
        if not text or not isinstance(text, str):
            return text
        # Remove commas, periods, and asterisks
        text = text.replace(',', '').replace('.', '').replace('*', '')
        # Remove other problematic characters if needed
        return text.strip()

    def __init__(self, calendar_year=None):
        """
        :param calendar_year: e.g. 2024, 2025, 2026. If None, defaults to the previous calendar year.
        """
        current_year = datetime.now().year
        self.calendar_year = calendar_year if calendar_year else (current_year - 1)

        self.start_date = f"{self.calendar_year}-01-01"
        self.end_date = f"{self.calendar_year}-12-31"

        # Dynamic Tax Schema Years
        self.indian_fy = f"{self.calendar_year}-{str(self.calendar_year + 1)[-2:]}"
        self.assessment_year = f"{self.calendar_year + 1}-{str(self.calendar_year + 2)[-2:]}"

        print(f"\n=======================================================")
        print(f"  Schedule FA - Web Scraping Mode")
        print(f"  Target Year: {self.calendar_year}")
        print(f"  Applicable FY: {self.indian_fy} | AY: {self.assessment_year}")
        print(f"=======================================================\n")

        # Datasets & Caches
        self.df_sbi = self._fetch_sbi_rates_web()  # Initial download (will reload with extra dates later)
        self.company_cache = {}
        self._extra_ttbr_dates_loaded = False  # Track if we've loaded pre-FY dates

    def _fetch_sbi_rates_web(self, extra_dates=None):
        """Downloads and filters SBI forex rates using our own fetcher (SBI PDF first, GitHub fallback).

        Args:
            extra_dates: List of specific dates before FY to include (e.g., ['2024-11-08', '2024-09-15'])
        """
        # Use our own SBI forex fetcher instead of direct GitHub download
        df = None
        try:
            from sbi_forex_fetcher import fetch_sbi_forex_rates

            print("[*] Using SBI forex fetcher (tries SBI PDF first, then GitHub fallback)")
            df = fetch_sbi_forex_rates()

            # Convert DATE column to Date and extract TT BUY as TTBR
            if 'DATE' in df.columns and 'TT BUY' in df.columns:
                df['Date'] = pd.to_datetime(df['DATE']).dt.strftime('%Y-%m-%d')
                df['TTBR'] = pd.to_numeric(df['TT BUY'], errors='coerce')
                # Keep only Date and TTBR for compatibility
                df = df[['Date', 'TTBR']].copy()
            else:
                raise ValueError(f"Fetcher returned unexpected format. Columns: {list(df.columns)}")

        except Exception as e:
            print(f"[!] SBI forex fetcher failed: {e}")
            print(f"[!] Falling back to local CSV or legacy GitHub download")

        # If our fetcher failed, try reading local CSV first
        if df is None:
            local_csv = os.path.join('data', 'SBI_FOREX_CARD_RATES_USD.csv')
            if os.path.exists(local_csv):
                try:
                    print(f"[*] Reading local SBI forex data from {local_csv}...")
                    df = pd.read_csv(local_csv)

                    if 'DATE' in df.columns and 'TT BUY' in df.columns:
                        df['Date'] = pd.to_datetime(df['DATE']).dt.strftime('%Y-%m-%d')
                        df['TTBR'] = pd.to_numeric(df['TT BUY'], errors='coerce')
                        df = df[['Date', 'TTBR']].copy()
                        print(f"[OK] Loaded {len(df)} records from local CSV")
                    else:
                        raise ValueError(f"Local CSV has unexpected format. Columns: {list(df.columns)}")
                except Exception as e:
                    print(f"[!] Failed to read local CSV: {e}")
                    df = None

        # If still no data, try legacy GitHub download
        if df is None:
            print("[*] Downloading SBI TTBR rates from GitHub (legacy method)...")

            url = "https://raw.githubusercontent.com/sahilgupta/sbi-fx-ratekeeper/main/csv_files/SBI_REFERENCE_RATES_USD.csv"

            try:
                # Download the CSV file
                response = requests.get(url, timeout=30, verify=False)
                response.raise_for_status()

                # Parse CSV
                from io import StringIO
                df = pd.read_csv(StringIO(response.text))

                print(f"[*] CSV columns found: {list(df.columns)[:5]}")

                # GitHub CSV uses: DATE, TT BUY (not Date, TTBR)
                # Normalize column names
                if 'DATE' in df.columns:
                    df['Date'] = pd.to_datetime(df['DATE']).dt.strftime('%Y-%m-%d')
                    df['TTBR'] = pd.to_numeric(df['TT BUY'], errors='coerce')
                    df = df[['Date', 'TTBR']].copy()
                elif 'Date' not in df.columns or 'TTBR' not in df.columns:
                    raise ValueError(f"CSV must have 'DATE'/'Date' and 'TT BUY'/'TTBR' columns. Found: {list(df.columns)}")

                df = df.sort_values('Date').dropna(subset=['TTBR'])
            except Exception as e:
                print(f"[ERROR] Failed to download from GitHub: {e}")
                raise

        # At this point, df should have columns: Date, TTBR (from either our fetcher, local CSV, or GitHub legacy)
        # Ensure data is clean
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
        df['TTBR'] = pd.to_numeric(df['TTBR'], errors='coerce')
        df = df.sort_values('Date').dropna(subset=['TTBR'])

        # Filter for target year only (for peak/closing calculations)
        df_year = df[(df['Date'] >= self.start_date) & (df['Date'] <= self.end_date)]

        # Add specific pre-FY dates for initial value calculations (if provided)
        if extra_dates:
            df_extra = df[df['Date'].isin(extra_dates)]
            df_combined = pd.concat([df_year, df_extra]).drop_duplicates(subset=['Date']).sort_values('Date')

            if len(df_extra) > 0:
                print(f"[OK] Downloaded {len(df_year)} SBI TTBR records for {self.calendar_year}")
                print(f"[OK] Plus {len(df_extra)} specific dates before FY: {', '.join(sorted(extra_dates))}")
            else:
                print(f"[!] WARNING: Could not find TTBR for pre-FY dates: {', '.join(sorted(extra_dates))}")
                print(f"[OK] Downloaded {len(df_year)} SBI TTBR records for {self.calendar_year}")

            print(f"[OK] TTBR range: {df_combined['TTBR'].min():.2f} to {df_combined['TTBR'].max():.2f}")
            return df_combined[['Date', 'TTBR']]
        else:
            if df_year.empty:
                print(f"[!] WARNING: No SBI TTBR data for {self.calendar_year}")
                print(f"[!] Available date range: {df['Date'].min()} to {df['Date'].max()}")
                print("[!] Using fallback interpolated rates")
                raise ValueError(f"No data for year {self.calendar_year}")

            print(f"[OK] Downloaded {len(df_year)} SBI TTBR records for {self.calendar_year}")
            print(f"[OK] TTBR range: {df_year['TTBR'].min():.2f} to {df_year['TTBR'].max():.2f}")
            return df_year[['Date', 'TTBR']]

    def _detect_country(self, address):
        """Detect country from address and return ITR-compliant country name and code."""
        address_upper = address.upper()

        # ITR Schedule FA Country Codes (common ones)
        country_mapping = {
            "UNITED STATES": {"name": "UNITED STATES OF AMERICA", "code": "2"},
            "USA": {"name": "UNITED STATES OF AMERICA", "code": "2"},
            "U.S.A": {"name": "UNITED STATES OF AMERICA", "code": "2"},
            "CANADA": {"name": "CANADA", "code": "3"},
            "UNITED KINGDOM": {"name": "UNITED KINGDOM", "code": "1"},
            "UK": {"name": "UNITED KINGDOM", "code": "1"},
            "GERMANY": {"name": "GERMANY", "code": "4"},
            "FRANCE": {"name": "FRANCE", "code": "5"},
            "JAPAN": {"name": "JAPAN", "code": "6"},
            "AUSTRALIA": {"name": "AUSTRALIA", "code": "7"},
            "SWITZERLAND": {"name": "SWITZERLAND", "code": "8"},
            "NETHERLANDS": {"name": "NETHERLANDS", "code": "9"},
            "SINGAPORE": {"name": "SINGAPORE", "code": "10"},
            "HONG KONG": {"name": "HONG KONG", "code": "11"},
            "CHINA": {"name": "CHINA", "code": "12"},
            "SOUTH KOREA": {"name": "SOUTH KOREA", "code": "13"},
            "KOREA": {"name": "SOUTH KOREA", "code": "13"},
            "TAIWAN": {"name": "TAIWAN", "code": "14"},
            "INDIA": {"name": "INDIA", "code": "15"},
            "BRAZIL": {"name": "BRAZIL", "code": "16"},
            "MEXICO": {"name": "MEXICO", "code": "17"},
            "ISRAEL": {"name": "ISRAEL", "code": "18"},
            "IRELAND": {"name": "IRELAND", "code": "19"},
            "SPAIN": {"name": "SPAIN", "code": "20"},
            "ITALY": {"name": "ITALY", "code": "21"},
        }

        # Try to detect country from address
        for country_key, country_data in country_mapping.items():
            if country_key in address_upper:
                return country_data["name"], country_data["code"]

        # Default to USA if not detected
        print(f"     [!] Could not detect country from address, defaulting to USA")
        return "UNITED STATES OF AMERICA", "2"

    def _scrape_company_profile(self, symbol):
        """Scrapes company profile information from Yahoo Finance."""
        print(f"[*] Fetching company profile for {symbol}...")

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        driver = None
        try:
            driver = webdriver.Chrome(options=chrome_options)

            # Go to Yahoo Finance profile page
            url = f"https://finance.yahoo.com/quote/{symbol}/profile"
            driver.get(url)
            time.sleep(3)  # Wait for page to load

            # Extract company info from asset-profile section
            company_name = f"{symbol} Inc."  # Default
            full_address = "United States"  # Default
            zip_code = "00000"  # Default
            country_name = "UNITED STATES OF AMERICA"  # Default
            country_code = "2"  # Default (USA)

            try:
                # Find the asset-profile section using data-testid
                profile_section = driver.find_element(By.CSS_SELECTOR, "[data-testid='asset-profile']")

                # Extract company name (h3 within asset-profile)
                try:
                    company_name = profile_section.find_element(By.CSS_SELECTOR, "h3").text.strip()
                except:
                    print(f"[!] Could not extract company name from h3")

                # Extract address from nested classes: .company-details > .company-info > .address
                # The .address class has 3 divs: [street, city+state+zip, country]
                # We only need first 2 divs (street + city/state/zip)
                try:
                    # All are class names, not tags
                    address_element = profile_section.find_element(By.CSS_SELECTOR, ".company-details .company-info .address")

                    # Get all div children (3 divs: street, city+state+zip, country)
                    address_divs = address_element.find_elements(By.TAG_NAME, "div")

                    # Extract country from third div (for country detection)
                    country_text = ""
                    if len(address_divs) >= 3:
                        country_text = address_divs[2].text.strip()

                    # Extract zip from the second div (city, state, zip)
                    # Zip is the last word (after last space) that's 5 digits
                    if len(address_divs) >= 2:
                        city_state_zip = address_divs[1].text.strip()
                        # Split by space and get last element as zip
                        parts = city_state_zip.split()
                        if parts and len(parts[-1]) == 5 and parts[-1].isdigit():
                            zip_code = parts[-1]
                            # Remove zip from city_state_zip for address
                            city_state_no_zip = " ".join(parts[:-1])
                        else:
                            city_state_no_zip = city_state_zip

                        # Build full address from first div (street) + second div without zip
                        street = address_divs[0].text.strip() if len(address_divs) >= 1 else ""
                        full_address = f"{street} {city_state_no_zip}".strip()

                        # Use country text for detection (more accurate than address)
                        if country_text:
                            country_name, country_code = self._detect_country(country_text)
                    else:
                        # Fallback if structure is different
                        address_parts = [address_divs[i].text.strip() for i in range(min(2, len(address_divs))) if address_divs[i].text.strip()]
                        full_address = " ".join(address_parts)

                except Exception as e:
                    print(f"[!] Could not extract address element: {e}")

            except Exception as e:
                print(f"[!] Could not find asset-profile section: {e}")

            # Detect country from address (only if not already detected from third div)
            try:
                if not country_name or country_name == "":
                    country_name, country_code = self._detect_country(full_address)
            except:
                country_name, country_code = self._detect_country(full_address)

            # If zip code wasn't extracted from div structure, try regex as fallback
            if not zip_code or zip_code == "":
                import re
                zip_match = re.search(r'\b(\d{5})\b', full_address)
                zip_code = zip_match.group(1) if zip_match else "00000"

            print(f"[OK] {company_name}")
            print(f"     Country: {country_name} (Code: {country_code})")
            print(f"     Address: {full_address}")
            print(f"     Zip: {zip_code}")

            driver.quit()

            return {
                "company_name": company_name,
                "company_address": full_address,
                "zip_code": zip_code,
                "country_name": country_name,
                "country_code": country_code
            }

        except Exception as e:
            print(f"[!] Error scraping profile for {symbol}: {e}")
            if driver:
                driver.quit()
            return {
                "company_name": f"{symbol} Inc.",
                "company_address": "United States",
                "zip_code": "00000",
                "country_name": "UNITED STATES OF AMERICA",
                "country_code": "2"
            }

    def _read_client_statement(self, client_statement_path=None):
        """Reads closing balance from E*TRADE ClientStatement PDF."""
        # Auto-find ClientStatements_*.pdf in inputs folder
        if client_statement_path is None:
            import glob
            pattern = "inputs/ClientStatements_*.pdf"
            matching_files = glob.glob(pattern)

            if matching_files:
                client_statement_path = matching_files[0]  # Use first match
                print(f"[*] Found ClientStatement: {client_statement_path}")
            else:
                print(f"[i] No ClientStatement PDF found (inputs/ClientStatements_*.pdf)")
                print(f"[i] Will use calculated closing balance from holdings")
                return None

        if not os.path.exists(client_statement_path):
            print(f"[i] ClientStatement not found: {client_statement_path}")
            print(f"[i] Will use calculated closing balance from holdings")
            return None

        try:
            import PyPDF2
            import re

            print(f"[*] Reading ClientStatement: {client_statement_path}")

            with open(client_statement_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)

                # Read first page
                page = pdf_reader.pages[0]
                text = page.extract_text()

                # Extract account number (E*TRADE format: "146 - 239025 - 205 - 4 - 1" at top of statement)
                # The main account number is the first 9 digits (146239025)
                account_match = re.search(r'(\d{3})\s*-\s*(\d{6})\s*-\s*\d{3}\s*-\s*\d\s*-\s*\d', text)
                if account_match:
                    account_number = account_match.group(1) + account_match.group(2)  # Combine first two parts
                    print(f"[OK] Found Account Number: {account_number}")
                    # Store it for later use
                    self.extracted_account_number = account_number
                else:
                    self.extracted_account_number = None
                    print(f"[!] Could not extract account number from ClientStatement")

                # Extract "Ending Total Value (as of MM/DD/YY) $XX,XXX.XX"
                match = re.search(r'Ending Total Value.*?\$([0-9,]+\.[0-9]{2})', text)

                if match:
                    ending_value_str = match.group(1).replace(',', '')
                    ending_value_usd = float(ending_value_str)

                    print(f"[OK] Found Ending Total Value: ${ending_value_usd:,.2f}")
                    return ending_value_usd
                else:
                    print(f"[!] Could not find Ending Total Value in ClientStatement")
                    return None

        except ImportError:
            print(f"[!] PyPDF2 not installed. Run: pip install PyPDF2")
            print(f"[i] Will use calculated closing balance from holdings")
            return None
        except Exception as e:
            print(f"[!] Error reading ClientStatement: {e}")
            print(f"[i] Will use calculated closing balance from holdings")
            return None

    def _scrape_yahoo_finance(self, symbol):
        """Scrapes stock price data from Yahoo Finance using Selenium."""
        print(f"[*] Scraping {symbol} stock prices from Yahoo Finance...")
        print(f"[*] This may take 30-60 seconds, please wait...")

        # Setup Chrome options
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # Run in background
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.page_load_strategy = 'eager'  # Don't wait for all resources

        driver = None
        try:
            # Initialize driver - use system Chrome directly (no ChromeDriverManager for corporate networks)
            driver = webdriver.Chrome(options=chrome_options)

            # Navigate to history page
            url = f"https://finance.yahoo.com/quote/{symbol}/history/"
            print(f"[*] Opening {url}")
            driver.get(url)
            time.sleep(5)

            print("[*] Setting custom date range...")
            try:
                # Click on the date picker to open it
                date_picker = driver.find_element(By.CSS_SELECTOR, "div[data-testid='history-date-picker']")
                date_picker.click()
                time.sleep(3)

                # Find and fill the startDate field (format: mm/dd/yyyy)
                # Yahoo Finance might exclude boundaries, so request one day earlier
                print("[*] Setting start date to 12/31/2024 (to ensure we get 01/01/2025)...")
                start_date_input = driver.find_element(By.CSS_SELECTOR, "input[name='startDate']")
                start_date_input.clear()
                time.sleep(0.5)
                start_date_input.send_keys("12/31/2024")
                time.sleep(1)

                # Find and fill the endDate field (format: mm/dd/yyyy)
                # Request one day later to ensure we get 12/31/2025
                print("[*] Setting end date to 01/01/2026 (to ensure we get 12/31/2025)...")
                end_date_input = driver.find_element(By.CSS_SELECTOR, "input[name='endDate']")
                end_date_input.clear()
                time.sleep(0.5)
                end_date_input.send_keys("01/01/2026")
                time.sleep(1)

                # Try multiple selectors for Done button
                print("[*] Looking for Done button...")
                done_clicked = False
                done_selectors = [
                    "//button[text()='Done']",
                    "//button[contains(text(), 'Done')]",
                    "//span[text()='Done']/parent::button",
                    "//button[@type='submit']",
                    "button[type='submit']"
                ]

                for selector in done_selectors:
                    try:
                        if selector.startswith('//'):
                            done_button = driver.find_element(By.XPATH, selector)
                        else:
                            done_button = driver.find_element(By.CSS_SELECTOR, selector)
                        done_button.click()
                        done_clicked = True
                        print("[OK] Set date range: 12/31/2024 - 01/01/2026 (will filter to 2025), clicked Done")
                        break
                    except:
                        continue

                if not done_clicked:
                    print("[!] Could not find Done button, pressing Enter on end date field...")
                    from selenium.webdriver.common.keys import Keys
                    end_date_input.send_keys(Keys.RETURN)

                # Wait for table to reload with full year data
                print("[*] Waiting for table to reload with full year data...")
                time.sleep(10)

            except Exception as e:
                print(f"[!] Could not set date range: {str(e)[:80]}")
                print("[!] Using default view (last 12 months only)")

            # Scrape table data
            print("[*] Scraping price table...")
            rows = driver.find_elements(By.XPATH, "//table[@data-test='historical-prices']//tbody//tr")

            if not rows:
                # Try alternative selector
                rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

            if not rows:
                raise Exception("Could not find price table")

            print(f"[*] Found {len(rows)} rows in table, filtering for year {self.calendar_year}...")

            data = []
            for row in rows:
                try:
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if len(cols) >= 6:
                        date_str = cols[0].text.strip()
                        # Use Adj Close (Column 5) instead of Close (Column 4)
                        # Adj Close accounts for splits, dividends, and other corporate actions
                        adj_close_price = cols[5].text.replace(',', '').strip()

                        # Parse date
                        date_obj = pd.to_datetime(date_str, errors='coerce')
                        if pd.isna(date_obj):
                            continue

                        date_formatted = date_obj.strftime('%Y-%m-%d')

                        # Filter for target year only
                        if self.start_date <= date_formatted <= self.end_date:
                            data.append({
                                'Date': date_formatted,
                                'Stock_Close_USD': float(adj_close_price)
                            })
                except:
                    continue

            if not data:
                raise Exception(f"No data found for {self.calendar_year}")

            df = pd.DataFrame(data)
            df = df.sort_values('Date')
            print(f"[OK] Scraped {len(df)} trading days for {symbol} in {self.calendar_year}")
            return df

        except Exception as e:
            print(f"[!] ERROR scraping Yahoo Finance: {str(e)}")
            print(f"[!] Falling back to yfinance API")

            # Fallback to yfinance API
            try:
                import yfinance as yf
                ticker = yf.Ticker(symbol)
                df_stock = yf.download(symbol, start=self.start_date, end=self.end_date, progress=False)

                if not df_stock.empty:
                    if isinstance(df_stock.columns, pd.MultiIndex):
                        df_stock = df_stock['Close']
                    df_stock = df_stock.reset_index()
                    df_stock['Date'] = pd.to_datetime(df_stock['Date']).dt.strftime('%Y-%m-%d')
                    df_stock = df_stock.rename(columns={'Close': 'Stock_Close_USD'})
                    df_stock = df_stock[['Date', 'Stock_Close_USD']]
                    print(f"[OK] Fetched {len(df_stock)} days via yfinance API")
                    return df_stock
            except:
                pass

            return pd.DataFrame()

        finally:
            if driver:
                driver.quit()

    def get_company_details(self, symbol):
        """Fetches company info and stock prices via web scraping."""
        if symbol in self.company_cache:
            return self.company_cache[symbol]

        print(f"\n[*] Processing stock: {symbol}")

        # Default company details
        company_name = f"{symbol} Inc."
        address = "United States"
        zip_code = ""

        # Get company details from config
        config = load_config()
        if symbol in config.get("table_a3_companies", {}):
            company_info = config["table_a3_companies"][symbol]
            company_name = company_info.get("company_name", f"{symbol} Inc.")
            address = company_info.get("company_address", "United States")
            zip_code = company_info.get("zip_code", "")
        else:
            # Scrape company profile from Yahoo Finance if not in config
            profile = self._scrape_company_profile(symbol)
            company_name = profile.get("company_name", f"{symbol} Inc.")
            address = profile.get("company_address", "United States")
            zip_code = profile.get("zip_code", "")

        # Scrape stock prices
        df_stock = self._scrape_yahoo_finance(symbol)

        if df_stock.empty:
            raise ValueError(f"Could not fetch stock prices for {symbol}")

        # Build valuation matrix - ONLY use actual trading days (no interpolation!)
        # Merge stock prices with TTBR rates
        df_matrix = pd.merge(
            df_stock[['Date', 'Stock_Close_USD']],
            self.df_sbi[['Date', 'TTBR']],
            on='Date',
            how='inner'  # INNER join = only dates that have BOTH stock price AND TTBR
        )

        # Calculate INR valuation per share (only for actual trading days)
        df_matrix['Valuation_Per_Share_INR'] = df_matrix['Stock_Close_USD'] * df_matrix['TTBR']

        # Sort by date
        df_matrix = df_matrix.sort_values('Date').reset_index(drop=True)

        data = {
            "name": company_name,
            "address": address,
            "zip": zip_code,
            "matrix": df_matrix
        }
        self.company_cache[symbol] = data
        return data

    def calculate_tranche_values(self, symbol, qty, acq_date_str, sell_date_str=None, unit_cost_usd=0.0):
        """Computes Initial, Peak, and Closing values in INR for a lot based on holding period."""
        comp_data = self.get_company_details(symbol)
        df_matrix = comp_data["matrix"]

        hold_start = max(self.start_date, acq_date_str)
        hold_end = sell_date_str if (sell_date_str and sell_date_str <= self.end_date) else self.end_date

        window = df_matrix[(df_matrix['Date'] >= hold_start) & (df_matrix['Date'] <= hold_end)]
        if window.empty:
            window = df_matrix.tail(1)

        # 1. Initial Value (Cost basis at acquisition)
        # IMPORTANT: Use ACTUAL acquisition date for initial value, not hold_start
        # For shares acquired before FY (e.g., Nov 2024), we need the Nov 2024 TTBR
        init_row = df_matrix[df_matrix['Date'] == acq_date_str]

        if not init_row.empty:
            ttbr_init = init_row['TTBR'].values[0]
        else:
            # Acquisition date not in our matrix (before FY start or non-trading day)
            # Try to get TTBR from SBI data directly
            acq_ttbr_row = self.df_sbi[self.df_sbi['Date'] == acq_date_str]
            if not acq_ttbr_row.empty:
                ttbr_init = acq_ttbr_row['TTBR'].values[0]
            else:
                # Acquisition date not found (weekend/holiday)
                # Use nearest previous trading day TTBR
                sbi_before = self.df_sbi[self.df_sbi['Date'] < acq_date_str]
                if not sbi_before.empty:
                    closest_date = sbi_before['Date'].max()
                    ttbr_init = sbi_before[sbi_before['Date'] == closest_date]['TTBR'].values[0]
                    print(f"[i] {acq_date_str} is weekend/holiday, using previous trading day {closest_date} TTBR: {ttbr_init:.2f}")
                else:
                    # Really can't find anything, use fallback
                    ttbr_init = 83.50
                    print(f"[!] WARNING: TTBR not found for {acq_date_str}, using fallback {ttbr_init}")

        if unit_cost_usd > 0:
            initial_val = round(qty * unit_cost_usd * ttbr_init, 2)
        else:
            # Use market price on acquisition date if no unit cost provided
            if not init_row.empty:
                initial_val = round(qty * init_row['Valuation_Per_Share_INR'].values[0], 2)
            else:
                # Can't calculate without acquisition date data
                initial_val = 0
                print(f"[!] WARNING: Cannot calculate initial value for {acq_date_str}")

        # 2. Peak Value during holding window
        peak_val = round(qty * window['Valuation_Per_Share_INR'].max(), 2)

        # 3. Closing Value
        # If sold within this FY: Closing = 0 (no longer holding)
        # If not sold or sold after FY: Closing = value on Dec 31 (still holding)
        if sell_date_str and sell_date_str <= self.end_date:
            # Sold within the FY -> Closing balance is 0
            closing_val = 0
        else:
            # Still holding on Dec 31 -> Use Dec 31 value
            dec31_row = df_matrix[df_matrix['Date'] == self.end_date]
            close_per_share = dec31_row['Valuation_Per_Share_INR'].values[0] if not dec31_row.empty else df_matrix['Valuation_Per_Share_INR'].iloc[-1]
            closing_val = round(qty * close_per_share, 2)

        return initial_val, peak_val, closing_val

    def _scan_and_reload_ttbr_if_needed(self, bystatus_path, gl_path):
        """Scan E*TRADE files for pre-FY acquisition dates and reload TTBR if needed."""
        if self._extra_ttbr_dates_loaded:
            return  # Already loaded

        extra_dates = []

        # Scan ByStatus for pre-FY acquisitions
        try:
            df_bystatus = pd.read_excel(bystatus_path, sheet_name='Sellable')
            for _, row in df_bystatus.iterrows():
                if row.get('Sellable Qty.', 0) > 0:
                    acq_date_raw = row.get('Date Acquired')
                    if pd.notna(acq_date_raw):
                        acq_date = pd.to_datetime(acq_date_raw)
                        if acq_date.strftime('%Y-%m-%d') < self.start_date:
                            # Add the date and surrounding dates (for weekends/holidays)
                            for offset in range(-3, 4):  # 3 days before/after
                                date_with_offset = (acq_date + pd.Timedelta(days=offset)).strftime('%Y-%m-%d')
                                extra_dates.append(date_with_offset)
        except Exception as e:
            print(f"[!] Could not scan ByStatus for pre-FY dates: {e}")

        # Scan G&L for pre-FY acquisitions
        if gl_path and os.path.exists(gl_path):
            try:
                df_gl = pd.read_excel(gl_path, sheet_name='G&L_Expanded')
                df_sold = df_gl[df_gl['Record Type'] == 'Sell']
                for _, row in df_sold.iterrows():
                    acq_date_raw = row.get('Date Acquired')
                    if pd.notna(acq_date_raw):
                        acq_date = pd.to_datetime(acq_date_raw)
                        if acq_date.strftime('%Y-%m-%d') < self.start_date:
                            # Add the date and surrounding dates (for weekends/holidays)
                            for offset in range(-3, 4):  # 3 days before/after
                                date_with_offset = (acq_date + pd.Timedelta(days=offset)).strftime('%Y-%m-%d')
                                extra_dates.append(date_with_offset)
            except Exception as e:
                print(f"[!] Could not scan G&L for pre-FY dates: {e}")

        # If we found pre-FY dates, reload TTBR with them
        if extra_dates:
            unique_extra = sorted(set(extra_dates))
            # Show only the core dates, not all surrounding dates
            core_dates = sorted(set([d for d in extra_dates[::7]]))  # Sample to show main dates
            print(f"[*] Found acquisition dates before FY {self.calendar_year}, downloading TTBR with ±3 day buffer for weekends")
            self.df_sbi = self._fetch_sbi_rates_web(extra_dates=unique_extra)
            self._extra_ttbr_dates_loaded = True

    def process_etrade_exports(self, bystatus_path=None, gl_path=None, account_no="146239025", config=None):
        # Default paths - check inputs folder first, then root folder
        if bystatus_path is None:
            if os.path.exists("inputs/ByStatus_expanded.xlsx"):
                bystatus_path = "inputs/ByStatus_expanded.xlsx"
            elif os.path.exists("ByStatus_expanded.xlsx"):
                bystatus_path = "ByStatus_expanded.xlsx"
            else:
                raise FileNotFoundError("ByStatus_expanded.xlsx not found in 'inputs/' or root folder.")

        # G&L file is OPTIONAL - only needed if stocks were sold during the financial year
        if gl_path is None:
            if os.path.exists("inputs/G&L_Expanded.xlsx"):
                gl_path = "inputs/G&L_Expanded.xlsx"
            elif os.path.exists("G&L_Expanded.xlsx"):
                gl_path = "G&L_Expanded.xlsx"
            else:
                gl_path = None  # G&L file not found - this is OK

        if not os.path.exists(bystatus_path):
            raise FileNotFoundError(f"ByStatus file missing: '{bystatus_path}' not found.")

        # IMPORTANT: Scan files for pre-FY acquisition dates and reload TTBR if needed
        # This ensures we have TTBR for dates like Nov 2024 for initial value calculations
        self._scan_and_reload_ttbr_if_needed(bystatus_path, gl_path)

        try:
            df_bystatus = pd.read_excel(bystatus_path, sheet_name='Sellable')
        except Exception as e:
            raise ValueError(f"Error reading ByStatus file. Check if 'Sellable' sheet exists: {e}")

        df_open = df_bystatus[df_bystatus['Record Type'].isin(['Purchase', 'Grant'])].copy()

        # Try to read G&L file if it exists
        df_sold = pd.DataFrame()  # Empty dataframe if no sales
        df_sold_future = pd.DataFrame()  # Shares that will be sold after this FY
        if gl_path and os.path.exists(gl_path):
            try:
                df_gl = pd.read_excel(gl_path, sheet_name='G&L_Expanded')
                df_sold_all = df_gl[df_gl['Record Type'] == 'Sell'].copy()

                # IMPORTANT: We need to handle TWO types of sold shares:
                # 1. Sold IN this FY: Show Closing Balance = 0, Gross Proceeds = actual proceeds
                # 2. Sold AFTER this FY: Show Closing Balance > 0, Gross Proceeds = 0

                df_sold_all['Date Acquired'] = pd.to_datetime(df_sold_all['Date Acquired'])
                df_sold_all['Date Sold'] = pd.to_datetime(df_sold_all['Date Sold'])

                # Split into two categories:
                # Category 1: Actually sold in THIS FY (start_date <= sold <= end_date)
                # IMPORTANT: Only include sales from THIS FY onwards (exclude old sales from previous years)
                df_sold_in_fy = df_sold_all[
                    (df_sold_all['Date Acquired'] <= self.end_date) &
                    (df_sold_all['Date Sold'] >= self.start_date) &
                    (df_sold_all['Date Sold'] <= self.end_date)
                ].copy()

                # Category 2: Held in THIS FY but sold AFTER FY ends
                # These are for advance tax planning (sold in future years)
                # IMPORTANT: Only include future sales, not old sales from before this FY
                df_sold_future = df_sold_all[
                    (df_sold_all['Date Acquired'] <= self.end_date) &
                    (df_sold_all['Date Sold'] > self.end_date)
                ].copy()

                # Exclude any sales that happened BEFORE this FY started (old/historical sales)
                # Example: For FY 2025, exclude sales from 2024, 2023, etc.
                df_sold_in_fy = df_sold_in_fy[df_sold_in_fy['Date Sold'] >= self.start_date]

                # For now, we'll process Category 1 (actually sold in FY)
                df_sold = df_sold_in_fy

                # Category 3: Sales BEFORE this FY (excluded from A3 for tracking)
                # These are shares that were sold in previous years and should NOT appear in A3
                df_sold_before_fy = df_sold_all[
                    (df_sold_all['Date Sold'] < pd.to_datetime(self.start_date))
                ].copy()

                print(f"[OK] Found {len(df_sold_in_fy)} sales WITHIN FY {self.indian_fy}")
                print(f"[OK] Found {len(df_sold_future)} holdings that will be sold AFTER FY {self.indian_fy}")

                excluded = len(df_sold_before_fy)
                if excluded > 0:
                    print(f"[i] Excluded {excluded} sales from BEFORE FY {self.indian_fy} (will appear in 'Excluded from A3' sheet)")

            except Exception as e:
                print(f"[!] WARNING: Error reading G&L file: {e}")
                print("[i] Continuing without sold shares data")
                df_sold = pd.DataFrame()
        else:
            print("[i] G&L_Expanded.xlsx not found - assuming no sales in this financial year")

        equity_tranches = []

        # 1. Parse Open Lots (Unsold shares)
        for _, row in df_open.iterrows():
            qty = int(row['Sellable Qty.'])
            if qty == 0:
                continue

            # IMPORTANT: Only include holdings acquired ON OR BEFORE the end of the calendar year
            acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
            if acq_date > self.end_date:
                # Skip holdings acquired after the calendar year ends (e.g., 2026-05-08 for FY 2025)
                continue

            symbol = str(row['Symbol']).strip() if pd.notna(row['Symbol']) else "AMD"
            comp_info = self.get_company_details(symbol)

            plan_type = str(row['Plan Type'])
            nature_prefix = "ESPP" if "ESPP" in plan_type else "RSU"
            nature = f"{nature_prefix} ({qty} shares)" if qty != 1 else f"{nature_prefix} ({qty} share)"

            # Use Purchase Date FMV for initial value (Fair Market Value on acquisition date)
            # This is the correct value for ITR Schedule FA
            purchase_fmv_str = str(row['Purchase Date FMV']).replace('$', '').replace(',', '').strip()
            unit_cost = float(purchase_fmv_str) if purchase_fmv_str and purchase_fmv_str != '--' else float(row['Est. Cost Basis (per share):'])

            init_val, peak_val, close_val = self.calculate_tranche_values(symbol, qty, acq_date, unit_cost_usd=unit_cost)

            equity_tranches.append({
                "CountryName": "UNITED STATES OF AMERICA",
                "CountryCodeExcludingIndia": 2,
                "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                "ZipCode": str(comp_info["zip"]),
                "NatureOfEntity": nature,
                "InterestAcquiringDate": acq_date,
                "InitialValOfInvstmnt": init_val,
                "PeakBalanceDuringPeriod": peak_val,
                "ClosingBalance": close_val,
                "_FMV_USD": unit_cost,  # Store FMV for reference (not exported to JSON)
                "TotGrossAmtPaidCredited": 0,
                "TotGrossProceeds": 0
            })

        # 2. Parse Sold Lots (Actually sold WITHIN this FY)
        # These have: Closing Balance = 0 (no longer holding)
        #            Gross Proceeds = actual proceeds from sale
        for _, row in df_sold.iterrows():
            qty = int(row['Quantity'])
            symbol = str(row['Symbol']).strip() if pd.notna(row['Symbol']) else "AMD"
            comp_info = self.get_company_details(symbol)

            # Determine if it's RSU or ESPP based on cost basis
            # ESPP typically has lower cost basis due to discount, but safest is to check G&L file
            # For now, assume sold shares are RSU (most common for equity awards)
            nature = f"RSU ({qty} share) Sold" if qty == 1 else f"RSU ({qty} shares) Sold"
            acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
            sell_date = pd.to_datetime(row['Date Sold']).strftime('%Y-%m-%d')

            # Use Adjusted Cost Basis Per Share for initial value
            # This represents the FMV on vesting/acquisition date and is used for tax purposes
            unit_cost = float(row['Adjusted Cost Basis Per Share'])

            proceeds_usd = float(row['Total Proceeds'])

            init_val, peak_val, close_val = self.calculate_tranche_values(symbol, qty, acq_date, sell_date_str=sell_date, unit_cost_usd=unit_cost)

            df_matrix = comp_info["matrix"]
            sell_row = df_matrix[df_matrix['Date'] == sell_date]
            sell_ttbr = sell_row['TTBR'].values[0] if not sell_row.empty else 89.47
            proceeds_inr = round(proceeds_usd * sell_ttbr, 2)

            equity_tranches.append({
                "CountryName": "UNITED STATES OF AMERICA",
                "CountryCodeExcludingIndia": 2,
                "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                "ZipCode": str(comp_info["zip"]),
                "NatureOfEntity": nature,
                "InterestAcquiringDate": acq_date,
                "InitialValOfInvstmnt": init_val,
                "PeakBalanceDuringPeriod": peak_val,
                "ClosingBalance": close_val,
                "_FMV_USD": unit_cost,  # Store FMV for reference (not exported to JSON)
                "_SaleDate": sell_date,  # Store sale date for Capital Gains sheet (not exported to JSON)
                "_GrossProceeds": proceeds_inr,  # Store proceeds for Capital Gains sheet
                "TotGrossAmtPaidCredited": 0,
                "TotGrossProceeds": proceeds_inr
            })

        # 3. Parse Future-Sold Lots (Held in this FY but sold AFTER FY ends)
        # These have: Closing Balance = closing value on Dec 31 (still holding)
        #            Gross Proceeds = 0 (not sold yet in this FY)
        for _, row in df_sold_future.iterrows():
            qty = int(row['Quantity'])
            symbol = str(row['Symbol']).strip() if pd.notna(row['Symbol']) else "AMD"
            comp_info = self.get_company_details(symbol)

            # Determine plan type from G&L
            plan_type = str(row.get('Plan Type', ''))
            nature_prefix = "ESPP" if "ESPP" in plan_type else "RSU"
            nature = f"{nature_prefix} ({qty} shares) - Sold" if qty != 1 else f"{nature_prefix} ({qty} share) - Sold"

            acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
            sell_date = pd.to_datetime(row['Date Sold']).strftime('%Y-%m-%d')
            # NOTE: We do NOT pass sell_date to calculate_tranche_values because we want closing balance on Dec 31, not 0
            unit_cost = float(row['Adjusted Cost Basis Per Share'])

            # Calculate sale proceeds in INR (for Capital Gains sheet)
            proceeds_usd = float(row['Total Proceeds'])

            # Get TTBR for sell date (may need to fetch if not in current year)
            df_matrix = comp_info["matrix"]
            sell_row = df_matrix[df_matrix['Date'] == sell_date]
            if not sell_row.empty:
                sell_ttbr = sell_row['TTBR'].values[0]
            else:
                # Sell date is outside FY, use a reasonable rate (or fetch from SBI)
                sell_ttbr = 89.47  # Fallback
            proceeds_inr = round(proceeds_usd * sell_ttbr, 2)

            # Calculate values WITHOUT sell date (so closing balance is > 0)
            init_val, peak_val, close_val = self.calculate_tranche_values(symbol, qty, acq_date, unit_cost_usd=unit_cost)

            equity_tranches.append({
                "CountryName": "UNITED STATES OF AMERICA",
                "CountryCodeExcludingIndia": 2,
                "NameOfEntity": self.clean_text_for_itr(comp_info["name"]),
                "AddressOfEntity": self.clean_text_for_itr(comp_info["address"]),
                "ZipCode": str(comp_info["zip"]),
                "NatureOfEntity": nature,
                "InterestAcquiringDate": acq_date,
                "InitialValOfInvstmnt": init_val,
                "PeakBalanceDuringPeriod": peak_val,
                "ClosingBalance": close_val,  # > 0 because still holding on Dec 31
                "_FMV_USD": unit_cost,
                "_SaleDate": sell_date,  # Store sale date for Capital Gains sheet
                "_GrossProceeds": proceeds_inr,  # Store proceeds for Capital Gains sheet
                "TotGrossAmtPaidCredited": 0,
                "TotGrossProceeds": 0  # 0 because not sold yet in this FY
            })

        # Table A2 Custodial Account Aggregation
        # CORRECT METHOD: Calculate daily total account value and find maximum
        # (Not sum of individual peaks, since they occur on different dates)

        # Get the daily matrix (stock prices and TTBR for each day)
        comp_info = self.get_company_details("AMD")
        df_daily = comp_info["matrix"].copy()

        # For each day, calculate total account value in USD and INR
        daily_account_usd = []
        daily_account_inr = []

        for _, day_row in df_daily.iterrows():
            date = day_row['Date']
            stock_price = day_row['Stock_Close_USD']
            ttbr = day_row['TTBR']

            # Sum all holdings owned on this date
            total_shares = 0
            for tranche in equity_tranches:
                acq_date = tranche['InterestAcquiringDate']

                # Determine if we owned this holding on this date
                # Owned if: acquired on/before this date AND (not sold OR sold after this date)
                if acq_date <= date:
                    # Check if sold
                    if "Sold" in tranche['NatureOfEntity']:
                        # For sold shares, we need to check the sale date
                        # We'll assume they were held through Dec 31 for peak calculation
                        # (Since sold shares also contribute to peak during the year)
                        pass  # Include them

                    # Extract quantity from nature string
                    import re
                    nature = tranche['NatureOfEntity']
                    qty_match = re.search(r'\((\d+)\s+shares?\)', nature)
                    if qty_match:
                        qty = int(qty_match.group(1))
                        total_shares += qty

            # Calculate total account value for this day
            account_value_usd = total_shares * stock_price
            account_value_inr = account_value_usd * ttbr

            daily_account_usd.append(account_value_usd)
            daily_account_inr.append(account_value_inr)

        # Add to dataframe for reference
        df_daily['Total Account Value (USD)'] = daily_account_usd
        df_daily['Total Account Value (INR)'] = daily_account_inr

        # Find peak account value
        peak_idx = df_daily['Total Account Value (INR)'].idxmax()
        peak_row = df_daily.loc[peak_idx]

        total_peak_account_inr = round(peak_row['Total Account Value (INR)'], 2)
        total_peak_account_usd = round(peak_row['Total Account Value (USD)'], 2)
        peak_date = peak_row['Date']
        peak_ttbr = round(peak_row['TTBR'], 2)

        print(f"[*] A2 Peak calculated from daily account values:")
        print(f"    Peak Date: {peak_date}")
        print(f"    Account Value: ${total_peak_account_usd:.2f} × {peak_ttbr:.2f} = {total_peak_account_inr:.2f} INR")

        # Store the daily matrix with account values for the reference sheet
        self._daily_account_matrix = df_daily

        # Try to get closing balance from ClientStatement PDF (more accurate)
        client_statement_closing_usd = self._read_client_statement()

        if client_statement_closing_usd:
            # Use ClientStatement value and convert to INR
            # Get TTBR rate for Dec 31
            comp_info = self.get_company_details("AMD")  # Use first symbol's matrix
            df_matrix = comp_info["matrix"]
            dec31_row = df_matrix[df_matrix['Date'] == self.end_date]
            closing_ttbr = dec31_row['TTBR'].values[0] if not dec31_row.empty else 89.47

            total_closing_account_inr = round(client_statement_closing_usd * closing_ttbr, 2)
            print(f"[OK] Using ClientStatement closing: ${client_statement_closing_usd:.2f} × {closing_ttbr:.2f} = {total_closing_account_inr:.2f} INR")
        else:
            # Fallback: sum of A3 closing balances
            total_closing_account_inr = sum(t["ClosingBalance"] for t in equity_tranches)
            print(f"[i] Using calculated closing from A3 sum: {total_closing_account_inr:,} INR")

        # Build Table A2 entry from config (single custodial account)
        acc_config = config.get("custodial_account", {})

        # Use extracted account number from ClientStatement if available, then config, then parameter, then empty
        extracted = getattr(self, 'extracted_account_number', None)
        from_config = acc_config.get("account_number")

        if extracted:
            final_account_no = extracted
            print(f"[i] Using account number from ClientStatement: {final_account_no}")
        elif from_config:
            final_account_no = from_config
            print(f"[i] Using account number from config.json: {final_account_no}")
        elif account_no:
            final_account_no = account_no
            print(f"[i] Using account number from parameter: {final_account_no}")
        else:
            final_account_no = ""
            print(f"[i] Account number not found - leaving empty")

        custodial_accounts = [{
            "CountryName": acc_config.get("country_name", "UNITED STATES OF AMERICA"),
            "CountryCodeExcludingIndia": int(acc_config.get("country_code", 2)),
            "FinancialInstName": self.clean_text_for_itr(acc_config.get("financial_institution_name", "E*TRADE Securities LLC")),
            "FinancialInstAddress": self.clean_text_for_itr(acc_config.get("financial_institution_address", "1271 Avenue of the Americas New York NY 10020 United States")),
            "ZipCode": str(acc_config.get("zip_code", "10020")),
            "AccountNumber": str(final_account_no),
            "Status": acc_config.get("status", "BENEFICIAL_OWNER"),
            "AccOpenDate": acc_config.get("account_opening_date", ""),
            "PeakBalanceDuringPeriod": total_peak_account_inr,
            "ClosingBalance": total_closing_account_inr,
            "GrossAmtPaidCredited": 0,
            "NatureOfAmount": "N"
        }]

        # Clean internal fields (those starting with _) before exporting to JSON
        equity_tranches_clean = []
        for tranche in equity_tranches:
            clean_tranche = {k: v for k, v in tranche.items() if not k.startswith('_')}
            equity_tranches_clean.append(clean_tranche)

        output_data = {
            "ScheduleFA": {
                "AssessmentYear": self.assessment_year,
                "IndianFY": self.indian_fy,
                "DtlsForeignCustodialAcc": custodial_accounts,
                "DtlsForeignEquityDebtInterest": equity_tranches_clean,
                "DetailsOfTrustOutIndiaTrustee": []
            }
        }

        # Create outputs folder
        config = load_config()
        output_dir = "outputs"
        os.makedirs(output_dir, exist_ok=True)

        json_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}.json")
        excel_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}.xlsx")

        with open(json_filename, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2)

        df_a2 = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignCustodialAcc"])
        df_a3 = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignEquityDebtInterest"])

        # Round all numeric values to integers for Excel display (A2 and A3)
        numeric_cols_a2 = ['PeakBalanceDuringPeriod', 'ClosingBalance', 'GrossAmtPaidCredited']
        for col in numeric_cols_a2:
            if col in df_a2.columns:
                df_a2[col] = df_a2[col].round(0).astype(int)

        numeric_cols_a3 = ['InitialValOfInvstmnt', 'PeakBalanceDuringPeriod', 'ClosingBalance',
                           'TotGrossAmtPaidCredited', 'TotGrossProceeds']
        for col in numeric_cols_a3:
            if col in df_a3.columns:
                df_a3[col] = df_a3[col].round(0).astype(int)

        # Force ZipCode, AccountNumber, and CountryCode to be text (prevent Excel auto-conversion)
        if 'ZipCode' in df_a2.columns:
            df_a2['ZipCode'] = df_a2['ZipCode'].astype(str)
        if 'AccountNumber' in df_a2.columns:
            df_a2['AccountNumber'] = df_a2['AccountNumber'].astype(str)

        if 'ZipCode' in df_a3.columns:
            df_a3['ZipCode'] = df_a3['ZipCode'].astype(str)
        if 'CountryCodeExcludingIndia' in df_a3.columns:
            df_a3['CountryCodeExcludingIndia'] = df_a3['CountryCodeExcludingIndia'].astype(str)

        # Create reference sheet with daily AMD prices and TTBR rates (simplified)
        df_reference = self._daily_account_matrix[['Date', 'Stock_Close_USD', 'TTBR',
                                                     'Valuation_Per_Share_INR']].copy()
        df_reference = df_reference.rename(columns={
            'Date': 'Date',
            'Stock_Close_USD': 'AMD Stock Price (USD)',
            'TTBR': 'SBI TTBR Rate (USD to INR)',
            'Valuation_Per_Share_INR': 'AMD Value per Share (INR)'
        })
        # Round to 2 decimal places for readability
        df_reference['AMD Stock Price (USD)'] = df_reference['AMD Stock Price (USD)'].round(2)
        df_reference['SBI TTBR Rate (USD to INR)'] = df_reference['SBI TTBR Rate (USD to INR)'].round(2)
        df_reference['AMD Value per Share (INR)'] = df_reference['AMD Value per Share (INR)'].round(2)

        # Add peak per-share info summary (from column D)
        # Find peak per-share value (maximum of AMD Value per Share INR)
        peak_share_idx = df_reference['AMD Value per Share (INR)'].idxmax()
        peak_share_row = df_reference.loc[peak_share_idx]

        # Add two spacer columns for visual separation
        df_reference['--'] = ''
        df_reference['--'] = ''

        # Create peak per-share info summary
        peak_labels = ['PEAK PER-SHARE INFO', 'Peak Date', 'Stock Price (USD)', 'TTBR', 'Peak INR Value']
        peak_values = [
            '',
            str(peak_share_row['Date']),
            float(peak_share_row['AMD Stock Price (USD)']),
            float(peak_share_row['SBI TTBR Rate (USD to INR)']),
            float(peak_share_row['AMD Value per Share (INR)'])
        ]

        # Pad to match dataframe length
        while len(peak_labels) < len(df_reference):
            peak_labels.append('')
            peak_values.append('')

        df_reference['Peak Info'] = peak_labels
        df_reference['Value'] = peak_values

        # Create Pre-FY Acquisitions sheet for holdings acquired before the financial year
        pre_fy_data = []
        for tranche in equity_tranches:
            acq_date = tranche['InterestAcquiringDate']
            if acq_date < self.start_date:  # Acquired before FY start
                # Get the quantity from nature field
                nature = tranche['NatureOfEntity']

                # Extract actual FMV from E*TRADE file (stored during processing)
                fmv_usd = tranche.get('_FMV_USD', 0)
                initial_inr = tranche['InitialValOfInvstmnt']

                # Try to get acquisition date from our TTBR data
                acq_ttbr_row = self.df_sbi[self.df_sbi['Date'] == acq_date]
                if not acq_ttbr_row.empty:
                    ttbr_used = acq_ttbr_row['TTBR'].values[0]
                    ttbr_source = acq_date
                else:
                    # It was a weekend/holiday, find what we used
                    sbi_before = self.df_sbi[self.df_sbi['Date'] < acq_date]
                    if not sbi_before.empty:
                        ttbr_source = sbi_before['Date'].max()
                        ttbr_used = sbi_before[sbi_before['Date'] == ttbr_source]['TTBR'].values[0]
                    else:
                        ttbr_source = 'Fallback'
                        ttbr_used = 83.50

                # Extract quantity from nature string for per-share FMV
                import re
                qty_match = re.search(r'\((\d+)\s+shares?\)', nature)
                qty = int(qty_match.group(1)) if qty_match else 1

                pre_fy_data.append({
                    'Nature': nature,
                    'Acquisition Date': acq_date,
                    'Day of Week': pd.to_datetime(acq_date).strftime('%A'),
                    'Quantity': qty,
                    'FMV per Share (USD)': round(fmv_usd, 2),
                    'Total FMV (USD)': round(fmv_usd * qty, 2),
                    'TTBR Used': round(ttbr_used, 2),
                    'TTBR Date': ttbr_source,
                    'Initial Value (INR)': initial_inr,
                    'Note': 'Weekend/Holiday' if ttbr_source != acq_date else 'Trading Day'
                })

        df_pre_fy = pd.DataFrame(pre_fy_data) if pre_fy_data else pd.DataFrame({
            'Note': ['No acquisitions before FY start']
        })

        # Create A2 Peak Calculation sheet - showing daily account values
        df_a2_peak = self._daily_account_matrix[['Date', 'Stock_Close_USD', 'TTBR',
                                                   'Total Account Value (USD)',
                                                   'Total Account Value (INR)']].copy()
        df_a2_peak = df_a2_peak.rename(columns={
            'Date': 'Date',
            'Stock_Close_USD': 'AMD Stock Price (USD)',
            'TTBR': 'SBI TTBR (USD to INR)',
            'Total Account Value (USD)': 'Total Account Value (USD)',
            'Total Account Value (INR)': 'Total Account Value (INR)'
        })
        # Round values
        df_a2_peak['AMD Stock Price (USD)'] = df_a2_peak['AMD Stock Price (USD)'].round(2)
        df_a2_peak['SBI TTBR (USD to INR)'] = df_a2_peak['SBI TTBR (USD to INR)'].round(2)
        df_a2_peak['Total Account Value (USD)'] = df_a2_peak['Total Account Value (USD)'].round(2)
        df_a2_peak['Total Account Value (INR)'] = df_a2_peak['Total Account Value (INR)'].round(2)

        # Add peak summary (using plain values, not formatted strings, to avoid Excel errors)
        peak_account_idx = df_a2_peak['Total Account Value (INR)'].idxmax()
        peak_account_row = df_a2_peak.loc[peak_account_idx]

        # Create summary columns with proper initialization
        summary_labels = []
        summary_values = []

        # Add peak summary data
        summary_items = [
            ('Peak Date', str(peak_account_row['Date'])),
            ('Stock Price (USD)', float(peak_account_row['AMD Stock Price (USD)'])),
            ('TTBR', float(peak_account_row['SBI TTBR (USD to INR)'])),
            ('Account Value (USD)', float(peak_account_row['Total Account Value (USD)'])),
            ('Account Value (INR)', round(peak_account_row['Total Account Value (INR)'], 2)),
            ('', ''),
            ('A2 Peak Balance (INR)', round(peak_account_row['Total Account Value (INR)'], 2))
        ]

        for label, value in summary_items:
            summary_labels.append(label)
            summary_values.append(value)

        # Pad with empty strings to match DataFrame length
        while len(summary_labels) < len(df_a2_peak):
            summary_labels.append('')
            summary_values.append('')

        # Add summary columns with proper data
        df_a2_peak['--'] = ''
        df_a2_peak['--'] = ''
        df_a2_peak['PEAK SUMMARY'] = summary_labels
        df_a2_peak['Value'] = summary_values

        # Create Capital Gains sheet (for stocks sold WITHIN this FY)
        # This calculates advance tax obligations based on sale date
        # Advance Tax Schedule (Income Tax Rule 234C):
        #   - By July 15: 15% of tax
        #   - By Sep 15: 45% of tax (cumulative)
        #   - By Dec 15: 75% of tax (cumulative)
        #   - By Mar 15: 100% of tax (cumulative)
        capital_gains_data = []

        for tranche in equity_tranches:
            # Include sales from THIS FY onwards (exclude old sales from previous years)
            # This includes: 1) Sales in current FY, 2) Future sales (for advance tax planning)
            if tranche.get('_SaleDate') and tranche.get('_GrossProceeds'):
                sale_date = pd.to_datetime(tranche.get('_SaleDate', ''))

                # FILTER: Only include sales from start of THIS FY onwards
                # Example: For FY 2025, only include sales from 2025-01-01 onwards
                if sale_date < pd.to_datetime(self.start_date):
                    continue  # Skip old sales from previous years

                # Extract quantity from NatureOfEntity (e.g., "RSU (26 shares) Sold" -> 26)
                qty_str = tranche['NatureOfEntity'].split('(')[1].split(' ')[0] if '(' in tranche['NatureOfEntity'] else ''

                acq_date = pd.to_datetime(tranche['InterestAcquiringDate'])

                # Calculate holding period in months
                holding_months = (sale_date.year - acq_date.year) * 12 + (sale_date.month - acq_date.month)

                # Determine tax type and rate
                # LTCG (Long Term Capital Gains): > 24 months = 12.5% tax
                # STCG (Short Term Capital Gains): <= 24 months = 31.2% tax
                if holding_months > 24:
                    tax_type = "LTCG"
                    tax_rate = 0.125  # 12.5%
                else:
                    tax_type = "STCG"
                    tax_rate = 0.312  # 31.2%

                # Use _GrossProceeds (from G&L) for ALL sales (not TotGrossProceeds which is 0 for future sales)
                import math
                gross_proceeds = math.ceil(tranche['_GrossProceeds'])  # Round UP proceeds
                cost_basis = math.ceil(tranche['InitialValOfInvstmnt'])  # Round UP cost basis
                capital_gain = gross_proceeds - cost_basis  # Already rounded up
                tax_amount = math.ceil(capital_gain * tax_rate)  # Round UP tax

                # Calculate advance tax schedule based on sale date (always round UP)
                sale_month = sale_date.month
                if sale_month <= 6:  # Sold before July 15
                    adv_tax_jul = math.ceil(tax_amount * 0.15)
                    adv_tax_sep = math.ceil(tax_amount * 0.45)
                    adv_tax_dec = math.ceil(tax_amount * 0.75)
                    adv_tax_mar = tax_amount
                elif sale_month <= 8:  # Sold between July 16 - Sep 15
                    adv_tax_jul = 0
                    adv_tax_sep = math.ceil(tax_amount * 0.45)
                    adv_tax_dec = math.ceil(tax_amount * 0.75)
                    adv_tax_mar = tax_amount
                elif sale_month <= 11:  # Sold between Sep 16 - Dec 15
                    adv_tax_jul = 0
                    adv_tax_sep = 0
                    adv_tax_dec = math.ceil(tax_amount * 0.75)
                    adv_tax_mar = tax_amount
                else:  # Sold between Dec 16 - Mar 15
                    adv_tax_jul = 0
                    adv_tax_sep = 0
                    adv_tax_dec = 0
                    adv_tax_mar = tax_amount

                capital_gains_data.append({
                    'Nature': tranche['NatureOfEntity'].replace(' Sold', '').replace(' - Sold', ''),
                    'Quantity': int(qty_str) if qty_str else 0,
                    'Acquisition Date': tranche['InterestAcquiringDate'],
                    'Sale Date': tranche.get('_SaleDate', ''),
                    'Holding Period (months)': holding_months,
                    'Tax Type': tax_type,
                    'Cost Basis (INR)': cost_basis,  # Rounded up
                    'Sale Proceeds (INR)': gross_proceeds,  # Rounded up
                    'Capital Gain (INR)': capital_gain,  # Rounded up
                    'Tax Rate': f"{tax_rate*100}%",
                    'Tax Amount (INR)': tax_amount,  # Rounded up
                    'Adv Tax by Jul 15 (15%)': adv_tax_jul,  # Rounded up
                    'Adv Tax by Sep 15 (45%)': adv_tax_sep,  # Rounded up
                    'Adv Tax by Dec 15 (75%)': adv_tax_dec,  # Rounded up
                    'Adv Tax by Mar 15 (100%)': adv_tax_mar  # Rounded up
                })

        # Create two separate tables for Capital Gains sheet

        # Table 1: Sale Details (without advance tax columns)
        df_sale_details = pd.DataFrame([{
            'Nature': item['Nature'],
            'Quantity': item['Quantity'],
            'Acquisition Date': item['Acquisition Date'],
            'Sale Date': item['Sale Date'],
            'Holding Period (months)': item['Holding Period (months)'],
            'Tax Type': item['Tax Type'],
            'Cost Basis (INR)': item['Cost Basis (INR)'],
            'Sale Proceeds (INR)': item['Sale Proceeds (INR)'],
            'Capital Gain (INR)': item['Capital Gain (INR)'],
            'Tax Rate': item['Tax Rate'],
            'Tax Amount (INR)': item['Tax Amount (INR)']
        } for item in capital_gains_data])

        # Table 2: Advance Tax Schedule Summary (single row with totals)
        if capital_gains_data:
            total_tax = sum(item['Tax Amount (INR)'] for item in capital_gains_data)
            total_adv_jul = sum(item['Adv Tax by Jul 15 (15%)'] for item in capital_gains_data)
            total_adv_sep = sum(item['Adv Tax by Sep 15 (45%)'] for item in capital_gains_data)
            total_adv_dec = sum(item['Adv Tax by Dec 15 (75%)'] for item in capital_gains_data)
            total_adv_mar = sum(item['Adv Tax by Mar 15 (100%)'] for item in capital_gains_data)

            df_advance_tax = pd.DataFrame([{
                'Description': 'TOTAL ADVANCE TAX SCHEDULE (Rule 234C)',
                'Total Tax (INR)': total_tax,
                'By Jul 15 (15%)': total_adv_jul,
                'By Sep 15 (45%)': total_adv_sep,
                'By Dec 15 (75%)': total_adv_dec,
                'By Mar 15 (100%)': total_adv_mar
            }])
        else:
            df_advance_tax = pd.DataFrame(columns=[
                'Description', 'Total Tax (INR)', 'By Jul 15 (15%)',
                'By Sep 15 (45%)', 'By Dec 15 (75%)', 'By Mar 15 (100%)'
            ])

        # If no sales, create empty DataFrames with columns
        if df_sale_details.empty:
            df_sale_details = pd.DataFrame(columns=[
                'Nature', 'Quantity', 'Acquisition Date', 'Sale Date', 'Holding Period (months)',
                'Tax Type', 'Cost Basis (INR)', 'Sale Proceeds (INR)', 'Capital Gain (INR)',
                'Tax Rate', 'Tax Amount (INR)'
            ])

        # Create "Excluded from A3" sheet - shows sales from previous years
        # This helps track what got removed from A3 as years progress
        excluded_a3_data = []
        if 'df_sold_before_fy' in locals() and not df_sold_before_fy.empty:
            for _, row in df_sold_before_fy.iterrows():
                qty = int(row['Quantity'])
                plan_type = str(row.get('Plan Type', ''))
                nature_prefix = "ESPP" if "ESPP" in plan_type else "RSU"

                acq_date = pd.to_datetime(row['Date Acquired']).strftime('%Y-%m-%d')
                sell_date = pd.to_datetime(row['Date Sold']).strftime('%Y-%m-%d')

                excluded_a3_data.append({
                    'Nature': f"{nature_prefix} ({qty} shares)" if qty != 1 else f"{nature_prefix} ({qty} share)",
                    'Quantity': qty,
                    'Acquisition Date': acq_date,
                    'Sale Date': sell_date,
                    'Year Sold': pd.to_datetime(sell_date).year,
                    'Reason': f"Sold in {pd.to_datetime(sell_date).year} (before FY {self.calendar_year})",
                    'Total Proceeds (USD)': round(float(row['Total Proceeds']), 2),
                    'Adjusted Cost Basis Per Share (USD)': round(float(row['Adjusted Cost Basis Per Share']), 2)
                })

        df_excluded_a3 = pd.DataFrame(excluded_a3_data) if excluded_a3_data else pd.DataFrame({
            'Note': [f'No sales excluded from A3 for FY {self.calendar_year}',
                     'All sales in G&L file are either in current FY or future years']
        })

        # Round numeric values in Excluded A3 sheet
        if excluded_a3_data:
            numeric_cols_excluded = ['Initial Value (INR)', 'Peak Value (INR)', 'Sale Proceeds (INR)',
                                      'Adjusted Cost Basis Per Share (USD)']
            for col in numeric_cols_excluded:
                if col in df_excluded_a3.columns:
                    if col == 'Adjusted Cost Basis Per Share (USD)':
                        df_excluded_a3[col] = df_excluded_a3[col].round(2)  # Keep 2 decimals for USD per share
                    else:
                        df_excluded_a3[col] = df_excluded_a3[col].round(0).astype(int)  # Round INR to integers

        with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
            df_a2.to_excel(writer, sheet_name="Table A2 Custodial Acc", index=False)
            df_a3.to_excel(writer, sheet_name="Table A3 Equity Interest", index=False)
            df_excluded_a3.to_excel(writer, sheet_name="Excluded from A3", index=False)

            # Write Capital Gains sheet with two tables
            # Table 1: Sale Details (starts at row 1)
            df_sale_details.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=0)

            # Table 2: Advance Tax Summary (starts after Table 1 + 3 blank rows)
            start_row_table2 = len(df_sale_details) + 4  # +1 for header, +3 for spacing
            df_advance_tax.to_excel(writer, sheet_name="Capital Gains", index=False, startrow=start_row_table2)
            df_reference.to_excel(writer, sheet_name="Reference - Daily Rates", index=False)
            df_a2_peak.to_excel(writer, sheet_name="A2 Peak Calculation", index=False)
            pre_fy_sheet_name = f"Pre-{self.calendar_year} Holdings Init Val"
            df_pre_fy.to_excel(writer, sheet_name=pre_fy_sheet_name, index=False)

            # Apply beautiful formatting to all sheets
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Color scheme
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
            header_font = Font(bold=True, color="FFFFFF", size=11)  # White bold
            alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
            total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
            total_font = Font(bold=True, size=11)
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]

                # Format header row
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border_thin

                # Format data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    # Alternating row colors (skip for Capital Gains table 2)
                    if sheet_name != "Capital Gains" or row_idx <= len(df_sale_details) + 1:
                        if row_idx % 2 == 0:
                            for cell in row:
                                if cell.value is not None:
                                    cell.fill = alt_row_fill

                    # Apply borders and alignment
                    for cell in row:
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical='center')

                # Format specific columns based on content
                for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                    header_value = col[0].value
                    col_letter = get_column_letter(col_idx)

                    # TTBR and Rate columns - show 2 decimal places (CHECK FIRST - these contain "USD" but aren't USD values!)
                    if header_value and any(keyword in str(header_value) for keyword in
                                             ['TTBR', 'Rate']):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'  # Show 2 decimal places

                    # USD columns - show 2 decimal places with $ symbol (CHECK AFTER TTBR to avoid conflicts)
                    elif header_value and '(USD)' in str(header_value):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '$#,##0.00'  # Show 2 decimal places

                    # INR Currency columns - different format for A2/A3 vs other sheets
                    elif header_value and any(keyword in str(header_value) for keyword in
                                           ['INR', 'Value', 'Balance', 'Amount', 'Proceeds', 'Cost', 'Gain', 'Tax', 'Invstmnt']):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and isinstance(cell.value, (int, float)):
                                # A2 and A3 sheets: integers only (ITR portal requirement)
                                # Other sheets: show decimals for accuracy
                                if sheet_name in ["Table A2 Custodial Acc", "Table A3 Equity Interest", "Excluded from A3"]:
                                    cell.number_format = '₹#,##0'  # No decimals for A2/A3
                                else:
                                    cell.number_format = '₹#,##0.00'  # Show 2 decimal places for reference sheets

                    # Date columns
                    elif header_value and 'Date' in str(header_value):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value and not isinstance(cell.value, str):
                                cell.number_format = 'YYYY-MM-DD'

                    # Text columns (ZipCode, AccountNumber)
                    elif header_value in ['ZipCode', 'AccountNumber']:
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value is not None:
                                cell.value = str(cell.value)
                                cell.number_format = '@'

                # Special formatting for Peak Info summary tables in Reference and A2 Peak sheets
                # This MUST run AFTER column formatting to override the "Value" column default
                if sheet_name in ["Reference - Daily Rates", "A2 Peak Calculation"]:
                    # Find where Peak Info columns start (look for "Peak Info" or "PEAK SUMMARY" header)
                    peak_info_col = None
                    value_col = None
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value in ['Peak Info', 'PEAK SUMMARY']:
                            peak_info_col = col_idx
                        elif cell.value == 'Value' and peak_info_col:
                            value_col = col_idx
                            break

                    if peak_info_col and value_col:
                        # Override formatting row by row based on label
                        for row_idx in range(2, ws.max_row + 1):
                            label_cell = ws.cell(row=row_idx, column=peak_info_col)
                            value_cell = ws.cell(row=row_idx, column=value_col)

                            if label_cell.value and value_cell.value and isinstance(value_cell.value, (int, float)):
                                label = str(label_cell.value)

                                # Apply specific formatting based on label content
                                if '(USD)' in label:
                                    # Stock Price (USD) or Account Value (USD)
                                    value_cell.number_format = '$#,##0.00'
                                elif label == 'TTBR':
                                    # Plain TTBR rate
                                    value_cell.number_format = '#,##0.00'
                                elif 'INR' in label or 'Peak Balance' in label:
                                    # Account Value (INR), Peak INR Value, A2 Peak Balance (INR)
                                    value_cell.number_format = '₹#,##0.00'

                # Special formatting for Capital Gains sheet
                if sheet_name == "Capital Gains":
                    # Table 2 header and data (bold + colored)
                    table2_header_row = start_row_table2 + 1
                    table2_data_row = start_row_table2 + 2

                    # Format Table 2 header
                    for cell in ws[table2_header_row]:
                        if cell.value:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Format Table 2 data (total row)
                    for cell in ws[table2_data_row]:
                        if cell.value:
                            cell.fill = total_fill
                            cell.font = total_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '₹#,##0.00'  # Show 2 decimal places

                # Freeze first row
                ws.freeze_panes = ws['A2']

            # Auto-adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    header_cell = column[0]

                    # Calculate width based on header and data
                    for cell in column:
                        try:
                            if cell.value is not None:
                                # For formatted numbers, account for the formatting
                                if isinstance(cell.value, (int, float)):
                                    # Currency with symbol (₹1,234.56 or $1,234.56)
                                    if cell.number_format and ('₹' in cell.number_format or '$' in cell.number_format):
                                        # Account for currency symbol, commas, and decimals
                                        formatted_length = len(f"{cell.value:,.2f}") + 2  # +2 for symbol and space
                                    # Plain numbers with decimals
                                    elif cell.number_format and '#,##0' in cell.number_format:
                                        formatted_length = len(f"{cell.value:,.2f}")
                                    else:
                                        formatted_length = len(str(cell.value))
                                    cell_length = formatted_length
                                else:
                                    cell_length = len(str(cell.value))

                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                    # Ensure minimum width for header text
                    if header_cell.value:
                        header_length = len(str(header_cell.value))
                        max_length = max(max_length, header_length)

                    # Set column width with padding, minimum 10, maximum 50
                    adjusted_width = min(max(max_length + 3, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        # Generate CSV files with custom headers matching ITR format
        csv_a2_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}_table_a2.csv")
        csv_a3_filename = os.path.join(output_dir, f"schedule_fa_{self.indian_fy}_table_a3.csv")

        # Table A2 CSV with ITR-compliant headers and column order
        # First, round all numeric values in the source data to integers
        for acc in output_data["ScheduleFA"]["DtlsForeignCustodialAcc"]:
            acc["PeakBalanceDuringPeriod"] = round(acc["PeakBalanceDuringPeriod"])
            acc["ClosingBalance"] = round(acc["ClosingBalance"])
            acc["GrossAmtPaidCredited"] = round(acc["GrossAmtPaidCredited"])

        df_a2_csv = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignCustodialAcc"])
        df_a2_csv = df_a2_csv.rename(columns={
            "CountryName": "Country/Region name",
            "CountryCodeExcludingIndia": "Country Name and Code",
            "FinancialInstName": "Name of financial institution",
            "FinancialInstAddress": "Address of financial institution",
            "ZipCode": "ZIP Code",
            "AccountNumber": "Account Number",
            "Status": "Status",
            "AccOpenDate": "Account opening date",
            "PeakBalanceDuringPeriod": "Peak Balance During the Period",
            "ClosingBalance": "Closing balance",
            "NatureOfAmount": "Nature of Amount",
            "GrossAmtPaidCredited": "Amount"
        })
        # Format dates to DD/MM/YYYY (ITR portal requirement)
        if "Account opening date" in df_a2_csv.columns:
            df_a2_csv["Account opening date"] = pd.to_datetime(df_a2_csv["Account opening date"], errors='coerce').dt.strftime('%d/%m/%Y')

        # Round numeric columns to integers (ITR portal requirement)
        df_a2_csv["Peak Balance During the Period"] = df_a2_csv["Peak Balance During the Period"].round(0).astype(int)
        df_a2_csv["Closing balance"] = df_a2_csv["Closing balance"].round(0).astype(int)
        df_a2_csv["Amount"] = df_a2_csv["Amount"].round(0).astype(int)

        # Reorder columns to match ITR format exactly
        df_a2_csv = df_a2_csv[[
            "Country/Region name", "Country Name and Code", "Name of financial institution",
            "Address of financial institution", "ZIP Code", "Account Number", "Status",
            "Account opening date", "Peak Balance During the Period", "Closing balance",
            "Nature of Amount", "Amount"
        ]]
        # Write A2 CSV (ITR portal format requirement)
        # Try simple format without quotes or trailing commas
        df_a2_csv.to_csv(csv_a2_filename, index=False, quoting=0)  # quoting=0 = QUOTE_MINIMAL

        # Table A3 CSV with ITR-compliant headers
        # First, round all numeric values in the source data to integers
        for holding in output_data["ScheduleFA"]["DtlsForeignEquityDebtInterest"]:
            holding["InitialValOfInvstmnt"] = round(holding["InitialValOfInvstmnt"])
            holding["PeakBalanceDuringPeriod"] = round(holding["PeakBalanceDuringPeriod"])
            holding["ClosingBalance"] = round(holding["ClosingBalance"])
            holding["TotGrossAmtPaidCredited"] = round(holding["TotGrossAmtPaidCredited"])
            holding["TotGrossProceeds"] = round(holding["TotGrossProceeds"])

        df_a3_csv = pd.DataFrame(output_data["ScheduleFA"]["DtlsForeignEquityDebtInterest"])
        df_a3_csv = df_a3_csv.rename(columns={
            "CountryName": "Country/Region name",
            "CountryCodeExcludingIndia": "Country Name and Code",
            "NameOfEntity": "Name of entity",
            "AddressOfEntity": "Address of entity",
            "ZipCode": "ZIP Code",
            "NatureOfEntity": "Nature of entity",
            "InterestAcquiringDate": "Date of acquiring the interest",
            "InitialValOfInvstmnt": "Initial value of the investment",
            "PeakBalanceDuringPeriod": "Peak value of investment during the Period",
            "ClosingBalance": "Closing balance",
            "TotGrossAmtPaidCredited": "Total gross amount paid/credited with respect to the holding during the period",
            "TotGrossProceeds": "Total gross proceeds from sale or redemption of investment during the period"
        })

        # Keep dates in YYYY-MM-DD ISO format (ITR portal requirement)
        # This avoids DD/MM vs MM/DD ambiguity
        if "Date of acquiring the interest" in df_a3_csv.columns:
            df_a3_csv["Date of acquiring the interest"] = pd.to_datetime(df_a3_csv["Date of acquiring the interest"], errors='coerce').dt.strftime('%Y-%m-%d')

        # Force text columns to be treated as text (prevent Excel "leading zeros" warning)
        # Add ="value" format for text fields
        text_cols_a3 = ["Country Name and Code", "ZIP Code"]
        for col in text_cols_a3:
            if col in df_a3_csv.columns:
                df_a3_csv[col] = df_a3_csv[col].astype(str)

        # Round numeric columns to integers (ITR portal requirement)
        df_a3_csv["Initial value of the investment"] = df_a3_csv["Initial value of the investment"].round(0).astype(int)
        df_a3_csv["Peak value of investment during the Period"] = df_a3_csv["Peak value of investment during the Period"].round(0).astype(int)
        df_a3_csv["Closing balance"] = df_a3_csv["Closing balance"].round(0).astype(int)
        df_a3_csv["Total gross amount paid/credited with respect to the holding during the period"] = df_a3_csv["Total gross amount paid/credited with respect to the holding during the period"].round(0).astype(int)
        df_a3_csv["Total gross proceeds from sale or redemption of investment during the period"] = df_a3_csv["Total gross proceeds from sale or redemption of investment during the period"].round(0).astype(int)

        # Write A3 CSV (ITR portal format requirement)
        # Try simple format without quotes or trailing commas
        df_a3_csv.to_csv(csv_a3_filename, index=False, quoting=0)  # quoting=0 = QUOTE_MINIMAL

        print(f"\n[SUCCESS] Finished processing calendar year {self.calendar_year}!")
        print(f"    - JSON Output:  {json_filename}")
        print(f"    - Excel Output: {excel_filename} (7 sheets)")
        print(f"        • Table A2 Custodial Acc")
        print(f"        • Table A3 Equity Interest")
        print(f"        • Excluded from A3 (Sales from previous years)")
        print(f"        • Capital Gains (Current + Future sales)")
        print(f"        • Reference - Daily Rates (AMD prices + SBI TTBR)")
        print(f"        • A2 Peak Calculation (Daily account values)")
        print(f"        • Pre-{self.calendar_year} Holdings Init Val")
        print(f"    - CSV A2:       {csv_a2_filename}")
        print(f"    - CSV A3:       {csv_a3_filename}")
        print(f"    - Total Equity Tranches: {len(equity_tranches)}")
        print(f"    - Open Holdings: {len([t for t in equity_tranches if t['ClosingBalance'] > 0])}")
        print(f"    - Sold Holdings: {len([t for t in equity_tranches if 'Sold' in t['NatureOfEntity']])}")
        if 'df_sold_before_fy' in locals():
            print(f"    - Excluded from A3: {len(df_sold_before_fy)} sales from previous years\n")
        else:
            print()
        return output_data

# =====================================================================
# EXECUTION ENTRY POINT
# =====================================================================
if __name__ == "__main__":

    # CONFIGURATION
    TARGET_YEAR = 2025
    # Load configuration from config.json
    config = load_config()

    TARGET_YEAR = config.get("target_year", 2025)
    # Get account number from custodial_account
    custodial_acc = config.get("custodial_account", {})
    ACCOUNT_NUMBER = custodial_acc.get("account_number", "")

    # Account number can be extracted from ClientStatement PDF, so it's optional in config
    if not ACCOUNT_NUMBER or ACCOUNT_NUMBER == "ENTER_YOUR_ETRADE_ACCOUNT_NUMBER":
        print("[i] Account number not in config.json - will extract from ClientStatement PDF")
        ACCOUNT_NUMBER = ""  # Will be extracted from ClientStatement

    # Input file paths - auto-detect from inputs folder
    BYSTATUS_FILE = "inputs/ByStatus_expanded.xlsx"
    GL_FILE = "inputs/G&L_Expanded.xlsx"

    # Fallback to root if inputs folder doesn't have them
    if not os.path.exists(BYSTATUS_FILE):
        BYSTATUS_FILE = "ByStatus_expanded.xlsx"
    if not os.path.exists(GL_FILE):
        GL_FILE = "G&L_Expanded.xlsx"

    print("\n[*] Starting Schedule FA generation (WEB SCRAPING MODE)...")
    print(f"[*] This will open Chrome browser in background")
    print(f"[*] Looking for E*TRADE files: {BYSTATUS_FILE}, {GL_FILE}")
    print()

    try:
        # Read E*TRADE files to discover company symbols
        print("[*] Reading E*TRADE export files to discover companies...")
        df_open = pd.read_excel(BYSTATUS_FILE)
        df_sold = pd.read_excel(GL_FILE)

        # Auto-discover unique symbols from input files
        symbols = set()
        if 'Symbol' in df_open.columns:
            symbols.update(df_open['Symbol'].dropna().unique())
        if 'Symbol' in df_sold.columns:
            symbols.update(df_sold['Symbol'].dropna().unique())

        print(f"[OK] Discovered {len(symbols)} unique symbols: {', '.join(sorted(str(s).strip() for s in symbols))}")

        # Initialize app to access scraping methods
        app = ScheduleFAApp(calendar_year=TARGET_YEAR)

        # Scrape company info for each discovered symbol not in config
        config_companies = config.get("table_a3_companies", {})
        # Filter out entries that start with underscore (comments/notes)
        config_companies = {k: v for k, v in config_companies.items() if not k.startswith('_')}

        for symbol in sorted(symbols):
            symbol = str(symbol).strip()
            if symbol and symbol not in config_companies:
                print(f"\n[+] Auto-discovering company info for: {symbol}")
                company_info = app._scrape_company_profile(symbol)
                config_companies[symbol] = company_info
                time.sleep(2)  # Be respectful to Yahoo Finance
            elif symbol:
                print(f"[>] Using existing config for: {symbol}")

        # Update config with discovered companies
        config["table_a3_companies"] = config_companies

        # Save updated config back to file for next run
        save_config(config)
        print("[OK] Saved company info to config.json for next run")
        print()

        # Now process E*TRADE exports with updated config
        app.process_etrade_exports(
            bystatus_path=BYSTATUS_FILE,
            gl_path=GL_FILE,
            account_no=ACCOUNT_NUMBER,
            config=config
        )

        print("[*] Process complete! Check the generated files:")
        print(f"    - schedule_fa_{app.indian_fy}.json")
        print(f"    - schedule_fa_{app.indian_fy}.xlsx")
        print("\n[*] You can now upload the JSON to the ITR e-filing portal.")
        print("[*] Review the Excel file to verify all values before filing.\n")

    except Exception as e:
        print(f"\n[ERROR] {str(e)}\n")
